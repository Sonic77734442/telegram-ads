from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Dict, List, Literal, Optional, Tuple
from enum import Enum

from fastapi import Depends, FastAPI, Header, HTTPException, Form, Request
import logging
import traceback
import time
import base64
from fastapi import File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, HTMLResponse, FileResponse, RedirectResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
import hashlib
import hmac
import secrets
import json
import os
import shutil
import httpx
import boto3
from botocore.config import Config as BotoConfig
from google.ads.googleads.client import GoogleAdsClient
from google.api_core import exceptions as google_api_exceptions
from dotenv import load_dotenv

from app.db import get_conn

load_dotenv()

_R2_CLIENT = None
_BCC_RATES_CACHE = {"ts": 0.0, "data": None}
_BCC_RATES_TTL_SEC = 900
_BCC_TOKEN_CACHE = {"token": None, "expires_at": 0.0}
_BCC_DEFAULT_MARKUP = float(os.getenv("BCC_DEFAULT_MARKUP", "10") or 10)
_LIVE_BILLING_CACHE: Dict[str, Dict[str, object]] = {}
_LIVE_BILLING_TTL_SEC = 300


def _default_fee_config() -> Dict[str, Optional[float]]:
    return {
        "meta": 5,
        "google": 5,
        "yandex": 3,
        "tiktok": 3,
        "telegram": 25,
        "monochrome": None,
    }


def _load_fee_config(raw: Optional[str]) -> Dict[str, Optional[float]]:
    base = _default_fee_config()
    if not raw:
        return base
    try:
        data = json.loads(raw)
    except Exception:
        return base
    if not isinstance(data, dict):
        return base
    for key in base.keys():
        if key in data:
            base[key] = data.get(key)
    return base


def _r2_enabled() -> bool:
    return bool(
        os.getenv("R2_ACCESS_KEY_ID")
        and os.getenv("R2_SECRET_ACCESS_KEY")
        and os.getenv("R2_BUCKET")
        and os.getenv("R2_ACCOUNT_ID")
    )


def _r2_client():
    global _R2_CLIENT
    if _R2_CLIENT:
        return _R2_CLIENT
    if not _r2_enabled():
        return None
    account_id = os.getenv("R2_ACCOUNT_ID")
    endpoint = os.getenv("R2_ENDPOINT") or f"https://{account_id}.r2.cloudflarestorage.com"
    _R2_CLIENT = boto3.client(
        "s3",
        endpoint_url=endpoint,
        aws_access_key_id=os.getenv("R2_ACCESS_KEY_ID"),
        aws_secret_access_key=os.getenv("R2_SECRET_ACCESS_KEY"),
        region_name="auto",
        config=BotoConfig(signature_version="s3v4"),
    )
    return _R2_CLIENT


def _r2_bucket() -> str:
    return os.getenv("R2_BUCKET", "")


def _r2_parse_path(path: str) -> Optional[Tuple[str, str]]:
    if not path or not path.startswith("r2://"):
        return None
    raw = path[5:]
    if "/" not in raw:
        return None
    bucket, key = raw.split("/", 1)
    return bucket, key


def _r2_presigned_url(key: str, bucket: Optional[str] = None, expires: int = 3600) -> Optional[str]:
    client = _r2_client()
    if not client:
        return None
    bucket_name = bucket or _r2_bucket()
    return client.generate_presigned_url(
        "get_object",
        Params={"Bucket": bucket_name, "Key": key},
        ExpiresIn=expires,
    )


def _r2_upload_bytes(key: str, data: bytes, content_type: str = "application/pdf") -> str:
    client = _r2_client()
    bucket_name = _r2_bucket()
    client.put_object(Bucket=bucket_name, Key=key, Body=data, ContentType=content_type)
    return f"r2://{bucket_name}/{key}"


def _fetch_bcc_rates() -> Dict[str, object]:
    now = time.time()
    cached = _BCC_RATES_CACHE.get("data")
    if cached and now - _BCC_RATES_CACHE.get("ts", 0) < _BCC_RATES_TTL_SEC:
        return cached
    rates_url = os.getenv("BCC_RATES_URL", "https://api.bcc.kz/bcc/production/v1/public/rates")
    token_url = os.getenv("BCC_TOKEN_URL", "https://api.bcc.kz/bcc/production/v2/oauth/token")
    client_id = os.getenv("BCC_CLIENT_ID")
    client_secret = os.getenv("BCC_CLIENT_SECRET")
    scope = os.getenv("BCC_SCOPE", "bcc.application.informational.api")

    auth_header = None
    if client_id and client_secret:
        cached_token = _BCC_TOKEN_CACHE.get("token")
        token_exp = float(_BCC_TOKEN_CACHE.get("expires_at", 0.0))
        if cached_token and token_exp - now > 30:
            auth_header = f"Bearer {cached_token}"
        else:
            basic_raw = f"{client_id}:{client_secret}".encode("utf-8")
            basic = base64.b64encode(basic_raw).decode("ascii")
            token_resp = httpx.post(
                token_url,
                headers={
                    "Authorization": f"Basic {basic}",
                    "Content-Type": "application/x-www-form-urlencoded",
                },
                data={"grant_type": "client_credentials", "scope": scope},
                timeout=15,
            )
            if token_resp.status_code != 200:
                raise HTTPException(status_code=502, detail="Failed to fetch BCC OAuth token")
            token_payload = token_resp.json()
            access_token = token_payload.get("access_token")
            if not access_token:
                raise HTTPException(status_code=502, detail="BCC OAuth token is empty")
            expires_in = int(token_payload.get("expires_in") or 3600)
            _BCC_TOKEN_CACHE["token"] = access_token
            _BCC_TOKEN_CACHE["expires_at"] = now + max(60, expires_in)
            auth_header = f"Bearer {access_token}"

    headers = {"Accept": "application/json"}
    if auth_header:
        headers["Authorization"] = auth_header
    resp = httpx.get(rates_url, headers=headers, timeout=15)
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail="Failed to fetch BCC rates from API")

    payload = resp.json()

    def _extract_rates_list(raw: object) -> Optional[List[Dict[str, object]]]:
        if isinstance(raw, list):
            return [item for item in raw if isinstance(item, dict)]
        if not isinstance(raw, dict):
            return None
        candidates = [
            raw.get("Rates"),
            raw.get("rates"),
            (raw.get("data") or {}).get("Rates") if isinstance(raw.get("data"), dict) else None,
            (raw.get("data") or {}).get("rates") if isinstance(raw.get("data"), dict) else None,
            (raw.get("result") or {}).get("Rates") if isinstance(raw.get("result"), dict) else None,
            (raw.get("result") or {}).get("rates") if isinstance(raw.get("result"), dict) else None,
        ]
        for candidate in candidates:
            if isinstance(candidate, list):
                return [item for item in candidate if isinstance(item, dict)]
        return None

    rates_list = _extract_rates_list(payload)
    if not isinstance(rates_list, list):
        if isinstance(payload, dict):
            logging.error("Unexpected BCC payload keys: %s", list(payload.keys()))
        raise HTTPException(status_code=502, detail="Unexpected BCC rates payload")

    rates: Dict[str, Optional[Dict[str, float]]] = {"USD": None, "EUR": None}
    updated_at = None
    for item in rates_list:
        if not isinstance(item, dict):
            continue
        code = str(item.get("currency") or item.get("currencyCode") or item.get("code") or "").upper()
        if code not in rates:
            continue
        purchase = item.get("purchase")
        if purchase is None:
            purchase = item.get("buy")
        sell = item.get("sell")
        if sell is None:
            sell = item.get("sale")
        try:
            buy_value = float(purchase) if purchase is not None else None
            sell_value = float(sell) if sell is not None else None
        except (TypeError, ValueError):
            continue
        rates[code] = {"sell": sell_value, "buy": buy_value}
        dt = item.get("dateTime") or item.get("datetime") or item.get("date")
        if isinstance(dt, str) and dt:
            if updated_at is None or dt > updated_at:
                updated_at = dt

    data = {
        "source": "bcc_api",
        "section": "public",
        "url": rates_url,
        "rates": rates,
        "fetched_at": datetime.utcnow().isoformat() + "Z",
        "dateTime": updated_at,
    }
    _BCC_RATES_CACHE["ts"] = now
    _BCC_RATES_CACHE["data"] = data
    return data


def _get_marked_bcc_sell_rate(code: str, rates_data: Optional[Dict[str, object]] = None) -> Optional[float]:
    code_upper = str(code or "").upper()
    if not code_upper:
        return None
    try:
        rates_payload = rates_data or _fetch_bcc_rates()
    except Exception:
        return None
    rates = rates_payload.get("rates") if isinstance(rates_payload, dict) else None
    if not isinstance(rates, dict):
        return None
    entry = rates.get(code_upper)
    if not isinstance(entry, dict):
        return None
    sell = entry.get("sell")
    try:
        sell_value = float(sell) if sell is not None else None
    except (TypeError, ValueError):
        return None
    if sell_value is None:
        return None
    return sell_value + _BCC_DEFAULT_MARKUP


def _convert_amount_to_usd(amount: object, currency: object, rates_data: Optional[Dict[str, object]] = None) -> Optional[float]:
    try:
        value = float(amount) if amount is not None else None
    except (TypeError, ValueError):
        return None
    if value is None:
        return None

    code = str(currency or "USD").upper()
    if code == "USD":
        return value

    usd_rate = _get_marked_bcc_sell_rate("USD", rates_data)
    if not usd_rate or usd_rate <= 0:
        return None

    if code == "KZT":
        return value / usd_rate

    if code == "EUR":
        eur_rate = _get_marked_bcc_sell_rate("EUR", rates_data)
        if not eur_rate or eur_rate <= 0:
            return None
        return (value * eur_rate) / usd_rate

    return None


def _convert_amount_to_kzt(amount: object, currency: object, rates_data: Optional[Dict[str, object]] = None) -> Optional[float]:
    try:
        value = float(amount) if amount is not None else None
    except (TypeError, ValueError):
        return None
    if value is None:
        return None

    code = str(currency or "KZT").upper()
    if code == "KZT":
        return value

    if code == "USD":
        usd_rate = _get_marked_bcc_sell_rate("USD", rates_data)
        if not usd_rate or usd_rate <= 0:
            return None
        return value * usd_rate

    if code == "EUR":
        eur_rate = _get_marked_bcc_sell_rate("EUR", rates_data)
        if not eur_rate or eur_rate <= 0:
            return None
        return value * eur_rate

    return None


def _r2_upload_fileobj(key: str, fileobj, content_type: str = "application/pdf") -> str:
    client = _r2_client()
    bucket_name = _r2_bucket()
    try:
        fileobj.seek(0)
    except Exception:
        pass
    client.upload_fileobj(fileobj, bucket_name, key, ExtraArgs={"ContentType": content_type})
    return f"r2://{bucket_name}/{key}"


def _send_telegram_alert(text: str, channel: str = "ops") -> None:
    if channel == "reg":
        token = os.getenv("TELEGRAM_REG_BOT_TOKEN")
        chat_id = os.getenv("TELEGRAM_REG_CHAT_ID")
        thread_id = os.getenv("TELEGRAM_REG_MESSAGE_THREAD_ID")
    else:
        token = os.getenv("TELEGRAM_BOT_TOKEN")
        chat_id = os.getenv("TELEGRAM_CHAT_ID")
        thread_id = os.getenv("TELEGRAM_MESSAGE_THREAD_ID")
    if not token or not chat_id:
        return
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload: Dict[str, object] = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "HTML",
        "disable_web_page_preview": True,
    }
    if thread_id:
        try:
            payload["message_thread_id"] = int(thread_id)
        except ValueError:
            pass
    try:
        httpx.post(url, json=payload, timeout=8)
    except Exception:
        # Telegram notifications should not break business flow.
        pass
Goal = Literal["reach", "traffic", "leads", "conversions"]
PlatformKey = Literal[
    "meta",
    "google_display_cpm",
    "google_display_cpc",
    "google_search",
    "google_shopping",
    "youtube",
    "youtube_6s",
    "youtube_15s",
    "youtube_30s",
    "tiktok",
    "telegrad_channels",
    "telegrad_users",
    "telegrad_bots",
    "telegrad_search",
    "yandex_search",
    "yandex_display",
]
TargetingDepth = Literal["broad", "balanced", "focused"]
Country = Literal["kz", "uz", "ru"]
PricingMode = Literal["auto", "cpm", "cpc"]
Market = Literal["kz", "uz", "ru"]
Industry = Literal[
    "fmcg",
    "pharma",
    "finance",
    "travel",
    "ecommerce",
    "auto",
    "real_estate",
    "education",
    "other",
]
Currency = Literal["USD", "KZT"]
PlanMode = Literal["smart", "strategy"]
BusinessType = Literal["services", "ecom", "b2b", "local", "content"]


class FactRow(BaseModel):
    date: date
    platform: PlatformKey
    ad_account_id: Optional[str] = None
    campaign_name: Optional[str] = None
    impressions: float = 0
    clicks: float = 0
    cost: float = 0
    leads: float = 0
    conversions: float = 0
    views: float = 0


class RateCard(BaseModel):
    key: PlatformKey
    name: str
    tagline: str
    cpm: float
    cpc: float
    cpv: float
    ctr: float
    cvr: float
    post_click: float
    strengths: List[str]
    pricing: Literal["cpm", "cpc", "mixed"] = "cpm"
    format: Optional[str] = None


class PlanLine(BaseModel):
    key: PlatformKey
    name: str
    role: Optional[str] = None
    rationale: Optional[str] = None
    share: float
    budget: float
    impressions: float
    reach: float
    clicks: float
    leads: float
    conversions: float
    cpm: float
    cpc: float
    cpv: float
    cvr: float


class PlanTotals(BaseModel):
    budget: float
    impressions: float
    reach: float
    clicks: float
    leads: float
    conversions: float


class PlanResponse(BaseModel):
    lines: List[PlanLine]
    totals: PlanTotals
    kpi: Optional[Dict[str, float]] = None
    budget_usd: float
    period_days: int
    planned_kpi: Optional[float] = None
    fact_weekly: Optional[List[Dict[str, object]]] = None
    fact_raw: Optional[List[FactRow]] = None
    unmatched_fact: Optional[List[FactRow]] = None


class PlanRequest(BaseModel):
    company: Optional[str] = Field(None, description="Client/company name (legacy)")
    client_name: Optional[str] = Field(None, description="Client legal name")
    brand: Optional[str] = Field(None, description="Brand name")
    product: Optional[str] = Field(None, description="Product name")
    market: Optional[Market] = Field(None, description="Primary market")
    geo_split: Optional[str] = Field(None, description="Geo split free form")
    plan_mode: Optional[PlanMode] = Field("strategy", description="Plan mode: smart or strategy")
    business_type: Optional[BusinessType] = Field(None, description="Business type for smart mode")
    budget: float = Field(..., gt=0, description="Total budget in entered currency")
    currency: Currency = Field("USD", description="Currency of input budget")
    fx_rate: Optional[float] = Field(None, gt=0, description="FX rate KZT->USD when currency is KZT")
    date_start: Optional[date] = Field(None, description="Start date")
    date_end: Optional[date] = Field(None, description="End date")
    goal: Goal = Field("leads")
    avg_frequency: float = Field(1.6, gt=0)
    period_days: int = Field(30, gt=0)
    targeting_depth: TargetingDepth = Field("balanced")
    seasonality: float = Field(1.0, gt=0)
    platforms: Optional[List[PlatformKey]] = Field(None, description="Platforms to include")
    placements: Optional[List[str]] = Field(None, description="Placement/slot toggles (e.g. fb_feed, ig_reels)")
    pricing_mode: PricingMode = Field("auto", description="auto/cpm/cpc for forecasting")
    country: Country = Field("kz", description="Country for rate adjustment")
    cities: Optional[List[str]] = Field(None, description="Cities in country")
    interests: Optional[List[str]] = Field(None, description="Audience interests")
    age_min: Optional[int] = Field(None, ge=13, description="Minimum age")
    age_max: Optional[int] = Field(None, ge=13, description="Maximum age")
    industry: Industry = Field("other", description="Industry for cost/CTR tuning")
    ltv_per_conversion: float = Field(100.0, gt=0, description="LTV per conversion in USD")
    kpi_type: Optional[str] = Field(None, description="KPI type (cpc/cpl/cpa/cpm)")
    kpi_target: Optional[float] = Field(None, description="KPI target value in USD")
    budget_split: Optional[Dict[PlatformKey, float]] = Field(None, description="Manual budget share per platform")
    audience_size: Optional[float] = Field(None, description="Known audience size if available")
    creative_count: Optional[int] = Field(None, description="Creative count")
    author: Optional[str] = Field(None, description="Who prepared the media plan")
    vat_percent: Optional[float] = Field(None, description="VAT percent, if applicable")
    agency_fee_percent: Optional[float] = Field(None, description="Agency commission percent")
    utm_template: Optional[str] = Field(None, description="UTM template")
    pixels_configured: Optional[bool] = Field(None, description="Are pixels/conversions configured?")
    budget_scenarios: Optional[Dict[str, float]] = Field(
        None,
        description="Named scenarios with budget multipliers, e.g. {'base':1, 'plus20':1.2, 'minus20':0.8}",
    )
    funnel_split: Optional[Dict[str, float]] = Field(
        None, description="Funnel split percentages, e.g. {'awareness':40, 'consideration':35, 'performance':25}"
    )
    channel_inputs: Optional[Dict[str, Dict[str, float]]] = Field(
        None,
        description="Channel-specific overrides, e.g. {'meta': {'cpm':2.5,'ctr':0.012,'cvr':0.02}, 'google_search':{'cpc':0.6,'cvr':0.03}}",
    )
    assumption_profile: Optional[str] = Field(None, description="Assumption profile key")
    assumptions: Optional[Dict[str, str]] = Field(None, description="Assumptions breakdown")
    match_strategy: Optional[Literal["account", "campaign", "platform"]] = Field(
        "account", description="How to match fact to plan: account->ad_account_id, campaign->name, platform"
    )
    telegrad_rich_media: Optional[bool] = Field(True, description="If True, +50% CPM uplift for Telegrad channels/users")
    monthly_platforms: Optional[List[List[PlatformKey]]] = Field(
        None, description="Per-month allowed platforms; index 0 = month1. Missing -> all platforms on."
    )


rate_cards: Dict[PlatformKey, RateCard] = {
    "meta": RateCard(
        key="meta",
        name="Meta (FB/IG)",
        tagline="Lead forms, reach, CPC/CPM",
        cpm=2.1,
        cpc=0.35,
        cpv=0.02,
        ctr=0.013,
        cvr=0.016,
        post_click=0.42,
        strengths=["Social", "Lead forms", "Retarget"],
        pricing="mixed",
        format="Feed/Reels",
    ),
    "google_display_cpm": RateCard(
        key="google_display_cpm",
        name="Google Display В· CPM",
        tagline="Баннерная сеть (показы)",
        cpm=2.8,
        cpc=0.45,
        cpv=0.03,
        ctr=0.008,
        cvr=0.012,
        post_click=0.32,
        strengths=["Awareness", "GDN", "Brand"],
        pricing="cpm",
        format="Display",
    ),
    "google_display_cpc": RateCard(
        key="google_display_cpc",
        name="Google Display В· CPC",
        tagline="Баннерная сеть (клики)",
        cpm=3.1,
        cpc=0.28,
        cpv=0.03,
        ctr=0.015,
        cvr=0.014,
        post_click=0.34,
        strengths=["Traffic", "Banner CPC", "Smart Display"],
        pricing="cpc",
        format="Display",
    ),
    "google_search": RateCard(
        key="google_search",
        name="Google Search · Контекст",
        tagline="Поиск с намерением",
        cpm=4.2,
        cpc=0.55,
        cpv=0.04,
        ctr=0.025,
        cvr=0.028,
        post_click=0.5,
        strengths=["High intent", "Text ads", "Performance"],
        pricing="cpc",
        format="Search",
    ),
    "google_shopping": RateCard(
        key="google_shopping",
        name="Google Shopping",
        tagline="Товарные кампании",
        cpm=3.6,
        cpc=0.42,
        cpv=0.04,
        ctr=0.022,
        cvr=0.032,
        post_click=0.44,
        strengths=["Feed", "Ecom", "PLA"],
        pricing="cpc",
        format="Shopping",
    ),
    "youtube": RateCard(
        key="youtube",
        name="YouTube (Generic)",
        tagline="Видео и бренд",
        cpm=2.4,
        cpc=0.48,
        cpv=0.015,
        ctr=0.009,
        cvr=0.011,
        post_click=0.35,
        strengths=["Video", "Awareness", "Masthead"],
        pricing="cpm",
        format="Video",
    ),
    "youtube_6s": RateCard(
        key="youtube_6s",
        name="YouTube 6s Bumper",
        tagline="Короткое видео (6s)",
        cpm=2.0,
        cpc=0.52,
        cpv=0.011,
        ctr=0.007,
        cvr=0.008,
        post_click=0.28,
        strengths=["Bumper", "Reach", "Awareness"],
        pricing="cpm",
        format="Video",
    ),
    "youtube_15s": RateCard(
        key="youtube_15s",
        name="YouTube 15s",
        tagline="Средняя длительность (15s)",
        cpm=2.3,
        cpc=0.5,
        cpv=0.013,
        ctr=0.0085,
        cvr=0.01,
        post_click=0.32,
        strengths=["Mid video", "Reach/Traffic", "Skippable"],
        pricing="cpm",
        format="Video",
    ),
    "youtube_30s": RateCard(
        key="youtube_30s",
        name="YouTube 30s",
        tagline="Длинное видео (30s)",
        cpm=2.6,
        cpc=0.46,
        cpv=0.017,
        ctr=0.0095,
        cvr=0.012,
        post_click=0.36,
        strengths=["Story", "Brand", "Skippable"],
        pricing="cpm",
        format="Video",
    ),
    "tiktok": RateCard(
        key="tiktok",
        name="TikTok",
        tagline="UGC и вовлечение",
        cpm=1.9,
        cpc=0.3,
        cpv=0.01,
        ctr=0.017,
        cvr=0.013,
        post_click=0.33,
        strengths=["UGC", "Video", "Broad audience"],
        pricing="mixed",
        format="Video",
    ),
    "telegrad_channels": RateCard(
        key="telegrad_channels",
        name="Telegrad В· Channels",
        tagline="Реклама в каналах",
        cpm=0.12,
        cpc=0.32,
        cpv=0.0,
        ctr=0.014,
        cvr=0.011,
        post_click=0.31,
        strengths=["Channels", "Messenger", "Feed"],
        pricing="mixed",
        format="Feed",
    ),
    "telegrad_users": RateCard(
        key="telegrad_users",
        name="Telegrad В· Users",
        tagline="Реклама на пользователя",
        cpm=0.12,
        cpc=0.32,
        cpv=0.0,
        ctr=0.014,
        cvr=0.011,
        post_click=0.31,
        strengths=["User ads", "Messenger"],
        pricing="mixed",
        format="Dialogs",
    ),
    "telegrad_bots": RateCard(
        key="telegrad_bots",
        name="Telegrad В· Bots",
        tagline="Боты / CPA-like",
        cpm=0.10,
        cpc=0.3,
        cpv=0.0,
        ctr=0.012,
        cvr=0.01,
        post_click=0.28,
        strengths=["Bots", "Automation"],
        pricing="mixed",
        format="Bots",
    ),
    "telegrad_search": RateCard(
        key="telegrad_search",
        name="Telegrad В· Search",
        tagline="Поиск в мессенджере",
        cpm=0.08,
        cpc=0.28,
        cpv=0.0,
        ctr=0.011,
        cvr=0.009,
        post_click=0.25,
        strengths=["Search", "Intent"],
        pricing="cpc",
        format="Search",
    ),
    "yandex_search": RateCard(
        key="yandex_search",
        name="Яндекс Поиск",
        tagline="Контекст, РСЯ/Поиск",
        cpm=3.9,
        cpc=0.5,
        cpv=0.0,
        ctr=0.022,
        cvr=0.026,
        post_click=0.48,
        strengths=["High intent", "Search", "CPC"],
        pricing="cpc",
        format="Search",
    ),
    "yandex_display": RateCard(
        key="yandex_display",
        name="Яндекс Директ · РСЯ",
        tagline="Баннеры/смарт-баннеры",
        cpm=2.6,
        cpc=0.36,
        cpv=0.0,
        ctr=0.012,
        cvr=0.013,
        post_click=0.3,
        strengths=["Display", "Smart-banner", "Awareness"],
        pricing="mixed",
        format="Display",
    ),
}


targeting_adjustments: Dict[TargetingDepth, Dict[str, float]] = {
    "broad": {"cost": 0.94, "ctr": 0.92, "cvr": 0.9},
    "balanced": {"cost": 1.0, "ctr": 1.0, "cvr": 1.0},
    "focused": {"cost": 1.12, "ctr": 1.06, "cvr": 1.12},
}

country_adjustments: Dict[Country, float] = {
    "kz": 1.0,
    "uz": 0.9,
    "ru": 1.15,
}

min_budget_thresholds: Dict[PlatformKey, float] = {
    "meta": 300,
    "google_display_cpm": 400,
    "google_display_cpc": 400,
    "google_search": 500,
    "google_shopping": 400,
    "youtube": 300,
    "youtube_6s": 250,
    "youtube_15s": 300,
    "youtube_30s": 350,
    "tiktok": 250,
    "telegrad_channels": 250,
    "telegrad_users": 250,
    "telegrad_bots": 200,
    "telegrad_search": 200,
    "yandex_search": 400,
    "yandex_display": 350,
}

industry_adjustments: Dict[Industry, Dict[str, float]] = {
    "fmcg": {"cost": 0.97, "ctr": 1.02, "cvr": 0.98},
    "pharma": {"cost": 1.12, "ctr": 0.95, "cvr": 0.92},
    "finance": {"cost": 1.25, "ctr": 0.93, "cvr": 1.05},
    "travel": {"cost": 1.05, "ctr": 1.0, "cvr": 1.02},
    "ecommerce": {"cost": 1.0, "ctr": 1.0, "cvr": 1.02},
    "auto": {"cost": 1.15, "ctr": 0.94, "cvr": 1.03},
    "real_estate": {"cost": 1.18, "ctr": 0.92, "cvr": 1.04},
    "education": {"cost": 1.08, "ctr": 0.98, "cvr": 1.05},
    "other": {"cost": 1.0, "ctr": 1.0, "cvr": 1.0},
}


empty_totals = PlanTotals(
    budget=0,
    impressions=0,
    reach=0,
    clicks=0,
    leads=0,
    conversions=0,
)


def adjust_rate(
    card: RateCard, depth: TargetingDepth, seasonality: float, country: Country, rich_media_telegrad: bool = True
) -> RateCard:
    adjust = targeting_adjustments.get(depth, targeting_adjustments["balanced"])
    country_factor = country_adjustments.get(country, 1.0)
    cpm = card.cpm
    # Telegrad channels/users: +50% CPM if rich media (image/video)
    if rich_media_telegrad and card.key in {"telegrad_channels", "telegrad_users"}:
        cpm = cpm * 1.5
    return card.model_copy(
        update={
            "cpm": cpm * adjust["cost"] * seasonality * country_factor,
            "cpc": card.cpc * adjust["cost"] * seasonality * country_factor,
            "cpv": card.cpv * adjust["cost"] * seasonality * country_factor,
            "ctr": card.ctr * adjust["ctr"],
            "cvr": min(card.cvr * adjust["cvr"], 0.35),
        }
    )


def apply_channel_overrides(card: RateCard, channel_inputs: Optional[Dict[str, Dict[str, float]]]) -> RateCard:
    if not channel_inputs:
        return card
    key = None
    if card.key == "meta":
        key = "meta"
    elif card.key == "google_search":
        key = "google_search"
    elif card.key in {"telegrad_channels", "telegrad_users", "telegrad_bots", "telegrad_search"}:
        key = "telegram"
    if not key or key not in channel_inputs:
        return card
    overrides = channel_inputs.get(key) or {}
    update: Dict[str, float] = {}
    for field in ["cpm", "cpc", "ctr", "cvr"]:
        val = overrides.get(field)
        if val is not None and val > 0:
            update[field] = val
    return card.model_copy(update=update) if update else card


def smart_media_mix(goal: Goal, business_type: Optional[BusinessType]) -> Tuple[List[PlatformKey], Dict[PlatformKey, float], Dict[PlatformKey, str]]:
    bt = business_type or "services"
    if goal == "conversions":
        if bt == "ecom":
            platforms = ["meta", "google_search", "google_display_cpc"]
            split = {"meta": 0.5, "google_search": 0.3, "google_display_cpc": 0.2}
            rationale = {
                "meta": "Трафик + догрев",
                "google_search": "Перехват горячего спроса",
                "google_display_cpc": "Ремаркетинг и добор",
            }
            return platforms, split, rationale
        platforms = ["meta", "google_search"]
        split = {"meta": 0.6, "google_search": 0.4}
        rationale = {"meta": "Генерация спроса", "google_search": "Перехват спроса"}
        return platforms, split, rationale
    if goal == "traffic":
        platforms = ["meta", "telegrad_channels"]
        split = {"meta": 0.6, "telegrad_channels": 0.4}
        rationale = {"meta": "Дешевый охват и клики", "telegrad_channels": "Доп. трафик и клики"}
        return platforms, split, rationale
    # leads / default
    platforms = ["meta", "google_search"]
    split = {"meta": 0.6, "google_search": 0.4}
    rationale = {"meta": "Генерация спроса", "google_search": "Перехват спроса"}
    if bt in {"local", "services"}:
        rationale["meta"] = "Генерация заявок"
        rationale["google_search"] = "Горячий спрос"
    if bt == "b2b":
        rationale["meta"] = "Узкая аудитория + вовлечение"
        rationale["google_search"] = "Спрос по запросам"
    return platforms, split, rationale


def parse_fact_csv(csv_text: str) -> List[FactRow]:
    import csv
    from io import StringIO

    rows: List[FactRow] = []
    reader = csv.DictReader(StringIO(csv_text))
    for raw in reader:
        try:
            rows.append(
                FactRow(
                    date=raw.get("date"),
                    platform=raw.get("platform"),
                    ad_account_id=raw.get("ad_account_id") or None,
                    campaign_name=raw.get("campaign_name") or None,
                    impressions=float(raw.get("impressions") or 0),
                    clicks=float(raw.get("clicks") or 0),
                    cost=float(raw.get("cost") or 0),
                    leads=float(raw.get("leads") or 0),
                    conversions=float(raw.get("conversions") or 0),
                    views=float(raw.get("views") or 0),
                )
            )
        except Exception:
            continue
    return rows


def aggregate_weekly(plan: PlanResponse, facts: List[FactRow], strategy: str = "account") -> Tuple[List[Dict[str, object]], List[FactRow]]:
    """Aggregate fact weekly and compute plan per week by platform using daily averages."""
    if plan.period_days <= 0:
        return [], []
    # daily plan by platform
    daily_plan: Dict[str, Dict[str, float]] = {}
    for line in plan.lines:
        daily_plan[line.key] = {
            "budget": line.budget / plan.period_days,
            "impressions": line.impressions / plan.period_days,
            "reach": line.reach / plan.period_days,
            "clicks": line.clicks / plan.period_days,
            "leads": line.leads / plan.period_days,
            "conversions": line.conversions / plan.period_days,
        }

    weekly: Dict[Tuple[int, int, str], Dict[str, float]] = {}
    days_in_week: Dict[Tuple[int, int, str], set] = {}
    unmatched: List[FactRow] = []

    for row in facts:
        iso_year, iso_week, _ = row.date.isocalendar()
        key_component = row.platform
        if strategy == "account" and row.ad_account_id:
            key_component = row.ad_account_id
        elif strategy == "campaign" and row.campaign_name:
            key_component = row.campaign_name.lower().strip()
        else:
            key_component = row.platform

        key = (iso_year, iso_week, key_component)
        if row.platform not in daily_plan:
            unmatched.append(row)
            continue
        bucket = weekly.setdefault(
            key,
            {"cost": 0, "impressions": 0, "clicks": 0, "leads": 0, "conversions": 0, "views": 0},
        )
        bucket["cost"] += row.cost
        bucket["impressions"] += row.impressions
        bucket["clicks"] += row.clicks
        bucket["leads"] += row.leads
        bucket["conversions"] += row.conversions
        bucket["views"] += row.views
        days = days_in_week.setdefault(key, set())
        days.add(row.date)

    result: List[Dict[str, object]] = []
    for key, fact in weekly.items():
        iso_year, iso_week, platform_key = key
        d_plan = daily_plan.get(platform_key if platform_key in daily_plan else str(platform_key), {})
        days_count = len(days_in_week.get(key, set()))
        plan_week = {
            "budget": d_plan.get("budget", 0) * days_count,
            "impressions": d_plan.get("impressions", 0) * days_count,
            "reach": d_plan.get("reach", 0) * days_count,
            "clicks": d_plan.get("clicks", 0) * days_count,
            "leads": d_plan.get("leads", 0) * days_count,
            "conversions": d_plan.get("conversions", 0) * days_count,
        }
        result.append(
            {
                "year": iso_year,
                "week": iso_week,
                "platform": platform_key,
                "plan": plan_week,
                "fact": fact,
            }
        )
    return sorted(result, key=lambda x: (x["year"], x["week"], str(x["platform"]))), unmatched


def estimate_audience_size(req: PlanRequest, platform: PlatformKey) -> Optional[float]:
    """Fallback оценка объёма ЦА (reach cap) в условиях отсутствия API."""
    if req.country == "kz":
        base_population = 10_000_000
    elif req.country == "uz":
        base_population = 18_000_000
    else:
        base_population = 35_000_000
    age_span = 52  # 65-13 базовый диапазон
    age_min = req.age_min or 18
    age_max = req.age_max or 55
    age_factor = max(0.1, min(1.0, (age_max - age_min) / age_span))

    city_factor = 1.0
    if req.cities:
        city_factor = max(0.25, min(1.0, 0.18 * len(req.cities)))

    interest_factor = 0.75 if req.interests else 1.0

    platform_factor: Dict[PlatformKey, float] = {
        "meta": 0.65,
        "google_display_cpm": 0.7,
        "google_display_cpc": 0.7,
        "google_search": 0.55,
        "google_shopping": 0.35,
        "youtube": 0.7,
        "tiktok": 0.6,
    }
    pf = platform_factor.get(platform, 0.6)

    return base_population * age_factor * city_factor * interest_factor * pf


def score_for_goal(goal: Goal, card: RateCard) -> float:
    if goal == "reach":
        return 1 / max(card.cpm, 1)
    if goal == "traffic":
        return 1 / max(card.cpc, 1)
    if goal == "leads":
        return card.cvr / max(card.cpc, 1)
    if goal == "conversions":
        return (card.cvr * card.post_click) / max(card.cpc, 1)
    return 1.0


def compute_metrics(
    goal: Goal,
    budget: float,
    card: RateCard,
    avg_frequency: float,
    pricing_mode: PricingMode,
    audience_cap: Optional[float],
) -> Dict[str, float]:
    pricing = pricing_mode if pricing_mode != "auto" else card.pricing

    impressions_from_budget = (budget / max(card.cpm, 1e-6)) * 1000
    impressions_from_clicks = (budget / max(card.cpc, 1e-6)) / max(card.ctr, 0.001)

    if pricing == "cpm":
        impressions = impressions_from_budget
    elif pricing == "cpc":
        impressions = impressions_from_clicks
    else:
        impressions = impressions_from_budget if goal == "reach" else max(impressions_from_budget, impressions_from_clicks)

    clicks_from_budget = budget / max(card.cpc, 1e-6)
    clicks_from_ctr = impressions * card.ctr
    if pricing == "cpc":
        clicks = clicks_from_budget
    elif pricing == "cpm" and goal == "traffic":
        clicks = clicks_from_ctr
    else:
        clicks = max(clicks_from_budget, clicks_from_ctr)

    leads = clicks * card.cvr
    conversions = leads * card.post_click
    reach = impressions / max(avg_frequency, 1.05)

    if audience_cap and audience_cap > 0:
        max_impressions = audience_cap * max(avg_frequency, 1.0) * 1.3
        impressions = min(impressions, max_impressions)
        reach = min(reach, audience_cap)
        clicks = min(clicks, impressions * card.ctr if card.ctr > 0 else clicks)
        leads = clicks * card.cvr
        conversions = leads * card.post_click

    return {
        "budget": budget,
        "impressions": impressions,
        "reach": reach,
        "clicks": clicks,
        "leads": leads,
        "conversions": conversions,
    }


def build_plan(req: PlanRequest) -> PlanResponse:
    if req.currency == "KZT":
        if not req.fx_rate:
            raise HTTPException(status_code=400, detail="fx_rate is required when currency=KZT")
        budget_usd = req.budget / req.fx_rate
    else:
        budget_usd = req.budget

    effective_period = req.period_days
    if req.date_start and req.date_end and req.date_end >= req.date_start:
        effective_period = (req.date_end - req.date_start).days or req.period_days

    meta_placement_labels = {
        "fb_feed": "Meta В· Feed",
        "fb_video_feeds": "Meta В· Video Feeds",
        "fb_instream": "Meta В· In-Stream",
        "fb_reels": "Meta В· Reels",
        "fb_stories": "Meta В· Stories",
        "fb_search": "Meta В· Search",
        "ig_feed": "Meta В· IG Feed",
        "ig_profile_feed": "Meta В· Profile",
        "ig_reels": "Meta В· IG Reels",
        "ig_explore": "Meta В· Explore",
        "ig_explore_home": "Meta В· Explore Home",
        "ig_stories": "Meta В· IG Stories",
    }
    meta_specific = [p for p in (req.placements or []) if p in meta_placement_labels]

    plan_mode = req.plan_mode or "strategy"
    smart_split: Optional[Dict[PlatformKey, float]] = None
    smart_rationale: Dict[PlatformKey, str] = {}
    if plan_mode == "smart":
        platforms, split, rationale = smart_media_mix(req.goal, req.business_type)
        active_keys = platforms
        smart_split = split
        smart_rationale = rationale
    else:
        active_keys = req.platforms or list(rate_cards.keys())
    active_cards: List[RateCard] = []
    meta_expansion_count = 0
    for key in active_keys:
        if key == "meta" and meta_specific:
            base = rate_cards.get("meta")
            if not base:
                continue
            for mp in meta_specific:
                active_cards.append(base.model_copy(update={"name": meta_placement_labels.get(mp, base.name)}))
                meta_expansion_count += 1
            continue
        card = rate_cards.get(key)
        if card:
            active_cards.append(card)
    if not active_cards:
        return PlanResponse(lines=[], totals=empty_totals)

    adjusted = [
        adjust_rate(card, req.targeting_depth, req.seasonality, req.country, rich_media_telegrad=req.telegrad_rich_media)
        for card in active_cards
    ]
    industry_adj = industry_adjustments.get(req.industry, industry_adjustments["other"])
    adjusted = [
        card.model_copy(
            update={
                "cpm": card.cpm * industry_adj["cost"],
                "cpc": card.cpc * industry_adj["cost"],
                "cpv": card.cpv * industry_adj["cost"],
                "ctr": card.ctr * industry_adj["ctr"],
                "cvr": min(card.cvr * industry_adj["cvr"], 0.35),
            }
        )
        for card in adjusted
    ]
    adjusted = [apply_channel_overrides(card, req.channel_inputs) for card in adjusted]
    scored = [score_for_goal(req.goal, card) for card in adjusted]
    total_score = sum(scored) or len(adjusted)

    manual_split = None
    if req.budget_split:
        manual_total = sum(req.budget_split.values())
        if manual_total > 0:
            manual_split = {k: v / manual_total for k, v in req.budget_split.items()}
    if smart_split:
        manual_split = smart_split

    lines: List[PlanLine] = []
    for idx, card in enumerate(adjusted):
        share = (
            (manual_split.get(card.key, 0) if card.key != "meta" else (manual_split.get("meta", 0) / max(meta_expansion_count, 1)))
            if manual_split is not None
            else scored[idx] / total_score if total_score else 1 / len(adjusted)
        )
        if share == 0 and manual_split is not None:
            continue
        line_budget = budget_usd * share
        audience_cap = estimate_audience_size(req, card.key)
        metrics = compute_metrics(
            req.goal,
            line_budget,
            card,
            req.avg_frequency,
            req.pricing_mode,
            audience_cap,
        )
        lines.append(
            PlanLine(
                key=card.key,
                name=card.name,
                role=None,
                rationale=smart_rationale.get(card.key) if plan_mode == "smart" else None,
                share=share,
                budget=metrics["budget"],
                impressions=metrics["impressions"],
                reach=metrics["reach"],
                clicks=metrics["clicks"],
                leads=metrics["leads"],
                conversions=metrics["conversions"],
                cpm=card.cpm,
                cpc=card.cpc,
                cpv=card.cpv,
                cvr=card.cvr,
            )
        )

    totals = PlanTotals(
        budget=sum(line.budget for line in lines),
        impressions=sum(line.impressions for line in lines),
        reach=sum(line.reach for line in lines),
        clicks=sum(line.clicks for line in lines),
        leads=sum(line.leads for line in lines),
        conversions=sum(line.conversions for line in lines),
    )

    planned_kpi = None
    if req.kpi_type == "cpl" and totals.leads > 0:
        planned_kpi = totals.budget / totals.leads
    if req.kpi_type == "cpa" and totals.conversions > 0:
        planned_kpi = totals.budget / totals.conversions
    if req.kpi_type == "cpc" and totals.clicks > 0:
        planned_kpi = totals.budget / totals.clicks
    if req.kpi_type == "cpm" and totals.impressions > 0:
        planned_kpi = (totals.budget / totals.impressions) * 1000

    return PlanResponse(
        lines=lines,
        totals=totals,
        budget_usd=budget_usd,
        period_days=effective_period,
        planned_kpi=planned_kpi,
        fact_weekly=None,
    )


def plan_to_workbook(
    plan: PlanResponse,
    req: Optional[PlanRequest] = None,
    fact_rows: Optional[List[FactRow]] = None,
    weekly_fact: Optional[List[Dict[str, object]]] = None,
) -> BytesIO:
    wb = Workbook()
    outputs = wb.active
    outputs.title = "Outputs"
    inputs = wb.create_sheet("Inputs", 0)
    calculations = wb.create_sheet("Calculations")
    # Inputs (core)
    audience_desc = ""
    audience_volume = ""
    period_days = ""
    fx = ""
    author = ""
    company = ""
    client_name = ""
    brand = ""
    product = ""
    market = ""
    geo_split = ""
    creative_count = ""
    audience_size = ""
    funnel_split = ""
    if req:
        company = req.company or ""
        client_name = req.client_name or ""
        brand = req.brand or ""
        product = req.product or ""
        market = (req.market or req.country).upper()
        geo_split = req.geo_split or ""
        period_days = str(req.period_days)
        if req.date_start and req.date_end and req.date_end >= req.date_start:
            period_days = str((req.date_end - req.date_start).days or req.period_days)
        age_part = f"{req.age_min or 18}-{req.age_max or 55}"
        interests_part = ", ".join(req.interests) if req.interests else "broad"
        geo_part = ", ".join(req.cities) if req.cities else (req.market or req.country).upper()
        audience_desc = f"{geo_part}; {age_part}; {interests_part}"
        if req.audience_size:
            audience_volume = f"{int(req.audience_size):,}".replace(",", " ")
        elif req.platforms:
            cap = estimate_audience_size(req, req.platforms[0])
            audience_volume = f"{int(cap):,}".replace(",", " ") if cap else ""
        author = req.author or ""
        if req.currency == "KZT" and req.fx_rate:
            fx = f"{req.fx_rate}"
        creative_count = f"{req.creative_count}" if req.creative_count is not None else ""
        audience_size = f"{req.audience_size}" if req.audience_size is not None else ""
        if req.funnel_split:
            aw = req.funnel_split.get("awareness")
            co = req.funnel_split.get("consideration")
            pf = req.funnel_split.get("performance")
            parts = []
            if aw is not None:
                parts.append(f"Awareness {aw}%")
            if co is not None:
                parts.append(f"Consideration {co}%")
            if pf is not None:
                parts.append(f"Performance {pf}%")
            funnel_split = " / ".join(parts)
    input_rows = [
        ["Client", client_name or company],
        ["Brand", brand],
        ["Product", product],
        ["Market", market],
        ["Geo split", geo_split],
        ["Objective", req.goal if req else ""],
        ["Budget (net)", req.budget if req else ""],
        ["Currency", req.currency if req else ""],
        ["FX rate (KZTв†’USD)", fx],
        ["Flight start", req.date_start.isoformat() if req and req.date_start else ""],
        ["Flight end", req.date_end.isoformat() if req and req.date_end else ""],
        ["Period (days)", period_days],
        ["Audience description", audience_desc],
        ["Audience size", audience_volume or audience_size],
        ["Creative count", creative_count],
        ["Funnel split", funnel_split],
        ["Agency fee, %", f"{req.agency_fee_percent}%" if req and req.agency_fee_percent else ""],
        ["VAT, %", f"{req.vat_percent}%" if req and req.vat_percent else ""],
        ["KPI type", req.kpi_type if req else ""],
        ["KPI target", req.kpi_target if req and req.kpi_target else ""],
        ["UTM template", req.utm_template if req and req.utm_template else ""],
        ["Pixels configured", "Да" if req and req.pixels_configured else "Нет"],
        ["Channel overrides", json.dumps(req.channel_inputs, ensure_ascii=False) if req and req.channel_inputs else ""],
        ["Prepared by", author],
    ]
    for idx, row in enumerate(input_rows, start=1):
        inputs.cell(row=idx, column=1, value=row[0])
        inputs.cell(row=idx, column=2, value=row[1])

    wb.remove(calculations)

    totals = plan.totals
    fee = (req.agency_fee_percent or 0) / 100 if req else 0
    vat = (req.vat_percent or 0) / 100 if req else 0

    headers = [
        "Платформа",
        "Доля, %",
        "Бюджет план, $",
        "Бюджет факт, $",
        "Охват план",
        "Охват факт",
        "Показы план",
        "Показы факт",
        "Клики план",
        "Клики факт",
        "Лиды план",
        "Лиды факт",
        "Конверсии план",
        "Конверсии факт",
        "Просмотры план",
        "Просмотры факт",
        "Viewable план",
        "Viewable факт",
        "CPM план, $",
        "CPM факт, $",
        "CPC план, $",
        "CPC факт, $",
        "CPV план, $",
        "CPV факт, $",
        "CTR план",
        "CTR факт",
        "CVR план",
        "CVR факт",
        "Post-click план",
        "Post-click факт",
        "VTR план",
        "VTR факт",
        "LTV план",
        "LTV факт",
        "Дней открутки",
        "Пополнение (gross), $",
    ]
    totals = plan.totals
    total_overhead = totals.budget * (fee + vat)
    total_gross = totals.budget + total_overhead
    cpm = (totals.budget / totals.impressions * 1000) if totals.impressions else None
    cpc = (totals.budget / totals.clicks) if totals.clicks else None
    cpl = (totals.budget / totals.leads) if totals.leads else None
    cpa = (totals.budget / totals.conversions) if totals.conversions else None
    freq = (totals.impressions / totals.reach) if totals.reach else None
    flight = ""
    if req and req.date_start and req.date_end:
        flight = f"{req.date_start.isoformat()} в†’ {req.date_end.isoformat()}"
    elif req:
        flight = f"{req.period_days} days"
    summary_rows = [
        ["Budget (net/client)", totals.budget, "Гарантия"],
        ["Комиссия/VAT", total_overhead, "Гарантия"],
        ["Budget (gross)", total_gross, "Гарантия"],
        ["CPM", cpm, "Прогноз"],
        ["CPC", cpc, "Прогноз"],
        ["CPL", cpl, "Прогноз"],
        ["CPA", cpa, "Прогноз"],
        ["Impressions", totals.impressions, "Прогноз"],
        ["Reach", totals.reach, "Прогноз"],
        ["Clicks", totals.clicks, "Прогноз"],
        ["Leads", totals.leads, "Прогноз"],
        ["Purchases", totals.conversions, "Прогноз"],
        ["Frequency", round(freq, 2) if freq else None, "Прогноз"],
        ["Flight", flight, "Прогноз"],
    ]
    outputs.append(["Outputs (standard)", "Value", "Type"])
    for row in summary_rows:
        outputs.append(row)
    outputs.append([])
    outputs.append(headers)
    start_row = outputs.max_row + 1
    fee = (req.agency_fee_percent or 0) / 100 if req else 0
    vat = (req.vat_percent or 0) / 100 if req else 0
    for idx, line in enumerate(plan.lines):
        row = start_row + idx
        ctr_val = (line.clicks / line.impressions) if line.impressions > 0 else 0
        cvr_val = (line.leads / line.clicks) if line.clicks > 0 else line.cvr
        post_click_val = (line.conversions / line.leads) if line.leads > 0 else 0.35
        is_video = 1 if (line.name.lower().find("youtube") >= 0 or line.name.lower().find("tiktok") >= 0) else 0
        outputs.append(
            [
                line.name,
                round(line.share * 100, 1),
                round(line.budget),
                None,  # Budget fact
                None,  # Reach (formula)
                None,  # Reach fact
                None,  # Impressions plan (formula)
                None,  # Impressions fact
                None,  # Clicks plan (formula)
                None,  # Clicks fact
                None,  # Leads plan (formula)
                None,  # Leads fact
                None,  # Conversions plan (formula)
                None,  # Conversions fact
                None,  # Views plan (formula)
                None,  # Views fact
                None,  # Viewable plan (formula)
                None,  # Viewable fact
                round(line.cpm, 3),  # CPM plan
                None,  # CPM fact
                round(line.cpc, 3),  # CPC plan
                None,  # CPC fact
                round(line.cpv, 4),  # CPV plan
                None,  # CPV fact
                round(ctr_val, 4),  # CTR plan
                None,  # CTR fact
                round(cvr_val, 4),  # CVR plan
                None,  # CVR fact
                round(post_click_val, 4),  # Post-click plan
                None,  # Post-click fact
                None,  # VTR plan
                None,  # VTR fact
                None,  # LTV plan
                None,  # LTV fact
                None,  # Days
                round(line.budget * (1 + fee + vat), 2),  # Top-up with fee/VAT
            ]
        )
        # Excel formulas so values обновляются цепочкой внутри файла
        # Column mapping for readability:
        # C budget plan, D budget fact, E reach plan, F reach fact, G impr plan, H impr fact,
        # I clicks plan, J clicks fact, K leads plan, L leads fact, M conv plan, N conv fact,
        # O views plan, P views fact, Q viewable plan, R viewable fact,
        # S CPM plan, T CPM fact, U CPC plan, V CPC fact, W CPV plan, X CPV fact,
        # Y CTR plan, Z CTR fact, AA CVR plan, AB CVR fact, AC Post-click plan, AD Post-click fact,
        # AE VTR plan, AF VTR fact, AG LTV plan, AH LTV fact, AI Days
        outputs[f"G{row}"] = (
            f"=IF(S{row}>0,C{row}/S{row}*1000,"
            f"IF(AND(U{row}>0,Y{row}>0),C{row}/U{row}/MAX(Y{row},0.0001),0))"
        )  # Impressions plan
        outputs[f"E{row}"] = f"=IF(G{row}>0,G{row}/1.6,0)"  # Reach plan (avg freq 1.6)
        outputs[f"I{row}"] = f"=IF(U{row}>0,C{row}/U{row},G{row}*Y{row})"  # Clicks plan
        outputs[f"K{row}"] = f"=I{row}*AA{row}"  # Leads plan
        outputs[f"M{row}"] = f"=K{row}*AC{row}"  # Conversions plan
        outputs[f"O{row}"] = f"=IF(W{row}>0,C{row}/W{row},0)"  # Views plan
        outputs[f"Q{row}"] = f"=O{row}*{0.7 if is_video else 0}"  # Viewable plan
        outputs[f"AE{row}"] = f"=IF(G{row}>0,O{row}/G{row},0)"  # VTR plan
        outputs[f"AG{row}"] = f"=M{row}*100"  # LTV plan with $100 per conversion

    if plan.lines:
        outputs.append([])
        totals_row = start_row + len(plan.lines) + 1
        last_data_row = start_row + len(plan.lines) - 1
        outputs.append(
            [
                "Итого",
                "",
                f"=SUM(C{start_row}:C{last_data_row})",
                f"=SUM(D{start_row}:D{last_data_row})",
                f"=SUM(E{start_row}:E{last_data_row})",
                f"=SUM(F{start_row}:F{last_data_row})",
                f"=SUM(G{start_row}:G{last_data_row})",
                f"=SUM(H{start_row}:H{last_data_row})",
                f"=SUM(I{start_row}:I{last_data_row})",
                f"=SUM(J{start_row}:J{last_data_row})",
                f"=SUM(K{start_row}:K{last_data_row})",
                f"=SUM(L{start_row}:L{last_data_row})",
                f"=SUM(M{start_row}:M{last_data_row})",
                f"=SUM(N{start_row}:N{last_data_row})",
                f"=SUM(O{start_row}:O{last_data_row})",
                f"=SUM(P{start_row}:P{last_data_row})",
                f"=SUM(Q{start_row}:Q{last_data_row})",
                f"=SUM(R{start_row}:R{last_data_row})",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                f"=SUM(AJ{start_row}:AJ{last_data_row})",
            ]
        )

    # KPI block
    current_row = start_row + len(plan.lines) + 3
    if req and req.kpi_type and req.kpi_target:
        outputs.cell(row=current_row, column=1, value="KPI контроль")
        kpi_label = req.kpi_type.upper()
        outputs.cell(row=current_row, column=2, value="Тип")
        outputs.cell(row=current_row, column=3, value=kpi_label)
        outputs.cell(row=current_row + 1, column=2, value="План")
        outputs.cell(row=current_row + 1, column=3, value=plan.planned_kpi)
        outputs.cell(row=current_row + 2, column=2, value="Цель")
        outputs.cell(row=current_row + 2, column=3, value=req.kpi_target)
        outputs.cell(row=current_row + 3, column=2, value="Отклонение")
        if plan.planned_kpi:
            outputs.cell(row=current_row + 3, column=3, value=plan.planned_kpi - req.kpi_target)
        current_row += 5

    # Flight plan sheet (monthly + weekly per platform)
    flight = wb.create_sheet("Flight Plan")
    total_days = plan.period_days or (req.period_days if req else 0)
    weeks = max(1, (total_days + 6) // 7)
    months = max(1, (total_days + 29) // 30)
    fee = (req.agency_fee_percent or 0) / 100 if req else 0
    vat = (req.vat_percent or 0) / 100 if req else 0

    def month_weights(line_key: PlatformKey) -> List[float]:
        weights = [1.0] * months
        if req and req.monthly_platforms:
            for i in range(months):
                if i < len(req.monthly_platforms):
                    enabled = req.monthly_platforms[i]
                    weights[i] = 1.0 if line_key in enabled else 0.0
        total = sum(weights)
        if total <= 0:
            return [0.0] * months
        return [w / total for w in weights]

    week_to_month = []
    weeks_in_month = [0] * months
    for w in range(weeks):
        start_day = w * 7
        m_idx = min(months - 1, start_day // 30)
        week_to_month.append(m_idx)
        weeks_in_month[m_idx] += 1

    flight.append(["Месячный сплит по платформам"])
    flight.append([])
    monthly_header = ["Платформа"] + [f"M{i+1}" for i in range(months)]
    flight.append(monthly_header)
    for line in plan.lines:
        weights = month_weights(line.key)
        row = [line.name] + [round(line.budget * w, 2) for w in weights]
        flight.append(row)

    flight.append([])
    flight.append(["Недельный сплит по платформам"])
    weekly_header = ["Платформа"] + [f"W{i+1}" for i in range(weeks)]
    flight.append(weekly_header)
    for line in plan.lines:
        weights = month_weights(line.key)
        month_budgets = [line.budget * w for w in weights]
        row_vals = []
        for w in range(weeks):
            m_idx = week_to_month[w]
            divisor = weeks_in_month[m_idx] if weeks_in_month[m_idx] else 1
            row_vals.append(round(month_budgets[m_idx] / divisor, 2))
        row = [line.name] + row_vals
        flight.append(row)

    total_gross = sum(l.budget * (1 + fee + vat) for l in plan.lines)
    flight.append([])
    flight.append(["Итого к оплате (с НДС/ком.)", round(total_gross, 2), f"{total_days} дней"])

    # Creatives sheet
    creatives = wb.create_sheet("Creatives")
    creatives.append(["Платформа", "Форматы / размеры", "Текст", "Файлы / примечания"])
    creatives.append(["Meta (FB/IG) Feed", "1080x1080 (1:1), 1080x1350 (4:5), 1200x628 (1.91:1)", "Заголовок 25–40 знаков, текст до 125", "PNG/JPG; текст на изображении <=20%"])
    creatives.append(["Meta (FB/IG) Reels/Stories", "1080x1920 (9:16)", "Короткий текст", "Видео 9:16 или 4:5, MP4/MOV, до 4 ГБ"])
    creatives.append(["Google Ads КМС", "1200x628, 1200x1200, 300x250, 728x90", "Заголовок до 30, описание до 90", "PNG/JPG; высокое разрешение"])
    creatives.append(["Google Ads YouTube", "16:9", "Короткий заголовок", "Видео MP4, 16:9"])
    creatives.append(["Яндекс Директ (РСЯ)", "16:9 от 450x257 до 1080x607; 1:1 от 450x450 до 1080x1080; 2:3", "Заголовок до 56, текст до 81", "PNG/JPG"])
    creatives.append(["TikTok Ads", "9:16 (720x1280 или 1080x1920), 1:1, 16:9", "Заголовок до 100", "Видео MP4/MOV; рекоменд. 9:16"])
    creatives.append(["Telegram Ads", "Только текст + ссылка", "Заголовок до 160, текст до 160", "Без баннеров"])

    # Brand Metrics sheet removed
    # Scenarios sheet removed

    # Fact raw sheet
    if fact_rows:
        fact_sheet = wb.create_sheet("Fact Raw")
        fact_sheet.append(
            ["date", "platform", "ad_account_id", "campaign_name", "impressions", "clicks", "cost", "leads", "conversions", "views"]
        )
        for r in fact_rows:
            fact_sheet.append(
                [
                    r.date.isoformat(),
                    r.platform,
                    r.ad_account_id or "",
                    r.campaign_name or "",
                    r.impressions,
                    r.clicks,
                    r.cost,
                    r.leads,
                    r.conversions,
                    r.views,
                ]
            )

    # Plan vs Fact weekly sheet
    if weekly_fact:
        pvf = wb.create_sheet("Plan vs Fact Weekly")
        pvf.append(
            [
                "Year",
                "Week",
                "Platform",
                "Plan Reach",
                "Fact Reach (optional)",
                "Plan Budget",
                "Fact Cost",
                "Plan Impr",
                "Fact Impr",
                "Plan Clicks",
                "Fact Clicks",
                "Plan Leads",
                "Fact Leads",
                "Plan Conv",
                "Fact Conv",
                "CPL Plan",
                "CPL Fact",
                "CPA Plan",
                "CPA Fact",
                "CPM Fact",
                "CPC Fact",
            ]
        )
        for item in weekly_fact:
            plan_week = item["plan"]
            fact_week = item["fact"]
            pvf.append(
                [
                    item["year"],
                    item["week"],
                    item["platform"],
                    round(plan_week.get("reach", 0)),
                    "",
                    round(plan_week.get("budget", 0), 2),
                    round(fact_week.get("cost", 0), 2),
                    round(plan_week.get("impressions", 0)),
                    round(fact_week.get("impressions", 0)),
                    round(plan_week.get("clicks", 0)),
                    round(fact_week.get("clicks", 0)),
                    round(plan_week.get("leads", 0)),
                    round(fact_week.get("leads", 0)),
                    round(plan_week.get("conversions", 0)),
                    round(fact_week.get("conversions", 0)),
                    round(plan_week.get("budget", 0) / plan_week.get("leads", 1), 2) if plan_week.get("leads") else "",
                    round(fact_week.get("cost", 0) / fact_week.get("leads", 1), 2) if fact_week.get("leads") else "",
                    round(plan_week.get("budget", 0) / plan_week.get("conversions", 1), 2)
                    if plan_week.get("conversions")
                    else "",
                    round(fact_week.get("cost", 0) / fact_week.get("conversions", 1), 2)
                    if fact_week.get("conversions")
                    else "",
                    round((fact_week.get("cost", 0) / fact_week.get("impressions", 1)) * 1000, 2)
                    if fact_week.get("impressions")
                    else "",
                    round(fact_week.get("cost", 0) / fact_week.get("clicks", 1), 3) if fact_week.get("clicks") else "",
                ]
            )
    _format_workbook(wb)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _format_workbook(wb: Workbook) -> None:
    thin = Side(style="thin", color="999999")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for ws in wb.worksheets:
        max_widths: Dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                max_widths[cell.column] = max(max_widths.get(cell.column, 0), len(val))
                cell.border = border
        for col_idx, width in max_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max(width + 2, 10), 60)


app = FastAPI(title="Envidicy Media Plan API", version="0.2.0")
_default_origins = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://127.0.0.1:8000",
    "http://localhost:8000",
    "https://envidicydashclientv20.vercel.app",
    "https://app.envidicy.kz",
    "https://www.envidicy.kz",
]
_extra_origins = [o.strip() for o in (os.getenv("FRONTEND_ORIGINS") or "").split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=[*_default_origins, *_extra_origins],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Log unhandled errors to Render runtime logs for debugging.
@app.middleware("http")
async def log_exceptions(request: Request, call_next):
    try:
        return await call_next(request)
    except Exception:
        logging.error("Unhandled error on %s %s", request.method, request.url.path)
        logging.error(traceback.format_exc())
        raise
try:
    from app.db import apply_schema

    apply_schema()
except Exception:
    pass

ADMIN_EMAILS = {"romant997@gmail.com", "kolyadov.denis@gmail.com"}
BENEFICIARY = {
    "name": "ИП Art Book Inc.",
    "bin": "960910300234",
    "iban": "KZ588562204108888284",
    "bank": "АО Банк ЦентрКредит",
    "bic": "KCJBKZKX",
    "kbe": "19",
    "currency": "KZT",
}


def _format_date_ru(dt: datetime) -> str:
    return (
        dt.strftime("%d %B %Y")
        .replace("January", "января")
        .replace("February", "февраля")
        .replace("March", "марта")
        .replace("April", "апреля")
        .replace("May", "мая")
        .replace("June", "июня")
        .replace("July", "июля")
        .replace("August", "августа")
        .replace("September", "сентября")
        .replace("October", "октября")
        .replace("November", "ноября")
        .replace("December", "декабря")
    )


def _repair_mojibake_text(value: object) -> object:
    if not isinstance(value, str):
        return value
    text = value
    if not text:
        return text
    # Typical case: UTF-8 text was decoded as cp1251 and later saved as unicode.
    try:
        repaired = text.encode("cp1251").decode("utf-8")
        if repaired and repaired != text:
            return repaired
    except Exception:
        pass
    return text


_CYR_TO_LAT_MAP = str.maketrans(
    {
        "А": "A",
        "В": "B",
        "С": "C",
        "Е": "E",
        "Н": "H",
        "К": "K",
        "М": "M",
        "О": "O",
        "Р": "P",
        "Т": "T",
        "Х": "X",
        "У": "Y",
        "а": "A",
        "в": "B",
        "с": "C",
        "е": "E",
        "н": "H",
        "к": "K",
        "м": "M",
        "о": "O",
        "р": "P",
        "т": "T",
        "х": "X",
        "у": "Y",
    }
)


def _normalize_ascii_code(value: object) -> str:
    text = str(_repair_mojibake_text(value) or "").strip().translate(_CYR_TO_LAT_MAP).upper()
    return "".join(ch for ch in text if ("A" <= ch <= "Z") or ch.isdigit())


def _get_company_profile(conn) -> Dict[str, object]:
    row = conn.execute("SELECT * FROM company_profile WHERE id=1").fetchone()
    base = {
        "name": BENEFICIARY["name"],
        "bin": BENEFICIARY["bin"],
        "iin": None,
        "legal_address": None,
        "factual_address": None,
        "bank": BENEFICIARY["bank"],
        "iban": BENEFICIARY["iban"],
        "bic": BENEFICIARY["bic"],
        "kbe": BENEFICIARY["kbe"],
        "currency": BENEFICIARY["currency"],
    }
    if row:
        for key in base:
            if row[key] is not None:
                base[key] = row[key]
    base["name"] = _repair_mojibake_text(base.get("name"))
    base["bank"] = _repair_mojibake_text(base.get("bank"))
    base["legal_address"] = _repair_mojibake_text(base.get("legal_address"))
    base["factual_address"] = _repair_mojibake_text(base.get("factual_address"))
    iban_normalized = _normalize_ascii_code(base.get("iban"))
    bic_normalized = _normalize_ascii_code(base.get("bic"))
    if iban_normalized:
        base["iban"] = iban_normalized
    if bic_normalized:
        base["bic"] = bic_normalized
    return base


def _next_invoice_number(conn) -> str:
    now = datetime.utcnow()
    year = now.year
    row = conn.execute("SELECT seq FROM invoice_counters WHERE year=?", (year,)).fetchone()
    if row:
        seq = int(row["seq"]) + 1
        conn.execute("UPDATE invoice_counters SET seq=? WHERE year=?", (seq, year))
    else:
        seq = 1
        conn.execute("INSERT INTO invoice_counters (year, seq) VALUES (?, ?)", (year, seq))
    return f"{year % 100:02d}{seq:09d}"


def _format_legal_entity_name(entity: Dict[str, object]) -> Optional[str]:
    full = (entity.get("full_name") or "").strip()
    short = (entity.get("short_name") or entity.get("name") or "").strip()
    if full and short and full != short:
        return f"{full} ({short})"
    return full or short or None


def _invoice_number(prefix: str, created_at, inv_id: int) -> str:
    if isinstance(created_at, datetime):
        dt = created_at
    else:
        try:
            dt = datetime.fromisoformat(created_at)
        except ValueError:
            dt = datetime.utcnow()
    return f"{prefix}-{dt.strftime('%Y%m%d')}-{inv_id:04d}"


def _format_amount(amount: float) -> str:
    return f"{amount:,.2f}".replace(",", " ")


def _amount_to_words_ru(amount: float) -> str:
    def _triad_to_words(n: int, feminine: bool = False) -> str:
        units_m = ["", "один", "два", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"]
        units_f = ["", "одна", "две", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"]
        teens = [
            "десять",
            "одиннадцать",
            "двенадцать",
            "тринадцать",
            "четырнадцать",
            "пятнадцать",
            "шестнадцать",
            "семнадцать",
            "восемнадцать",
            "девятнадцать",
        ]
        tens = ["", "", "двадцать", "тридцать", "сорок", "пятьдесят", "шестьдесят", "семьдесят", "восемьдесят", "девяносто"]
        hundreds = ["", "сто", "двести", "триста", "четыреста", "пятьсот", "шестьсот", "семьсот", "восемьсот", "девятьсот"]
        words = []
        words.append(hundreds[n // 100])
        n = n % 100
        if 10 <= n <= 19:
            words.append(teens[n - 10])
        else:
            words.append(tens[n // 10])
            unit = n % 10
            words.append((units_f if feminine else units_m)[unit])
        return " ".join(w for w in words if w)

    def _group_word(n: int, forms: Tuple[str, str, str]) -> str:
        if n % 10 == 1 and n % 100 != 11:
            return forms[0]
        if 2 <= n % 10 <= 4 and not (12 <= n % 100 <= 14):
            return forms[1]
        return forms[2]

    if amount is None:
        return ""
    rub = int(amount)
    kop = int(round((amount - rub) * 100))
    if kop == 100:
        rub += 1
        kop = 0
    parts = []
    millions = rub // 1_000_000
    thousands = (rub // 1_000) % 1_000
    remainder = rub % 1_000
    if millions:
        parts.append(_triad_to_words(millions))
        parts.append(_group_word(millions, ("миллион", "миллиона", "миллионов")))
    if thousands:
        parts.append(_triad_to_words(thousands, feminine=True))
        parts.append(_group_word(thousands, ("тысяча", "тысячи", "тысяч")))
    if remainder or not parts:
        parts.append(_triad_to_words(remainder))
    words = " ".join(p for p in parts if p).strip()
    if words:
        words = words[0].upper() + words[1:]
    return f"{words} тенге {kop:02d} тиын"


def _invoice_1c_html(payload: Dict[str, object]) -> str:
    amount = payload.get("amount", "0.00")
    currency = payload.get("currency", "KZT")
    number = payload.get("number", "—")
    date = payload.get("date", "—")
    beneficiary_name = payload.get("beneficiary_name", "")
    beneficiary_bin = payload.get("beneficiary_bin", "")
    beneficiary_bank = payload.get("beneficiary_bank", "")
    beneficiary_iban = payload.get("beneficiary_iban", "")
    beneficiary_bic = payload.get("beneficiary_bic", "")
    beneficiary_kbe = payload.get("beneficiary_kbe", "")
    beneficiary_address = payload.get("beneficiary_address", "")
    beneficiary_phone = payload.get("beneficiary_phone", "")
    payment_code = payload.get("payment_code", "853")
    payer_name = payload.get("payer_name", "")
    payer_bin = payload.get("payer_bin", "")
    payer_address = payload.get("payer_address", "")
    description = payload.get("description", "")
    contract_note = payload.get("contract_note", "")
    amount_words = payload.get("amount_words", "")
    request_id = payload.get("request_id", "")
    token = payload.get("token", "")
    token_query = f"?token={token}" if token else ""
    pdf_url = f"/wallet/topup-requests/{request_id}/pdf-generated{token_query}"
    return f"""
<!doctype html>
<html lang="ru">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Счет на оплату</title>
    <style>
      @page {{
        size: A4;
        margin: 12mm;
      }}
      body {{
        font-family: "Arial", sans-serif;
        color: #111;
        margin: 0;
        background: #fff;
      }}
      .wrap {{
        width: 190mm;
        max-width: 100%;
        margin: 0 auto;
        padding: 8px 16px 24px;
      }}
      h1 {{
        font-size: 16px;
        margin: 12px 0 6px;
        font-weight: 700;
      }}
      .title-line {{
        border-bottom: 1px solid #111;
        margin: 6px 0 8px;
      }}
      .note {{
        font-size: 11px;
        margin: 6px 0 8px;
      }}
      .alert-line {{
        color: #d9251d;
        font-weight: 700;
        font-size: 11px;
        margin: 6px 0 4px;
      }}
      .print-btn {{
        display: inline-block;
        margin: 4px 0 10px;
        padding: 8px 12px;
        border: 1px solid #333;
        border-radius: 8px;
        text-decoration: none;
        color: #111;
        font-size: 11px;
      }}
      @media print {{
        .print-btn {{
          display: none;
        }}
      }}
      table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 11px;
      }}
      td, th {{
        border: 1px solid #333;
        padding: 5px 6px;
        vertical-align: top;
      }}
      .bank-table td {{
        padding: 4px 6px;
      }}
      .no-border td {{
        border: none;
        padding: 2px 0;
      }}
      .small {{
        font-size: 11px;
      }}
      .warning {{
        border: 2px solid #d9251d;
        padding: 6px 8px;
        font-size: 11px;
        margin: 6px 0 6px;
      }}
      .right {{
        text-align: right;
      }}
      .nowrap {{
        white-space: nowrap;
      }}
      .center {{
        text-align: center;
      }}
    </style>
  </head>
  <body>
    <div class="wrap">
      <a class="print-btn" href="{pdf_url}">Скачать PDF</a>
      <table class="bank-table">
        <tr>
          <td>
            <strong>Образец платежного поручения</strong><br />
            Бенефициар: {beneficiary_name}<br />
            БИН/ИИН: {beneficiary_bin}
          </td>
          <td>
            ИИК<br />
            <strong>{beneficiary_iban}</strong>
          </td>
          <td class="center">
            КБе<br />
            <strong>{beneficiary_kbe}</strong>
          </td>
        </tr>
        <tr>
          <td>Банк бенефициара:<br />{beneficiary_bank}</td>
          <td>
            БИК<br />
            <strong>{beneficiary_bic}</strong>
          </td>
          <td class="center">
            Код назначения платежа<br />
            <strong>{payment_code}</strong>
          </td>
        </tr>
      </table>

      <p class="note">Счет действителен в течение 5 рабочих дней</p>

      <h1>Счет на оплату № {number} от {date}</h1>
      <div class="title-line"></div>

      <table class="no-border">
        <tr>
          <td>Исполнитель</td>
          <td><strong>
            БИН/ИИН {beneficiary_bin}, {beneficiary_name}
            {f", {beneficiary_address}" if beneficiary_address else ""}{f", тел.: {beneficiary_phone}" if beneficiary_phone else ""}
          </strong></td>
        </tr>
        <tr>
          <td>Заказчик</td>
          <td><strong>БИН/ИИН {payer_bin}, {payer_name}, {payer_address}</strong></td>
        </tr>
        <tr>
          <td>Договор</td>
          <td><strong>{contract_note}</strong></td>
        </tr>
      </table>

      <div class="alert-line">Внимание! В назначение платежа скопируйте данные, указанные ниже.</div>
      <div class="warning">{description}</div>
      <div class="alert-line">
        Если назначение платежа будет указано некорректно, платеж может быть возвращен как ошибочный либо время поступления денег на счет может занять до 3-х рабочих дней
      </div>

      <table>
        <thead>
          <tr>
            <th class="center">№</th>
            <th>Наименование</th>
            <th class="center">Ед.</th>
            <th class="center">Кол-во</th>
            <th class="right nowrap">Цена</th>
            <th class="right nowrap">Сумма</th>
          </tr>
        </thead>
        <tbody>
          <tr>
            <td class="center">1</td>
            <td>{description}</td>
            <td class="center">услуга</td>
            <td class="center">1</td>
            <td class="right nowrap">{amount}</td>
            <td class="right nowrap">{amount}</td>
          </tr>
        </tbody>
      </table>

      <table class="no-border">
        <tr>
          <td class="right"><strong>Итого:</strong></td>
          <td class="right nowrap" style="width:160px;"><strong>{amount}</strong></td>
        </tr>
      </table>

      <p class="small">Всего наименований 1, на сумму {amount} {currency}</p>
      <p class="small"><strong>Всего к оплате:</strong> {amount_words}. Услуги Исполнителя НДС не облагаются (п.п. 46 ст.394 Налогового кодекса Казахстана).</p>
    </div>
  </body>
</html>
"""


def _invoice_html(payload: Dict[str, object]) -> str:
    items_html = "".join(
        f"""
          <tr>
            <td class="center">{i + 1}</td>
            <td>{item["description"]}</td>
            <td class="center">{item["qty"]}</td>
            <td class="center">{item["unit"]}</td>
            <td class="right">{item["price"]}</td>
            <td class="right">{item["amount"]}</td>
          </tr>
        """
        for i, item in enumerate(payload["items"])
    )
    return f"""
<!doctype html>
<html lang="ru">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Счет на оплату {payload["number"]}</title>
    <style>
      body {{
        font-family: "Times New Roman", serif;
        background: #f7f7f7;
        color: #111;
        margin: 0;
        padding: 24px;
      }}
      .sheet {{
        max-width: 820px;
        margin: 0 auto;
        background: #fff;
        border: 1px solid #000;
        padding: 20px;
      }}
      h1 {{
        font-size: 18px;
        margin: 12px 0 8px;
        text-align: center;
      }}
      .muted {{
        color: #111;
        font-size: 12px;
      }}
      .bank-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 12px;
      }}
      .bank-table td {{
        border: 1px solid #000;
        padding: 4px 6px;
        vertical-align: top;
      }}
      table {{
        width: 100%;
        border-collapse: collapse;
        margin-top: 12px;
      }}
      th, td {{
        border: 1px solid #000;
        padding: 6px;
        font-size: 13px;
      }}
      th {{
        text-align: left;
        background: #f5f5f5;
      }}
      .right {{
        text-align: right;
      }}
      .center {{
        text-align: center;
      }}
      .section {{
        margin-top: 10px;
        font-size: 13px;
      }}
    </style>
  </head>
  <body>
      <div class="sheet">
        <table class="bank-table">
          <tr>
            <td rowspan="2">
              Банк получателя<br />
              {payload["beneficiary_bank"]}
            </td>
            <td>БИК</td>
            <td>{payload["beneficiary_bic"]}</td>
          </tr>
          <tr>
            <td>ИИК</td>
            <td>{payload["beneficiary_iban"]}</td>
          </tr>
          <tr>
            <td>
              Бенефициар<br />
              {payload["beneficiary_name"]}
            </td>
            <td>БИН</td>
            <td>{payload["beneficiary_bin"]}</td>
          </tr>
          <tr>
            <td>КБе</td>
            <td colspan="2">{payload["beneficiary_kbe"]}</td>
          </tr>
        </table>

        <h1>Счет на оплату № {payload["number"]} от {payload["date"]}</h1>

        <div class="section">
          <strong>Поставщик:</strong> {payload["beneficiary_name"]}, ИИН/БИН {payload["beneficiary_bin"]}
        </div>
        <div class="section">
          <strong>Покупатель:</strong> {payload["payer_name"]}, {payload["payer_bin"]}, {payload["payer_address"]}
        </div>

        <table>
          <thead>
            <tr>
              <th>№</th>
              <th>Наименование</th>
              <th>Кол-во</th>
              <th>Ед.</th>
              <th>Цена</th>
              <th>Сумма</th>
            </tr>
          </thead>
          <tbody>
            {items_html}
            <tr>
              <td colspan="5" class="right"><strong>Итого</strong></td>
              <td class="right"><strong>{payload["amount"]} {payload["currency"]}</strong></td>
            </tr>
            <tr>
              <td colspan="5" class="right">НДС</td>
              <td class="right">Без НДС</td>
            </tr>
            <tr>
              <td colspan="5" class="right"><strong>Всего к оплате</strong></td>
              <td class="right"><strong>{payload["amount"]} {payload["currency"]}</strong></td>
            </tr>
          </tbody>
        </table>

        <div class="section">
          Всего наименований {payload["items_count"]}, на сумму {payload["amount"]} {payload["currency"]}.
        </div>
      </div>
    </body>
  </html>
"""


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.get("/rates/bcc")
def bcc_rates() -> Dict[str, object]:
    try:
        return _fetch_bcc_rates()
    except Exception as exc:
        cached = _BCC_RATES_CACHE.get("data")
        if isinstance(cached, dict):
            return {
                **cached,
                "source": "bcc_cache_fallback",
                "warning": str(exc),
                "fetched_at": datetime.utcnow().isoformat() + "Z",
            }
        return {
            "source": "bcc_unavailable",
            "rates": {"USD": None, "EUR": None},
            "fetched_at": datetime.utcnow().isoformat() + "Z",
            "warning": str(exc),
        }


@app.get("/rate-cards", response_model=List[RateCard])
def list_rate_cards() -> List[RateCard]:
    return list(rate_cards.values())


@app.post("/plans/estimate", response_model=PlanResponse)
def estimate_plan(payload: PlanRequest) -> PlanResponse:
    if payload.budget <= 0:
        raise HTTPException(status_code=400, detail="Budget must be positive")
    if payload.avg_frequency <= 0:
        raise HTTPException(status_code=400, detail="avg_frequency must be positive")
    return build_plan(payload)


@app.post("/plans/estimate/excel")
def estimate_plan_excel(payload: PlanRequest):
    plan = estimate_plan(payload)
    workbook = plan_to_workbook(plan, payload)
    headers = {
        "Content-Disposition": 'attachment; filename="mediaplan.xlsx"'
    }
    return StreamingResponse(
        workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/fact/weekly")
async def fact_weekly(
    plan_payload: PlanRequest,
    file: UploadFile = File(...),
):
    content = await file.read()
    csv_text = content.decode("utf-8")
    fact_rows = parse_fact_csv(csv_text)
    plan = estimate_plan(plan_payload)
    weekly, unmatched = aggregate_weekly(plan, fact_rows, plan_payload.match_strategy or "account")
    return {
        "weekly": weekly,
        "plan_budget": plan.budget_usd,
        "period_days": plan.period_days,
        "unmatched": [r.model_dump() for r in unmatched],
    }


@app.post("/fact/weekly/excel")
async def fact_weekly_excel(
    plan_payload: PlanRequest,
    file: UploadFile = File(...),
):
    content = await file.read()
    csv_text = content.decode("utf-8")
    fact_rows = parse_fact_csv(csv_text)
    plan = estimate_plan(plan_payload)
    weekly, unmatched = aggregate_weekly(plan, fact_rows, plan_payload.match_strategy or "account")
    plan.fact_weekly = weekly
    plan.fact_raw = fact_rows
    plan.unmatched_fact = unmatched
    workbook = plan_to_workbook(plan, plan_payload, fact_rows=fact_rows, weekly_fact=weekly)
    headers = {
        "Content-Disposition": 'attachment; filename="mediaplan_plan_vs_fact.xlsx"'
    }
    return StreamingResponse(
        workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/campaigns")
def create_campaign(name: str, currency: str = "USD"):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        cur = conn.execute("INSERT INTO campaigns (name, currency) VALUES (?, ?)", (name, currency))
        conn.commit()
        return {"id": cur.lastrowid, "name": name, "currency": currency}


@app.post("/plans/save")
def save_plan(campaign_id: int, payload: PlanRequest):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    plan = estimate_plan(payload)
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO plans (campaign_id, payload, result) VALUES (?, json(?), json(?))",
            (campaign_id, payload.model_dump_json(), plan.model_dump_json()),
        )
        conn.commit()
    return {"status": "ok", "plan": plan}


@app.post("/fact/import")
async def import_fact(campaign_id: int, file: UploadFile = File(...)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    content = await file.read()
    csv_text = content.decode("utf-8")
    fact_rows = parse_fact_csv(csv_text)
    with get_conn() as conn:
        for r in fact_rows:
            conn.execute(
                """
                INSERT INTO fact_rows (campaign_id, date, platform, ad_account_id, campaign_name, impressions, clicks, cost, leads, conversions, views, raw)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, json(?))
                """,
                (
                    campaign_id,
                    r.date.isoformat(),
                    r.platform,
                    r.ad_account_id,
                    r.campaign_name,
                    r.impressions,
                    r.clicks,
                    r.cost,
                    r.leads,
                    r.conversions,
                    r.views,
                    r.model_dump_json(),
                ),
            )
        conn.commit()
    return {"status": "ok", "rows": len(fact_rows)}


@app.get("/reports/weekly")
def report_weekly(campaign_id: int, plan_id: Optional[int] = None):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        if plan_id:
            row = conn.execute("SELECT result FROM plans WHERE id=? AND campaign_id=?", (plan_id, campaign_id)).fetchone()
        else:
            row = conn.execute(
                "SELECT result FROM plans WHERE campaign_id=? ORDER BY id DESC LIMIT 1",
                (campaign_id,),
            ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Plan not found")
        plan = PlanResponse.model_validate_json(row["result"])
        fact_rows_db = conn.execute(
            "SELECT * FROM fact_rows WHERE campaign_id=?",
            (campaign_id,),
        ).fetchall()
    fact_rows = [
        FactRow(
            date=x["date"],
            platform=x["platform"],
            ad_account_id=x["ad_account_id"],
            campaign_name=x["campaign_name"],
            impressions=x["impressions"],
            clicks=x["clicks"],
            cost=x["cost"],
            leads=x["leads"],
            conversions=x["conversions"],
            views=x["views"],
        )
        for x in fact_rows_db
    ]
    weekly = aggregate_weekly(plan, fact_rows)
    return {"weekly": weekly}


class TopUpStatus(str, Enum):
    pending = "pending"
    completed = "completed"
    failed = "failed"


class AdminTopupUpdate(BaseModel):
    amount_net: Optional[float] = None
    fx_rate: Optional[float] = None


class TopUpRequest(BaseModel):
    platform: Literal["meta", "google", "tiktok", "yandex", "telegram", "monochrome"]
    account_id: str
    email: str
    company: str
    budget: Optional[float] = None
    fee_percent: Optional[float] = None
    vat_percent: Optional[float] = None


class AuthPayload(BaseModel):
    email: str
    password: str


class ChangePasswordPayload(BaseModel):
    current_password: str
    new_password: str


class MetaInsightsResponse(BaseModel):
    summary: Dict[str, object]
    campaigns: List[Dict[str, object]]


class GoogleInsightsResponse(BaseModel):
    summary: Dict[str, object]
    campaigns: List[Dict[str, object]]


class TikTokInsightsResponse(BaseModel):
    summary: Dict[str, object]
    campaigns: List[Dict[str, object]]
    adgroups: List[Dict[str, object]]
    ads: List[Dict[str, object]]


class ProfilePayload(BaseModel):
    name: Optional[str] = None
    company: Optional[str] = None
    language: Optional[str] = None
    whatsapp_phone: Optional[str] = None
    telegram_handle: Optional[str] = None


class InvoicePendingPayload(BaseModel):
    request_id: Optional[int] = None
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    amount: Optional[float] = None
    currency: Optional[str] = None
    client_name: Optional[str] = None
    client_bin: Optional[str] = None
    client_address: Optional[str] = None
    order_ref: Optional[str] = None


def _hash_password(password: str, salt: str) -> str:
    return hashlib.sha256(f"{salt}{password}".encode("utf-8")).hexdigest()


def _verify_password(password: str, salt: str, stored_hash: str) -> bool:
    return hmac.compare_digest(_hash_password(password, salt), stored_hash)


def _get_bearer_token(authorization: Optional[str]) -> Optional[str]:
    if not authorization:
        return None
    parts = authorization.split(" ", 1)
    if len(parts) != 2 or parts[0].lower() != "bearer":
        return None
    return parts[1].strip()


def _get_user_by_token(token: Optional[str]):
    if not token:
        return None
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT u.* FROM user_tokens ut
            JOIN users u ON u.id = ut.user_id
            WHERE ut.token=?
            """,
            (token,),
        ).fetchone()
        return dict(row) if row else None


def get_current_user(authorization: Optional[str] = Header(None)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    token = _get_bearer_token(authorization)
    if not token:
        raise HTTPException(status_code=401, detail="Missing token")
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT u.* FROM user_tokens ut
            JOIN users u ON u.id = ut.user_id
            WHERE ut.token=?
            """,
            (token,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=401, detail="Invalid token")
        return dict(row)


def get_optional_user(authorization: Optional[str] = Header(None)):
    token = _get_bearer_token(authorization)
    return _get_user_by_token(token)


def get_admin_user(current_user=Depends(get_current_user)):
    if current_user["email"] not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


@app.get("/admin/check-key")
def admin_check_key(key: str):
    admin_key = os.getenv("ADMIN_PORTAL_KEY")
    if not admin_key:
        raise HTTPException(status_code=500, detail="Admin portal key not set")
    if key != admin_key:
        raise HTTPException(status_code=403, detail="Invalid key")
    return {"status": "ok"}


class AccountRequestCreate(BaseModel):
    platform: Literal["meta", "google", "tiktok", "yandex", "telegram", "monochrome"]
    name: str
    payload: Dict[str, object]


class AccountRequestUpdate(BaseModel):
    status: Literal["new", "processing", "approved", "rejected"]
    account_code: Optional[str] = None
    manager_email: Optional[str] = None
    comment: Optional[str] = None
    budget_total: Optional[float] = None


class AccountRequestEventCreate(BaseModel):
    type: Literal["comment", "status"]
    status: Optional[Literal["new", "processing", "approved", "rejected"]] = None
    manager_email: Optional[str] = None
    comment: Optional[str] = None


class AdminAccountCreate(BaseModel):
    user_id: int
    platform: Literal["meta", "google", "tiktok", "yandex", "telegram", "monochrome"]
    name: str
    external_id: Optional[str] = None
    account_code: Optional[str] = None
    currency: str = "USD"
    status: Optional[str] = None


class AdminAccountUpdate(BaseModel):
    user_id: Optional[int] = None
    platform: Optional[str] = None
    name: Optional[str] = None
    external_id: Optional[str] = None
    account_code: Optional[str] = None
    currency: Optional[str] = None
    status: Optional[str] = None


class PasswordReset(BaseModel):
    email: str
    new_password: str


class WalletAdjust(BaseModel):
    user_email: str
    amount: float
    note: Optional[str] = None


class FeeConfigPayload(BaseModel):
    meta: Optional[float] = None
    google: Optional[float] = None
    yandex: Optional[float] = None
    tiktok: Optional[float] = None
    telegram: Optional[float] = None
    monochrome: Optional[float] = None


class WalletTopupRequestPayload(BaseModel):
    amount: float = Field(..., gt=0)
    currency: str = "KZT"
    note: Optional[str] = None
    legal_entity_id: Optional[int] = None
    client_name: Optional[str] = None
    client_bin: Optional[str] = None
    client_address: Optional[str] = None
    client_email: Optional[str] = None
    order_ref: Optional[str] = None


class LegalEntityPayload(BaseModel):
    name: str
    short_name: Optional[str] = None
    full_name: Optional[str] = None
    bin: Optional[str] = None
    address: Optional[str] = None
    legal_address: Optional[str] = None
    email: Optional[str] = None
    bank: Optional[str] = None
    iban: Optional[str] = None
    bic: Optional[str] = None
    kbe: Optional[str] = None


class AdminLegalEntityPayload(BaseModel):
    user_email: str
    bin: str
    short_name: str
    full_name: str
    legal_address: str


class CompanyProfilePayload(BaseModel):
    name: Optional[str] = None
    bin: Optional[str] = None
    iin: Optional[str] = None
    legal_address: Optional[str] = None
    factual_address: Optional[str] = None


@app.post("/topup/request")
def topup_request(payload: TopUpRequest):
    # Stub: in production push to queue/CRM
    return {"status": "ok", "message": "Top-up request received", "payload": payload}


@app.post("/auth/register")
def register(payload: AuthPayload):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    email = payload.email.strip().lower()
    password = payload.password.strip()
    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and password are required")
    with get_conn() as conn:
        existing = conn.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
        if existing:
            raise HTTPException(status_code=400, detail="User already exists")
        salt = secrets.token_hex(8)
        password_hash = _hash_password(password, salt)
        cur = conn.execute(
            "INSERT INTO users (email, password_hash, salt) VALUES (?, ?, ?)",
            (email, password_hash, salt),
        )
        token = secrets.token_hex(24)
        conn.execute("INSERT INTO user_tokens (user_id, token) VALUES (?, ?)", (cur.lastrowid, token))
        conn.commit()
        user_id = cur.lastrowid
        _send_telegram_alert(
            "\n".join(
                [
                    "👤 <b>Новая регистрация</b>",
                    f"ID: <code>{user_id}</code>",
                    f"Email: <code>{email}</code>",
                ]
            ),
            channel="reg",
        )
        return {"id": user_id, "email": email, "token": token}


@app.post("/auth/login")
def login(payload: AuthPayload):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    email = payload.email.strip().lower()
    password = payload.password.strip()
    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and password are required")
    with get_conn() as conn:
        user = conn.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
        if not user or not user["password_hash"] or not user["salt"]:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        if not _verify_password(password, user["salt"], user["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        token = secrets.token_hex(24)
        conn.execute("INSERT INTO user_tokens (user_id, token) VALUES (?, ?)", (user["id"], token))
        conn.commit()
        return {"id": user["id"], "email": user["email"], "token": token}


@app.post("/admin/reset-password")
def admin_reset_password(payload: PasswordReset, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    email = payload.email.strip().lower()
    password = payload.new_password.strip()
    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and new_password are required")
    salt = secrets.token_hex(8)
    password_hash = _hash_password(password, salt)
    with get_conn() as conn:
        user = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        conn.execute(
            "UPDATE users SET password_hash=?, salt=? WHERE id=?",
            (password_hash, salt, user["id"]),
        )
        conn.execute("DELETE FROM user_tokens WHERE user_id=?", (user["id"],))
        conn.commit()
        return {"status": "ok"}


def _get_or_create_wallet(conn, user_id: int) -> Dict[str, object]:
    row = conn.execute("SELECT * FROM wallets WHERE user_id=?", (user_id,)).fetchone()
    if row:
        return dict(row)
    conn.execute(
        "INSERT INTO wallets (user_id, balance, currency, low_threshold) VALUES (?, ?, ?, ?)",
        (user_id, 0.0, "KZT", 50000),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM wallets WHERE user_id=?", (user_id,)).fetchone()
    return dict(row)


def _get_or_create_profile(conn, user_id: int) -> Dict[str, object]:
    row = conn.execute("SELECT * FROM user_profiles WHERE user_id=?", (user_id,)).fetchone()
    if row:
        return dict(row)
    try:
        conn.execute(
            """
            INSERT INTO user_profiles (user_id, name, company, language, fee_config, notifications_seen_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (user_id, None, None, "ru", json.dumps(_default_fee_config(), ensure_ascii=False), None),
        )
        conn.commit()
    except Exception:
        # Fallback for older schema without new columns.
        try:
            conn.execute(
                """
                INSERT INTO user_profiles (user_id, name, company, language)
                VALUES (?, ?, ?, ?)
                """,
                (user_id, None, None, "ru"),
            )
            conn.commit()
        except Exception:
            pass
    row = conn.execute("SELECT * FROM user_profiles WHERE user_id=?", (user_id,)).fetchone()
    return dict(row) if row else {"user_id": user_id}


def _document_storage_dir() -> str:
    base_dir = os.path.join(os.path.dirname(__file__), "..", "storage", "documents")
    os.makedirs(base_dir, exist_ok=True)
    return base_dir


def _save_document(file: UploadFile) -> str:
    safe_ext = os.path.splitext(file.filename or "")[1] or ".pdf"
    filename = f"doc_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(6)}{safe_ext}"
    path = os.path.join(_document_storage_dir(), filename)
    with open(path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return path


def _avatar_storage_dir() -> str:
    base_dir = os.path.join(os.path.dirname(__file__), "..", "storage", "avatars")
    os.makedirs(base_dir, exist_ok=True)
    return base_dir


def _save_avatar(file: UploadFile) -> str:
    safe_ext = os.path.splitext(file.filename or "")[1] or ".jpg"
    filename = f"avatar_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(6)}{safe_ext}"
    path = os.path.join(_avatar_storage_dir(), filename)
    with open(path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return path


def _meta_fetch_insights(account_external_id: str, date_from: str, date_to: str) -> List[Dict[str, object]]:
    token = os.getenv("META_ACCESS_TOKEN")
    if not token:
        raise HTTPException(status_code=500, detail="META_ACCESS_TOKEN is not set")
    url = f"https://graph.facebook.com/v20.0/act_{account_external_id}/insights"
    params = {
        "access_token": token,
        "level": "campaign",
        "fields": "campaign_id,campaign_name,account_id,account_currency,spend,ctr,cpc,cpm,reach,impressions,clicks",
        "time_range": json.dumps({"since": date_from, "until": date_to}),
    }
    resp = httpx.get(url, params=params, timeout=20)
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Meta API error: {resp.text}")
    data = resp.json()
    return data.get("data", [])

def _meta_fetch_breakdowns(
    account_external_id: str,
    date_from: str,
    date_to: str,
    breakdowns: List[str],
    level: str = "account",
) -> List[Dict[str, object]]:
    token = os.getenv("META_ACCESS_TOKEN")
    if not token:
        raise HTTPException(status_code=500, detail="META_ACCESS_TOKEN is not set")
    url = f"https://graph.facebook.com/v20.0/act_{account_external_id}/insights"
    params = {
        "access_token": token,
        "level": level,
        "fields": "impressions,clicks,spend,reach",
        "time_range": json.dumps({"since": date_from, "until": date_to}),
        "breakdowns": ",".join(breakdowns),
    }
    resp = httpx.get(url, params=params, timeout=30)
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Meta API error: {resp.text}")
    data = resp.json()
    return data.get("data", [])
def _meta_fetch_daily(account_external_id: str, date_from: str, date_to: str) -> List[Dict[str, object]]:
    token = os.getenv("META_ACCESS_TOKEN")
    if not token:
        raise HTTPException(status_code=500, detail="META_ACCESS_TOKEN is not set")
    url = f"https://graph.facebook.com/v20.0/act_{account_external_id}/insights"
    params = {
        "access_token": token,
        "level": "account",
        "fields": "spend,impressions,clicks",
        "time_increment": 1,
        "time_range": json.dumps({"since": date_from, "until": date_to}),
    }
    resp = httpx.get(url, params=params, timeout=20)
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Meta API error: {resp.text}")
    data = resp.json()
    return data.get("data", [])


def _live_billing_cache_get(key: str) -> Optional[Dict[str, object]]:
    item = _LIVE_BILLING_CACHE.get(key)
    if not item:
        return None
    if time.time() - float(item.get("ts") or 0) > _LIVE_BILLING_TTL_SEC:
        _LIVE_BILLING_CACHE.pop(key, None)
        return None
    data = item.get("data")
    return dict(data) if isinstance(data, dict) else None


def _live_billing_cache_set(key: str, data: Dict[str, object]) -> Dict[str, object]:
    payload = dict(data)
    _LIVE_BILLING_CACHE[key] = {"ts": time.time(), "data": payload}
    return dict(payload)


def _meta_fetch_account_billing(account_external_id: str, force_refresh: bool = False) -> Dict[str, object]:
    cache_key = f"meta:{account_external_id}"
    cached = _live_billing_cache_get(cache_key)
    if cached and not force_refresh:
        return cached

    token = os.getenv("META_ACCESS_TOKEN")
    if not token:
        raise HTTPException(status_code=500, detail="META_ACCESS_TOKEN is not set")

    url = f"https://graph.facebook.com/v20.0/act_{account_external_id}"
    params = {
        "access_token": token,
        "fields": "account_id,amount_spent,spend_cap,currency",
    }
    resp = httpx.get(url, params=params, timeout=20)
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Meta account billing error: {resp.text}")

    data = resp.json()
    currency = data.get("currency") or "USD"

    def _meta_money_from_minor(value: object) -> float:
        try:
            raw = float(value or 0)
        except (TypeError, ValueError):
            return 0.0
        zero_decimal = {"JPY", "KRW", "VND", "CLP", "PYG", "UGX", "XAF", "XOF", "XPF"}
        divisor = 1 if currency in zero_decimal else 100
        return raw / divisor

    spend = _meta_money_from_minor(data.get("amount_spent"))
    spend_cap_raw = data.get("spend_cap")
    spend_cap = None
    try:
        spend_cap_value = _meta_money_from_minor(spend_cap_raw) if spend_cap_raw not in (None, "") else None
        if spend_cap_value and spend_cap_value > 0:
            spend_cap = spend_cap_value
    except (TypeError, ValueError):
        spend_cap = None

    payload = {
        "provider": "meta",
        "currency": currency,
        "spend": spend,
        "limit": spend_cap,
        "balance": (spend_cap - spend) if (spend_cap is not None) else None,
        "source": "meta_api",
        "updated_at": datetime.utcnow().isoformat() + "Z",
    }
    return _live_billing_cache_set(cache_key, payload)


def _google_ads_client() -> GoogleAdsClient:
    developer_token = os.getenv("GOOGLE_ADS_DEVELOPER_TOKEN")
    client_id = os.getenv("GOOGLE_ADS_CLIENT_ID")
    client_secret = os.getenv("GOOGLE_ADS_CLIENT_SECRET")
    refresh_token = os.getenv("GOOGLE_ADS_REFRESH_TOKEN")
    login_customer_id = os.getenv("GOOGLE_ADS_LOGIN_CUSTOMER_ID") or None
    if not developer_token or not client_id or not client_secret or not refresh_token:
        raise HTTPException(status_code=500, detail="Google Ads API credentials are not set")
    config = {
        "developer_token": developer_token,
        "client_id": client_id,
        "client_secret": client_secret,
        "refresh_token": refresh_token,
        "use_proto_plus": True,
    }
    if login_customer_id:
        config["login_customer_id"] = login_customer_id
    return GoogleAdsClient.load_from_dict(config)


def _google_fetch_insights(customer_id: str, date_from: str, date_to: str) -> Tuple[List[Dict[str, object]], Optional[str]]:
    client = _google_ads_client()
    ga_service = client.get_service("GoogleAdsService")
    currency = None
    currency_query = "SELECT customer.currency_code FROM customer LIMIT 1"
    currency_resp = ga_service.search(customer_id=customer_id, query=currency_query)
    for row in currency_resp:
        currency = row.customer.currency_code
        break
    query = f"""
        SELECT
          campaign.id,
          campaign.name,
          metrics.impressions,
          metrics.clicks,
          metrics.ctr,
          metrics.average_cpc,
          metrics.average_cpm,
          metrics.cost_micros,
          metrics.conversions
        FROM campaign
        WHERE segments.date BETWEEN '{date_from}' AND '{date_to}'
    """
    rows = ga_service.search(customer_id=customer_id, query=query)
    campaigns: List[Dict[str, object]] = []
    for row in rows:
        metrics = row.metrics
        campaigns.append(
            {
                "campaign_id": row.campaign.id,
                "campaign_name": row.campaign.name,
                "impressions": int(metrics.impressions or 0),
                "clicks": int(metrics.clicks or 0),
                "ctr": float(metrics.ctr or 0),
                "cpc": float(metrics.average_cpc or 0) / 1_000_000 if metrics.average_cpc else 0,
                "cpm": float(metrics.average_cpm or 0) / 1_000_000 if metrics.average_cpm else 0,
                "spend": float(metrics.cost_micros or 0) / 1_000_000,
                "conversions": float(metrics.conversions or 0),
            }
        )
    return campaigns, currency


def _google_normalize_customer_id(customer_id: str) -> str:
    return "".join(ch for ch in str(customer_id or "") if ch.isdigit())


def _tiktok_normalize_advertiser_id(advertiser_id: object) -> str:
    raw = str(advertiser_id or "").strip()
    digits = "".join(ch for ch in raw if ch.isdigit())
    return digits or raw


def _google_fetch_account_billing(customer_id: str, force_refresh: bool = False) -> Dict[str, object]:
    normalized_customer_id = _google_normalize_customer_id(customer_id)
    cache_key = f"google:{normalized_customer_id}"
    cached = _live_billing_cache_get(cache_key)
    if cached and not force_refresh:
        return cached
    if not normalized_customer_id:
        payload = {
            "provider": "google",
            "currency": "USD",
            "spend": None,
            "limit": None,
            "balance": None,
            "error": "Google customer ID не задан или указан неверно",
            "source": "google_ads_api",
            "updated_at": datetime.utcnow().isoformat() + "Z",
        }
        return _live_billing_cache_set(cache_key, payload)

    client = _google_ads_client()
    ga_service = client.get_service("GoogleAdsService")

    currency = "USD"
    currency_query = "SELECT customer.currency_code FROM customer LIMIT 1"
    for row in ga_service.search(customer_id=normalized_customer_id, query=currency_query):
        currency = row.customer.currency_code or currency
        break

    query = """
        SELECT
          account_budget.id,
          account_budget.approved_spending_limit_micros,
          account_budget.adjusted_spending_limit_micros,
          account_budget.amount_served_micros
        FROM account_budget
        ORDER BY account_budget.id DESC
        LIMIT 1
    """
    rows = []
    try:
        rows = list(ga_service.search(customer_id=normalized_customer_id, query=query))
    except Exception:
        rows = []

    spend = None

    if not rows:
        payload = {
            "provider": "google",
            "currency": currency,
            "spend": spend,
            "limit": None,
            "balance": None,
            "source": "google_ads_api",
            "updated_at": datetime.utcnow().isoformat() + "Z",
        }
        return _live_billing_cache_set(cache_key, payload)

    budget = rows[0].account_budget
    spend_budget = float(budget.amount_served_micros or 0) / 1_000_000
    spend = spend_budget
    adjusted_limit = float(budget.adjusted_spending_limit_micros or 0) / 1_000_000
    approved_limit = float(budget.approved_spending_limit_micros or 0) / 1_000_000
    limit = adjusted_limit or approved_limit or None

    payload = {
        "provider": "google",
        "currency": currency,
        "spend": spend,
        "limit": limit,
        "balance": (limit - spend) if (limit is not None and spend is not None) else None,
        "source": "google_ads_api",
        "updated_at": datetime.utcnow().isoformat() + "Z",
    }
    return _live_billing_cache_set(cache_key, payload)


def _tiktok_fetch_account_billing(advertiser_id: str, force_refresh: bool = False) -> Dict[str, object]:
    normalized_advertiser_id = _tiktok_normalize_advertiser_id(advertiser_id)
    cache_key = f"tiktok:{normalized_advertiser_id}"
    cached = _live_billing_cache_get(cache_key)
    if cached and not force_refresh:
        return cached

    if not normalized_advertiser_id:
        payload = {
            "provider": "tiktok",
            "currency": "USD",
            "spend": None,
            "limit": None,
            "balance": None,
            "error": "TikTok advertiser_id не задан или указан неверно",
            "source": "tiktok_api",
            "updated_at": datetime.utcnow().isoformat() + "Z",
        }
        return _live_billing_cache_set(cache_key, payload)

    end_date = datetime.utcnow().date()
    start_date_raw = str(os.getenv("TIKTOK_SPEND_START_DATE") or "2020-01-01").strip()
    try:
        start_date = datetime.strptime(start_date_raw, "%Y-%m-%d").date()
    except Exception:
        start_date = date(2020, 1, 1)
    if start_date > end_date:
        start_date = end_date

    spend = None
    spend_error = None
    try:
        spend_rows = _tiktok_fetch_report(
            normalized_advertiser_id,
            start_date.strftime("%Y-%m-%d"),
            end_date.strftime("%Y-%m-%d"),
            "AUCTION_CAMPAIGN",
            ["campaign_id"],
            ["spend"],
        )
        spend = sum(float(row.get("spend") or 0) for row in spend_rows)
    except Exception:
        try:
            total_spend = 0.0
            cursor = start_date
            while cursor <= end_date:
                chunk_end = min(cursor + timedelta(days=89), end_date)
                chunk_rows = _tiktok_fetch_report(
                    normalized_advertiser_id,
                    cursor.strftime("%Y-%m-%d"),
                    chunk_end.strftime("%Y-%m-%d"),
                    "AUCTION_CAMPAIGN",
                    ["campaign_id"],
                    ["spend"],
                )
                total_spend += sum(float(row.get("spend") or 0) for row in chunk_rows)
                cursor = chunk_end + timedelta(days=1)
            spend = total_spend
        except Exception as exc:
            spend_error = str(exc)
            spend = None

    currency = "USD"
    limit = None
    balance = None
    try:
        url = "https://business-api.tiktok.com/open_api/v1.3/advertiser/balance/get/"
        params = {"advertiser_id": normalized_advertiser_id}
        headers = {"Access-Token": _tiktok_access_token()}
        resp = httpx.get(url, params=params, headers=headers, timeout=20)
        if resp.status_code == 200:
            payload = resp.json()
            if payload.get("code") in (0, None):
                data = payload.get("data") or {}
                entries: List[Dict[str, object]] = []
                if isinstance(data, dict):
                    entries.append(data)
                    if isinstance(data.get("list"), list):
                        entries.extend(item for item in data.get("list") if isinstance(item, dict))

                for entry in entries:
                    entry_currency = entry.get("currency") or entry.get("account_currency")
                    if entry_currency:
                        currency = str(entry_currency).upper()
                        break

                def _pick_numeric(keys: List[str]) -> Optional[float]:
                    for entry in entries:
                        for key in keys:
                            raw = entry.get(key)
                            try:
                                if raw is None or raw == "":
                                    continue
                                return float(raw)
                            except (TypeError, ValueError):
                                continue
                    return None

                balance = _pick_numeric(
                    [
                        "balance",
                        "available_balance",
                        "cash_balance",
                        "valid_cash_balance",
                        "remain_cash",
                    ]
                )
                limit = _pick_numeric(
                    [
                        "spend_cap",
                        "budget",
                        "total_budget",
                        "total_balance",
                    ]
                )
    except Exception:
        pass

    if limit is None and balance is not None and spend is not None:
        limit = balance + spend
    if balance is None and limit is not None and spend is not None:
        balance = limit - spend

    result_payload = {
        "provider": "tiktok",
        "currency": currency,
        "spend": spend,
        "limit": limit,
        "balance": balance,
        "source": "tiktok_api",
        "updated_at": datetime.utcnow().isoformat() + "Z",
    }
    if spend_error and spend is None and limit is None and balance is None:
        result_payload["error"] = spend_error
    return _live_billing_cache_set(cache_key, result_payload)


def _attach_live_billing(account: Dict[str, object], force_refresh: bool = False) -> Dict[str, object]:
    payload = dict(account)
    platform = str(payload.get("platform") or "").lower().strip()
    external_id = payload.get("external_id") or payload.get("account_code") or payload.get("name")
    payload["live_billing"] = None
    if not external_id or platform not in {"meta", "google", "tiktok"}:
        return payload

    try:
        if platform == "meta":
            payload["live_billing"] = _meta_fetch_account_billing(str(external_id), force_refresh=force_refresh)
        elif platform == "google":
            payload["live_billing"] = _google_fetch_account_billing(str(external_id), force_refresh=force_refresh)
        elif platform == "tiktok":
            payload["live_billing"] = _tiktok_fetch_account_billing(str(external_id), force_refresh=force_refresh)
    except Exception as exc:
        logging.exception("Failed to fetch live billing for %s account %s", platform, external_id)
        payload["live_billing"] = {
            "provider": platform,
            "error": str(exc),
            "source": f"{platform}_api",
            "updated_at": datetime.utcnow().isoformat() + "Z",
        }
    return payload


def _attach_live_billing_many(rows: List[Dict[str, object]], force_refresh: bool = False) -> List[Dict[str, object]]:
    return [_attach_live_billing(row, force_refresh=force_refresh) for row in rows]


def _resolve_topup_account_amount(row: Dict[str, object]) -> Optional[float]:
    amount_input = row.get("amount_input")
    amount_net = row.get("amount_net")
    fx_rate = row.get("fx_rate")
    input_currency = str(row.get("currency") or "").upper()
    account_currency = str(row.get("account_currency") or row.get("currency") or "").upper()

    try:
        amount_input_value = float(amount_input) if amount_input is not None else None
    except (TypeError, ValueError):
        amount_input_value = None
    try:
        amount_net_value = float(amount_net) if amount_net is not None else None
    except (TypeError, ValueError):
        amount_net_value = None
    try:
        fx_rate_value = float(fx_rate) if fx_rate is not None else None
    except (TypeError, ValueError):
        fx_rate_value = None

    if account_currency and input_currency and account_currency == input_currency:
        return amount_net_value if amount_net_value is not None else amount_input_value

    if fx_rate_value and fx_rate_value > 0 and amount_input_value is not None:
        calculated = amount_input_value / fx_rate_value
        if amount_net_value is None:
            return calculated
        if amount_net_value > amount_input_value * 0.95:
            return calculated
        return amount_net_value

    return amount_net_value if amount_net_value is not None else amount_input_value


def _attach_topup_account_amount(rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    try:
        rates_data = _fetch_bcc_rates()
    except Exception:
        rates_data = None
    prepared = []
    for row in rows:
        payload = dict(row)
        payload["amount_account"] = _resolve_topup_account_amount(payload)
        account_currency = payload.get("account_currency") or payload.get("currency") or "USD"
        payload["amount_account_usd"] = _convert_amount_to_usd(payload.get("amount_account"), account_currency, rates_data)
        payload["amount_account_kzt"] = _convert_amount_to_kzt(payload.get("amount_account"), account_currency, rates_data)
        def _num(value: object, default: Optional[float] = 0.0) -> Optional[float]:
            if value is None:
                return default
            try:
                return float(value)
            except (TypeError, ValueError):
                text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
                if text == "":
                    return default
                try:
                    return float(text)
                except (TypeError, ValueError):
                    return default

        amount_input_value = _num(payload.get("amount_input"), 0.0) or 0.0
        amount_account_value = _num(payload.get("amount_account"), 0.0) or 0.0
        fx_rate_value = _num(payload.get("fx_rate"), None)
        fee_percent_value = _num(payload.get("fee_percent"), 0.0) or 0.0
        amount_account_kzt_value = _num(payload.get("amount_account_kzt"), 0.0) or 0.0
        input_currency = str(payload.get("currency") or "KZT").upper()

        our_rate = None
        fx_profit_kzt = 0.0
        if fx_rate_value and fx_rate_value > 0:
            if fx_rate_value > 10:
                our_rate = fx_rate_value - 10.0
            else:
                our_rate = fx_rate_value
            if amount_account_value > 0 and amount_input_value > 0:
                fx_profit_kzt = amount_input_value - (amount_account_value * our_rate)
            if fx_profit_kzt < 0:
                fx_profit_kzt = 0.0

        fee_base_kzt = amount_input_value
        if amount_account_value > 0:
            if fx_rate_value and fx_rate_value > 0:
                fee_base_kzt = amount_account_value * fx_rate_value
            elif input_currency == "KZT":
                fee_base_kzt = amount_account_value
            elif amount_account_kzt_value > 0:
                fee_base_kzt = amount_account_kzt_value

        fee_amount_kzt = fee_base_kzt * (fee_percent_value / 100.0) if fee_base_kzt > 0 and fee_percent_value > 0 else 0.0
        payload["our_rate"] = our_rate
        payload["fee_base_kzt"] = round(fee_base_kzt, 2)
        payload["fx_profit_kzt"] = round(fx_profit_kzt, 2)
        payload["fee_amount_kzt"] = round(fee_amount_kzt, 2)
        payload["profit_total_kzt"] = round(fx_profit_kzt + fee_amount_kzt, 2)
        prepared.append(payload)
    return prepared


def _google_fetch_audience_age_gender(customer_id: str, date_from: str, date_to: str) -> Tuple[List[Dict[str, object]], Optional[str]]:
    client = _google_ads_client()
    ga_service = client.get_service("GoogleAdsService")
    query = f"""
        SELECT
          segments.age_range,
          segments.gender,
          metrics.impressions,
          metrics.clicks,
          metrics.cost_micros
        FROM campaign
        WHERE segments.date BETWEEN '{date_from}' AND '{date_to}'
    """
    rows = ga_service.search(customer_id=customer_id, query=query)
    data: List[Dict[str, object]] = []
    for row in rows:
        data.append(
            {
                "age_range": str(row.segments.age_range),
                "gender": str(row.segments.gender),
                "impressions": int(row.metrics.impressions or 0),
                "clicks": int(row.metrics.clicks or 0),
                "spend": float(row.metrics.cost_micros or 0) / 1_000_000,
            }
        )
    return data, None


def _google_fetch_audience_device(customer_id: str, date_from: str, date_to: str) -> List[Dict[str, object]]:
    client = _google_ads_client()
    ga_service = client.get_service("GoogleAdsService")
    query = f"""
        SELECT
          segments.device,
          metrics.impressions,
          metrics.clicks,
          metrics.cost_micros
        FROM customer
        WHERE segments.date BETWEEN '{date_from}' AND '{date_to}'
    """
    rows = ga_service.search(customer_id=customer_id, query=query)
    data: List[Dict[str, object]] = []
    for row in rows:
        data.append(
            {
                "device": str(row.segments.device),
                "impressions": int(row.metrics.impressions or 0),
                "clicks": int(row.metrics.clicks or 0),
                "spend": float(row.metrics.cost_micros or 0) / 1_000_000,
            }
        )
    return data


def _google_fetch_audience_geo(customer_id: str, date_from: str, date_to: str, level: str) -> List[Dict[str, object]]:
    client = _google_ads_client()
    ga_service = client.get_service("GoogleAdsService")
    segment = {
        "country": "segments.geo_target_country",
        "region": "segments.geo_target_region",
        "city": "segments.geo_target_city",
    }.get(level)
    if not segment:
        return []
    query = f"""
        SELECT
          {segment},
          metrics.impressions,
          metrics.clicks,
          metrics.cost_micros
        FROM campaign
        WHERE segments.date BETWEEN '{date_from}' AND '{date_to}'
    """
    rows = ga_service.search(customer_id=customer_id, query=query)
    data: List[Dict[str, object]] = []
    resource_names: List[str] = []
    for row in rows:
        geo_value = getattr(row.segments, segment.split(".")[1])
        resource_name = str(geo_value)
        resource_names.append(resource_name)
        data.append(
            {
                "geo": resource_name,
                "impressions": int(row.metrics.impressions or 0),
                "clicks": int(row.metrics.clicks or 0),
                "spend": float(row.metrics.cost_micros or 0) / 1_000_000,
            }
        )
    if not resource_names:
        return data
    try:
        name_map = _google_resolve_geo_names(client, customer_id, resource_names)
        for row in data:
            row["geo"] = name_map.get(row["geo"], row["geo"])
    except Exception:
        pass
    return data


def _google_resolve_geo_names(client: GoogleAdsClient, customer_id: str, resource_names: List[str]) -> Dict[str, str]:
    ga_service = client.get_service("GoogleAdsService")
    unique = sorted(set(resource_names))
    if not unique:
        return {}
    placeholders = ", ".join([f"'{name}'" for name in unique])
    query = f"""
        SELECT
          geo_target_constant.resource_name,
          geo_target_constant.name,
          geo_target_constant.target_type
        FROM geo_target_constant
        WHERE geo_target_constant.resource_name IN ({placeholders})
    """
    rows = ga_service.search(customer_id=customer_id, query=query)
    mapping: Dict[str, str] = {}
    for row in rows:
        mapping[row.geo_target_constant.resource_name] = row.geo_target_constant.name
    return mapping


def _google_fetch_daily(customer_id: str, date_from: str, date_to: str) -> List[Dict[str, object]]:
    client = _google_ads_client()
    ga_service = client.get_service("GoogleAdsService")
    query = f"""
        SELECT
          segments.date,
          metrics.impressions,
          metrics.clicks,
          metrics.ctr,
          metrics.average_cpc,
          metrics.average_cpm,
          metrics.cost_micros
        FROM customer
        WHERE segments.date BETWEEN '{date_from}' AND '{date_to}'
    """
    rows = ga_service.search(customer_id=customer_id, query=query)
    daily: List[Dict[str, object]] = []
    for row in rows:
        metrics = row.metrics
        daily.append(
            {
                "date": str(row.segments.date),
                "impressions": int(metrics.impressions or 0),
                "clicks": int(metrics.clicks or 0),
                "ctr": float(metrics.ctr or 0),
                "cpc": float(metrics.average_cpc or 0) / 1_000_000 if metrics.average_cpc else 0,
                "cpm": float(metrics.average_cpm or 0) / 1_000_000 if metrics.average_cpm else 0,
                "spend": float(metrics.cost_micros or 0) / 1_000_000,
            }
        )
    return daily


def _tiktok_access_token() -> str:
    token = os.getenv("TIKTOK_ACCESS_TOKEN")
    if not token:
        raise HTTPException(status_code=500, detail="TIKTOK_ACCESS_TOKEN is not set")
    return token


def _tiktok_fetch_report(
    advertiser_id: str,
    date_from: str,
    date_to: str,
    data_level: str,
    dimensions: List[str],
    metrics: List[str],
) -> List[Dict[str, object]]:
    def _sanitize_dimensions(values: List[str]) -> List[str]:
        blocked = {"campaign_name", "adgroup_name", "ad_name"}
        cleaned = [v for v in values if v not in blocked]
        return cleaned or [v for v in values if v]

    def _request_with_dimensions(current_dimensions: List[str]) -> Dict[str, object]:
        params = {
            "advertiser_id": advertiser_id,
            "report_type": "BASIC",
            "data_level": data_level,
            "dimensions": json.dumps(current_dimensions),
            "metrics": json.dumps(metrics),
            "start_date": date_from,
            "end_date": date_to,
            "page_size": 1000,
        }
        headers = {"Access-Token": _tiktok_access_token()}
        resp = httpx.get(url, params=params, headers=headers, timeout=30)
        if resp.status_code != 200:
            raise HTTPException(status_code=502, detail=f"TikTok API error: {resp.text}")
        return resp.json()

    url = "https://business-api.tiktok.com/open_api/v1.3/report/integrated/get/"
    payload = _request_with_dimensions(dimensions)
    if payload.get("code") not in (0, None):
        message = str(payload.get("message") or "")
        if int(payload.get("code") or 0) == 40002 and "dimensions" in message.lower():
            sanitized = _sanitize_dimensions(dimensions)
            if sanitized != dimensions:
                payload = _request_with_dimensions(sanitized)
        if payload.get("code") not in (0, None):
            raise HTTPException(status_code=502, detail=f"TikTok API error: {payload}")
    data = payload.get("data") or {}
    rows = data.get("list") or []
    results: List[Dict[str, object]] = []
    for row in rows:
        merged = {}
        merged.update(row.get("dimensions") or {})
        merged.update(row.get("metrics") or {})
        results.append(merged)
    return results


def _tiktok_fetch_daily(advertiser_id: str, date_from: str, date_to: str) -> List[Dict[str, object]]:
    rows = _tiktok_fetch_report(
        advertiser_id,
        date_from,
        date_to,
        "AUCTION_ADVERTISER",
        ["stat_time_day"],
        ["spend", "impressions", "clicks", "ctr", "cpc", "cpm"],
    )
    daily: List[Dict[str, object]] = []
    for row in rows:
        daily.append(
            {
                "date": row.get("stat_time_day"),
                "spend": row.get("spend"),
                "impressions": row.get("impressions"),
                "clicks": row.get("clicks"),
                "ctr": row.get("ctr"),
                "cpc": row.get("cpc"),
                "cpm": row.get("cpm"),
            }
        )
    return daily

def _invoice_storage_dir() -> str:
    base_dir = os.path.join(os.path.dirname(__file__), "..", "storage", "invoices")
    os.makedirs(base_dir, exist_ok=True)
    return base_dir


def _wallet_invoice_pdf_path(request_id: int) -> str:
    filename = f"wallet_invoice_{request_id}.pdf"
    return os.path.join(_invoice_storage_dir(), filename)


def _save_invoice_pdf(pdf: UploadFile) -> str:
    suffix = ".pdf"
    if pdf.filename and pdf.filename.lower().endswith(".pdf"):
        suffix = ".pdf"
    filename = f"invoice_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(6)}{suffix}"
    if _r2_enabled():
        key = f"invoices/{filename}"
        return _r2_upload_fileobj(key, pdf.file, "application/pdf")
    path = os.path.join(_invoice_storage_dir(), filename)
    with open(path, "wb") as f:
        shutil.copyfileobj(pdf.file, f)
    return path


def _format_ru_date(date_str: str) -> str:
    try:
        dt = datetime.fromisoformat(date_str)
    except Exception:
        return date_str
    months = [
        "января",
        "февраля",
        "марта",
        "апреля",
        "мая",
        "июня",
        "июля",
        "августа",
        "сентября",
        "октября",
        "ноября",
        "декабря",
    ]
    return f"{dt.day} {months[dt.month - 1]} {dt.year} г."


def _wallet_invoice_page_html(
    request_row: Dict[str, object],
    invoice_number: str,
    invoice_date: str,
    company: Dict[str, object],
    customer: Dict[str, object],
) -> str:
    amount_val = float(request_row["amount"])
    amount = _format_amount(amount_val)
    currency = request_row.get("currency") or "KZT"
    amount_words = _amount_to_words_ru(amount_val)
    date_ru = _format_ru_date(invoice_date)
    company_name = company.get("name") or "—"
    company_bin = company.get("bin") or "—"
    company_iin = company.get("iin") or ""
    company_address = company.get("legal_address") or company.get("factual_address") or ""
    company_bank = company.get("bank") or "—"
    company_iban = company.get("iban") or "—"
    company_bic = company.get("bic") or "—"
    company_kbe = company.get("kbe") or "—"

    customer_name = customer.get("name") or "—"
    customer_bin = customer.get("bin") or "—"
    customer_address = customer.get("address") or "—"

    purpose = (
        f"За услуги по использованию Программного обеспечения Исполнителя \"{company_name}\" "
        f"по счету {invoice_number} от {date_ru}, согласно Публичному договору возмездного оказания услуг от 22.04.2025 г."
    )

    return f"""
<!doctype html>
<html lang="ru">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Счет на оплату</title>
    <style>
      body {{
        font-family: "Arial", sans-serif;
        background: #ffffff;
        color: #111;
        margin: 0;
        padding: 24px;
      }}
      .page {{
        max-width: 900px;
        margin: 0 auto;
        background: #fff;
        border: 1px solid #ddd;
        padding: 18px 20px 28px;
      }}
      .header {{
        text-align: center;
        font-weight: 700;
        font-size: 12px;
        margin-bottom: 6px;
      }}
      .bank-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 12px;
      }}
      .bank-table td {{
        border: 1px solid #333;
        padding: 6px 8px;
        vertical-align: top;
      }}
      .section-title {{
        font-weight: 700;
        font-size: 16px;
        margin: 12px 0 8px;
      }}
      .subline {{
        font-size: 12px;
        margin: 4px 0;
      }}
      .alert {{
        color: #d32f2f;
        font-weight: 700;
        margin: 10px 0 6px;
        font-size: 12px;
      }}
      .purpose {{
        border: 1px solid #d32f2f;
        padding: 6px 8px;
        font-size: 12px;
        margin-bottom: 10px;
      }}
      .items {{
        width: 100%;
        border-collapse: collapse;
        font-size: 12px;
      }}
      .items th, .items td {{
        border: 1px solid #333;
        padding: 6px;
      }}
      .items th {{
        background: #f3f3f3;
      }}
      .total {{
        text-align: right;
        font-weight: 700;
        margin-top: 6px;
        font-size: 12px;
      }}
      .footnote {{
        font-size: 12px;
        margin-top: 10px;
      }}
      .sign {{
        margin-top: 24px;
        border-top: 1px solid #111;
        height: 40px;
      }}
    </style>
  </head>
  <body>
    <div class="page">
      <div class="header">Образец платежного поручения</div>
      <table class="bank-table">
        <tr>
          <td rowspan="2">
            Бенефициар:<br />
            {company_name}<br />
            БИН: {company_bin}
          </td>
          <td>ИИК<br />{company_iban}</td>
          <td>КБе<br />{company_kbe}</td>
        </tr>
        <tr>
          <td>БИК<br />{company_bic}</td>
          <td>Код назначения платежа<br />853</td>
        </tr>
        <tr>
          <td colspan="3">Банк бенефициара: {company_bank}</td>
        </tr>
      </table>

      <div class="subline">Счет действителен в течение 5 рабочих дней</div>

      <div class="section-title">Счет на оплату № {invoice_number} от {date_ru}</div>

      <div class="subline">
        Исполнитель: БИН / ИИН {company_bin}{f", {company_iin}" if company_iin else ""}, {company_name}, {company_address}
      </div>
      <div class="subline">
        Заказчик: БИН / ИИН {customer_bin}, {customer_name}, {customer_address}
      </div>
      <div class="subline">Договор: Публичный договор возмездного оказания услуг от 22.04.2025 г.</div>

      <div class="alert">Внимание! В назначение платежа скопируйте данные, указанные ниже.</div>
      <div class="purpose">{purpose}</div>

      <table class="items">
        <thead>
          <tr>
            <th>№</th>
            <th>Наименование</th>
            <th>Ед.</th>
            <th>Кол-во</th>
            <th>Цена</th>
            <th>Сумма</th>
          </tr>
        </thead>
        <tbody>
          <tr>
            <td>1</td>
            <td>За услуги по использованию Программного обеспечения Исполнителя "{company_name}"</td>
            <td>услуга</td>
            <td>1</td>
            <td>{amount}</td>
            <td>{amount}</td>
          </tr>
        </tbody>
      </table>

      <div class="total">Итого: {amount} {currency}</div>

      <div class="footnote">
        Всего наименований 1, на сумму {amount} {currency}<br />
        Всего к оплате: {amount_words} {currency}. Услуги Исполнителя НДС не облагаются (п.п. 46 ст. 394 Налогового кодекса Казахстана).
      </div>

      <div class="sign"></div>
    </div>
  </body>
</html>
"""


@app.get("/wallet")
def get_wallet(current_user=Depends(get_current_user)):
    if not get_conn:
        return {}
    with get_conn() as conn:
        wallet = _get_or_create_wallet(conn, current_user["id"])
        reserved_row = conn.execute(
            """
            SELECT COALESCE(SUM(amount_input + amount_input * (fee_percent / 100.0) + amount_input * (vat_percent / 100.0)), 0) AS reserved
            FROM topups
            WHERE user_id=? AND status='pending' AND COALESCE(hold_applied, 0)=1
            """,
            (current_user["id"],),
        ).fetchone()
        reserved = float((reserved_row["reserved"] if reserved_row else 0) or 0)
        payload = dict(wallet)
        payload["available_balance"] = float(wallet.get("balance") or 0)
        payload["reserved_balance"] = reserved
        return payload


@app.get("/admin/wallets")
def admin_list_wallets(admin_user=Depends(get_admin_user), low_only: bool = False):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT w.*, u.email as user_email
            FROM wallets w
            JOIN users u ON u.id = w.user_id
            ORDER BY w.balance ASC
            """
        ).fetchall()
        data = [dict(row) for row in rows]
        if low_only:
            data = [row for row in data if row["balance"] <= row["low_threshold"]]
        return data


@app.post("/admin/wallets/adjust")
def admin_adjust_wallet(payload: WalletAdjust, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    email = payload.user_email.strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="user_email is required")
    with get_conn() as conn:
        user = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        wallet = _get_or_create_wallet(conn, user["id"])
        new_balance = float(wallet["balance"]) + payload.amount
        conn.execute(
            "UPDATE wallets SET balance=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
            (new_balance, user["id"]),
        )
        conn.execute(
            """
            INSERT INTO wallet_transactions (user_id, account_id, amount, currency, type, note)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (user["id"], None, payload.amount, wallet["currency"], "adjustment", payload.note),
        )
        conn.commit()
        return {"user_id": user["id"], "balance": new_balance}


@app.get("/admin/wallet-transactions")
def admin_list_wallet_transactions(admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT wt.*, u.email as user_email, a.name as account_name, a.platform as account_platform
            FROM wallet_transactions wt
            JOIN users u ON u.id = wt.user_id
            LEFT JOIN ad_accounts a ON a.id = wt.account_id
            ORDER BY wt.created_at DESC
            """
        ).fetchall()
        return [dict(row) for row in rows]


@app.get("/wallet/transactions")
def list_wallet_transactions(current_user=Depends(get_current_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT wt.*, a.name as account_name, a.platform as account_platform
            FROM wallet_transactions wt
            LEFT JOIN ad_accounts a ON a.id = wt.account_id
            WHERE wt.user_id=?
            ORDER BY wt.created_at DESC
            """,
            (current_user["id"],),
        ).fetchall()
        return [dict(row) for row in rows]


@app.get("/profile")
def get_profile(current_user=Depends(get_current_user)):
    if not get_conn:
        return {}
    with get_conn() as conn:
        profile = _get_or_create_profile(conn, current_user["id"])
        profile["email"] = current_user["email"]
        if profile.get("avatar_path"):
            profile["avatar_url"] = f"/profile/avatar?token={_ensure_token(conn, current_user['id'])}"
        return profile


@app.put("/profile")
def update_profile(payload: ProfilePayload, current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        _get_or_create_profile(conn, current_user["id"])
        conn.execute(
            """
            UPDATE user_profiles
            SET name=?, company=?, language=?, whatsapp_phone=?, telegram_handle=?, updated_at=CURRENT_TIMESTAMP
            WHERE user_id=?
            """,
            (
                payload.name,
                payload.company,
                payload.language,
                payload.whatsapp_phone,
                payload.telegram_handle,
                current_user["id"],
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM user_profiles WHERE user_id=?", (current_user["id"],)).fetchone()
        result = dict(row) if row else {}
        result["email"] = current_user["email"]
        if result.get("avatar_path"):
            result["avatar_url"] = f"/profile/avatar?token={_ensure_token(conn, current_user['id'])}"
        return result


@app.get("/fees")
def get_fees(current_user=Depends(get_current_user)):
    if not get_conn:
        return _default_fee_config()
    with get_conn() as conn:
        profile = _get_or_create_profile(conn, current_user["id"])
        return _load_fee_config(profile.get("fee_config"))


def _ensure_token(conn, user_id: int) -> str:
    row = conn.execute("SELECT token FROM user_tokens WHERE user_id=? ORDER BY created_at DESC LIMIT 1", (user_id,)).fetchone()
    if row and row["token"]:
        return row["token"]
    token = secrets.token_hex(32)
    conn.execute("INSERT INTO user_tokens (user_id, token) VALUES (?, ?)", (user_id, token))
    conn.commit()
    return token


@app.post("/profile/avatar")
def upload_avatar(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if not file.filename:
        raise HTTPException(status_code=400, detail="File is required")
    with get_conn() as conn:
        profile = _get_or_create_profile(conn, current_user["id"])
        path = _save_avatar(file)
        conn.execute("UPDATE user_profiles SET avatar_path=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?", (path, current_user["id"]))
        conn.commit()
        token = _ensure_token(conn, current_user["id"])
        return {"status": "ok", "avatar_url": f"/profile/avatar?token={token}"}


@app.get("/profile/avatar")
def get_avatar(token: Optional[str] = None, current_user=Depends(get_optional_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if token:
        current_user = _get_user_by_token(token)
    if not current_user:
        raise HTTPException(status_code=401, detail="Missing token")
    with get_conn() as conn:
        row = conn.execute("SELECT avatar_path FROM user_profiles WHERE user_id=?", (current_user["id"],)).fetchone()
        if not row or not row["avatar_path"]:
            raise HTTPException(status_code=404, detail="Avatar not found")
        return FileResponse(row["avatar_path"])


@app.get("/notifications")
def list_notifications(current_user=Depends(get_current_user)):
    if not get_conn:
        return {"items": [], "unread": 0}
    with get_conn() as conn:
        try:
            profile = _get_or_create_profile(conn, current_user["id"])
            seen_at = profile.get("notifications_seen_at")
        except Exception:
            seen_at = None
        topups = conn.execute(
            """
            SELECT id, created_at, status, amount_input, amount_net, currency
            FROM topups
            WHERE user_id=? AND status='completed'
            ORDER BY created_at DESC
            LIMIT 10
            """,
            (current_user["id"],),
        ).fetchall()
        requests = conn.execute(
            """
            SELECT id, created_at, status, platform, name
            FROM account_requests
            WHERE user_id=? AND status='approved'
            ORDER BY created_at DESC
            LIMIT 10
            """,
            (current_user["id"],),
        ).fetchall()
        try:
            unread_topups = conn.execute(
                """
                SELECT COUNT(1) as cnt
                FROM topups
                WHERE user_id=? AND status='completed' AND (? IS NULL OR created_at > ?)
                """,
                (current_user["id"], seen_at, seen_at),
            ).fetchone()
            unread_requests = conn.execute(
                """
                SELECT COUNT(1) as cnt
                FROM account_requests
                WHERE user_id=? AND status='approved' AND (? IS NULL OR created_at > ?)
                """,
                (current_user["id"], seen_at, seen_at),
            ).fetchone()
        except Exception:
            unread_topups = None
            unread_requests = None
    items: List[Dict[str, object]] = []
    for row in topups:
        items.append(
            {
                "type": "topup",
                "id": row["id"],
                "created_at": row["created_at"],
                "title": "Пополнение",
                "status": row["status"],
                "amount": row["amount_net"] or row["amount_input"],
                "currency": row["currency"],
            }
        )
    for row in requests:
        items.append(
            {
                "type": "account_request",
                "id": row["id"],
                "created_at": row["created_at"],
                "title": "Аккаунт открыт",
                "status": row["status"],
                "platform": row["platform"],
                "name": row["name"],
            }
        )
    items.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
    unread = 0
    if unread_topups:
        unread += unread_topups["cnt"] if isinstance(unread_topups, dict) else unread_topups[0]
    if unread_requests:
        unread += unread_requests["cnt"] if isinstance(unread_requests, dict) else unread_requests[0]
    return {"items": items[:10], "unread": int(unread)}



@app.post("/notifications/read")
def mark_notifications_read(current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        _get_or_create_profile(conn, current_user["id"])
        conn.execute(
            "UPDATE user_profiles SET notifications_seen_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
            (current_user["id"],),
        )
        conn.commit()
        return {"status": "ok"}


@app.get("/admin/notifications")
def admin_notifications(admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        requests = conn.execute(
            """
            SELECT r.id, r.created_at, r.status, r.platform, r.name, u.email as user_email
            FROM account_requests r
            JOIN users u ON u.id = r.user_id
            WHERE r.status='new'
            ORDER BY r.created_at DESC
            LIMIT 10
            """
        ).fetchall()
        topups = conn.execute(
            """
            SELECT t.id, t.created_at, t.status, t.amount_input, t.amount_net, t.currency, u.email as user_email, a.platform, a.name
            FROM topups t
            JOIN users u ON u.id = t.user_id
            JOIN ad_accounts a ON a.id = t.account_id
            WHERE t.status='pending'
            ORDER BY t.created_at DESC
            LIMIT 10
            """
        ).fetchall()
    items: List[Dict[str, object]] = []
    for row in requests:
        items.append(
            {
                "type": "account_request",
                "id": row["id"],
                "created_at": row["created_at"],
                "title": "Новая заявка",
                "status": row["status"],
                "platform": row["platform"],
                "name": row["name"],
                "user_email": row["user_email"],
            }
        )
    for row in topups:
        items.append(
            {
                "type": "topup",
                "id": row["id"],
                "created_at": row["created_at"],
                "title": "Новая заявка на пополнение",
                "status": row["status"],
                "amount": row["amount_net"] or row["amount_input"],
                "currency": row["currency"],
                "platform": row["platform"],
                "name": row["name"],
                "user_email": row["user_email"],
            }
        )
    items.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
    return items[:12]


@app.post("/auth/change-password")
def change_password(payload: ChangePasswordPayload, current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if not payload.current_password or not payload.new_password:
        raise HTTPException(status_code=400, detail="current_password and new_password are required")
    with get_conn() as conn:
        user = conn.execute("SELECT * FROM users WHERE id=?", (current_user["id"],)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        if not _verify_password(payload.current_password, user["salt"], user["password_hash"]):
            raise HTTPException(status_code=400, detail="Invalid current password")
        salt = secrets.token_hex(8)
        password_hash = _hash_password(payload.new_password, salt)
        conn.execute(
            "UPDATE users SET password_hash=?, salt=? WHERE id=?",
            (password_hash, salt, current_user["id"]),
        )
        conn.execute("DELETE FROM user_tokens WHERE user_id=?", (current_user["id"],))
        new_token = secrets.token_hex(24)
        conn.execute("INSERT INTO user_tokens (user_id, token) VALUES (?, ?)", (current_user["id"], new_token))
        conn.commit()
        return {"status": "ok", "token": new_token}


@app.get("/documents")
def list_documents(current_user=Depends(get_current_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, title, created_at FROM user_documents WHERE user_id=? ORDER BY created_at DESC",
            (current_user["id"],),
        ).fetchall()
        return [dict(row) for row in rows]


@app.get("/documents/{doc_id}")
def download_document(
    doc_id: int,
    token: Optional[str] = None,
    current_user=Depends(get_optional_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if token:
        current_user = _get_user_by_token(token)
    if not current_user:
        raise HTTPException(status_code=401, detail="Missing token")
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM user_documents WHERE id=? AND user_id=?",
            (doc_id, current_user["id"]),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Document not found")
        return FileResponse(row["file_path"], filename=os.path.basename(row["file_path"]))


@app.get("/meta/insights", response_model=MetaInsightsResponse)
def meta_insights(
    date_from: str,
    date_to: str,
    account_id: Optional[int] = None,
    current_user=Depends(get_current_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="date_from and date_to are required")
    with get_conn() as conn:
        if account_id:
            row = conn.execute(
                "SELECT * FROM ad_accounts WHERE id=? AND user_id=? AND platform='meta'",
                (account_id, current_user["id"]),
            ).fetchone()
            accounts = [dict(row)] if row else []
        else:
            rows = conn.execute(
                "SELECT * FROM ad_accounts WHERE user_id=? AND platform='meta'",
                (current_user["id"],),
            ).fetchall()
            accounts = [dict(r) for r in rows]
    if not accounts:
        return {"summary": {"spend": 0, "ctr": 0, "cpc": 0, "cpm": 0, "reach": 0}, "campaigns": []}

    campaigns: List[Dict[str, object]] = []
    total_spend = 0.0
    total_impressions = 0.0
    total_clicks = 0.0
    total_reach = 0.0
    currency = None

    for acc in accounts:
        external_id = acc.get("external_id") or acc.get("account_code")
        if not external_id:
            continue
        rows = _meta_fetch_insights(external_id, date_from, date_to)
        for row in rows:
            spend = float(row.get("spend") or 0)
            impressions = float(row.get("impressions") or 0)
            clicks = float(row.get("clicks") or 0)
            reach = float(row.get("reach") or 0)
            total_spend += spend
            total_impressions += impressions
            total_clicks += clicks
            total_reach += reach
            currency = currency or row.get("account_currency")
            raw_ctr = float(row.get("ctr") or 0)
            ctr = raw_ctr / 100 if raw_ctr > 1 else raw_ctr
            campaigns.append(
                {
                    "campaign_id": row.get("campaign_id"),
                    "campaign_name": row.get("campaign_name"),
                    "account_id": row.get("account_id"),
                    "account_currency": row.get("account_currency"),
                    "spend": spend,
                    "ctr": ctr,
                    "cpc": float(row.get("cpc") or 0),
                    "cpm": float(row.get("cpm") or 0),
                    "reach": reach,
                    "impressions": impressions,
                    "clicks": clicks,
                }
            )

    ctr = (total_clicks / total_impressions) if total_impressions else 0.0
    cpc = (total_spend / total_clicks) if total_clicks else 0.0
    cpm = (total_spend / total_impressions * 1000) if total_impressions else 0.0

    summary = {
        "spend": total_spend,
        "ctr": ctr,
        "cpc": cpc,
        "cpm": cpm,
        "reach": total_reach,
        "impressions": total_impressions,
        "clicks": total_clicks,
        "currency": currency or "USD",
    }
    return {"summary": summary, "campaigns": campaigns}


@app.get("/google/insights", response_model=GoogleInsightsResponse)
def google_insights(
    date_from: str,
    date_to: str,
    account_id: Optional[int] = None,
    current_user=Depends(get_current_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="date_from and date_to are required")
    with get_conn() as conn:
        if account_id:
            row = conn.execute(
                "SELECT * FROM ad_accounts WHERE id=? AND user_id=? AND platform='google'",
                (account_id, current_user["id"]),
            ).fetchone()
            accounts = [dict(row)] if row else []
        else:
            rows = conn.execute(
                "SELECT * FROM ad_accounts WHERE user_id=? AND platform='google'",
                (current_user["id"],),
            ).fetchall()
            accounts = [dict(r) for r in rows]
    if not accounts:
        return {"summary": {"spend": 0, "ctr": 0, "cpc": 0, "cpm": 0, "impressions": 0, "clicks": 0}, "campaigns": []}

    campaigns: List[Dict[str, object]] = []
    total_spend = 0.0
    total_impressions = 0.0
    total_clicks = 0.0
    total_conversions = 0.0
    currency = None

    for acc in accounts:
        external_id = acc.get("external_id") or acc.get("account_code")
        if not external_id:
            continue
        try:
            rows, acc_currency = _google_fetch_insights(str(external_id), date_from, date_to)
        except google_api_exceptions.GoogleAPICallError as exc:
            message = getattr(exc, "message", None) or str(exc)
            raise HTTPException(status_code=502, detail=f"Google Ads API error: {message}")
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"Google Ads error: {exc}")
        currency = currency or acc_currency
        for row in rows:
            total_spend += float(row.get("spend") or 0)
            total_impressions += float(row.get("impressions") or 0)
            total_clicks += float(row.get("clicks") or 0)
            total_conversions += float(row.get("conversions") or 0)
            campaigns.append(row)

    ctr = (total_clicks / total_impressions) if total_impressions else 0.0
    cpc = (total_spend / total_clicks) if total_clicks else 0.0
    cpm = (total_spend / total_impressions * 1000) if total_impressions else 0.0

    summary = {
        "spend": total_spend,
        "ctr": ctr,
        "cpc": cpc,
        "cpm": cpm,
        "impressions": total_impressions,
        "clicks": total_clicks,
        "conversions": total_conversions,
        "currency": currency or "USD",
    }
    return {"summary": summary, "campaigns": campaigns}


@app.get("/tiktok/insights", response_model=TikTokInsightsResponse)
def tiktok_insights(
    date_from: str,
    date_to: str,
    account_id: Optional[int] = None,
    current_user=Depends(get_current_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="date_from and date_to are required")
    with get_conn() as conn:
        if account_id:
            row = conn.execute(
                "SELECT * FROM ad_accounts WHERE id=? AND user_id=? AND platform='tiktok'",
                (account_id, current_user["id"]),
            ).fetchone()
            accounts = [dict(row)] if row else []
        else:
            rows = conn.execute(
                "SELECT * FROM ad_accounts WHERE user_id=? AND platform='tiktok'",
                (current_user["id"],),
            ).fetchall()
            accounts = [dict(r) for r in rows]
    if not accounts:
        return {"summary": {"spend": 0, "ctr": 0, "cpc": 0, "cpm": 0, "impressions": 0, "clicks": 0}, "campaigns": [], "adgroups": [], "ads": []}

    def _to_float(value: object) -> float:
        try:
            return float(value)
        except Exception:
            return 0.0

    total_spend = 0.0
    total_impressions = 0.0
    total_clicks = 0.0

    campaigns: List[Dict[str, object]] = []
    adgroups: List[Dict[str, object]] = []
    ads: List[Dict[str, object]] = []

    metrics = ["spend", "impressions", "clicks", "ctr", "cpc", "cpm"]
    summary_currency = None

    for acc in accounts:
        advertiser_id = acc.get("external_id") or acc.get("account_code")
        if not advertiser_id:
            if account_id:
                raise HTTPException(
                    status_code=400,
                    detail=f"Для аккаунта TikTok id={acc.get('id')} не указан advertiser id (external_id/account_code).",
                )
            continue
        summary_currency = summary_currency or acc.get("currency")
        advertiser_id = _tiktok_normalize_advertiser_id(advertiser_id)
        campaign_rows = _tiktok_fetch_report(
            str(advertiser_id),
            date_from,
            date_to,
            "AUCTION_CAMPAIGN",
            ["campaign_id"],
            metrics,
        )
        adgroup_rows = _tiktok_fetch_report(
            str(advertiser_id),
            date_from,
            date_to,
            "AUCTION_ADGROUP",
            ["adgroup_id"],
            metrics,
        )
        ad_rows = _tiktok_fetch_report(
            str(advertiser_id),
            date_from,
            date_to,
            "AUCTION_AD",
            ["ad_id"],
            metrics,
        )
        for row in campaign_rows:
            spend = _to_float(row.get("spend"))
            impressions = _to_float(row.get("impressions"))
            clicks = _to_float(row.get("clicks"))
            total_spend += spend
            total_impressions += impressions
            total_clicks += clicks
            campaigns.append(
                {
                    "campaign_id": row.get("campaign_id"),
                    "campaign_name": row.get("campaign_name"),
                    "spend": spend,
                    "impressions": impressions,
                    "clicks": clicks,
                    "ctr": _to_float(row.get("ctr")),
                    "cpc": _to_float(row.get("cpc")),
                    "cpm": _to_float(row.get("cpm")),
                }
            )
        for row in adgroup_rows:
            adgroups.append(
                {
                    "adgroup_id": row.get("adgroup_id"),
                    "adgroup_name": row.get("adgroup_name"),
                    "campaign_id": row.get("campaign_id"),
                    "campaign_name": row.get("campaign_name"),
                    "spend": _to_float(row.get("spend")),
                    "impressions": _to_float(row.get("impressions")),
                    "clicks": _to_float(row.get("clicks")),
                    "ctr": _to_float(row.get("ctr")),
                    "cpc": _to_float(row.get("cpc")),
                    "cpm": _to_float(row.get("cpm")),
                }
            )
        for row in ad_rows:
            ads.append(
                {
                    "ad_id": row.get("ad_id"),
                    "ad_name": row.get("ad_name"),
                    "adgroup_id": row.get("adgroup_id"),
                    "adgroup_name": row.get("adgroup_name"),
                    "campaign_id": row.get("campaign_id"),
                    "campaign_name": row.get("campaign_name"),
                    "spend": _to_float(row.get("spend")),
                    "impressions": _to_float(row.get("impressions")),
                    "clicks": _to_float(row.get("clicks")),
                    "ctr": _to_float(row.get("ctr")),
                    "cpc": _to_float(row.get("cpc")),
                    "cpm": _to_float(row.get("cpm")),
                }
            )

    ctr = (total_clicks / total_impressions) if total_impressions else 0.0
    cpc = (total_spend / total_clicks) if total_clicks else 0.0
    cpm = (total_spend / total_impressions * 1000) if total_impressions else 0.0

    summary = {
        "spend": total_spend,
        "ctr": ctr,
        "cpc": cpc,
        "cpm": cpm,
        "impressions": total_impressions,
        "clicks": total_clicks,
        "currency": summary_currency or "USD",
    }
    return {"summary": summary, "campaigns": campaigns, "adgroups": adgroups, "ads": ads}


@app.get("/meta/audience")
def meta_audience(
    date_from: str,
    date_to: str,
    group: str,
    account_id: Optional[int] = None,
    current_user=Depends(get_current_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="date_from and date_to are required")
    if group not in {"age_gender", "geo", "placement_device", "device"}:
        raise HTTPException(status_code=400, detail="Unsupported group")
    with get_conn() as conn:
        if account_id:
            row = conn.execute(
                "SELECT * FROM ad_accounts WHERE id=? AND user_id=? AND platform='meta'",
                (account_id, current_user["id"]),
            ).fetchone()
            accounts = [dict(row)] if row else []
        else:
            rows = conn.execute(
                "SELECT * FROM ad_accounts WHERE user_id=? AND platform='meta'",
                (current_user["id"],),
            ).fetchall()
            accounts = [dict(r) for r in rows]
    results = []
    for acc in accounts:
        external_id = acc.get("external_id") or acc.get("account_code")
        if not external_id:
            continue
        payload: Dict[str, object] = {"account_id": acc.get("id"), "name": acc.get("name") or external_id}
        if group == "age_gender":
            payload["age_gender"] = _meta_fetch_breakdowns(str(external_id), date_from, date_to, ["age", "gender"])
        elif group == "geo":
            payload["country"] = _meta_fetch_breakdowns(str(external_id), date_from, date_to, ["country"])
            payload["region"] = _meta_fetch_breakdowns(str(external_id), date_from, date_to, ["region"])
        else:
            payload["publisher_platform"] = _meta_fetch_breakdowns(
                str(external_id), date_from, date_to, ["publisher_platform"]
            )
            payload["impression_device"] = _meta_fetch_breakdowns(
                str(external_id), date_from, date_to, ["impression_device"]
            )
            payload["device_platform"] = _meta_fetch_breakdowns(
                str(external_id), date_from, date_to, ["device_platform"]
            )
        results.append(payload)
    return {"accounts": results}


@app.get("/google/audience")
def google_audience(
    date_from: str,
    date_to: str,
    group: str,
    account_id: Optional[int] = None,
    current_user=Depends(get_current_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="date_from and date_to are required")
    if group not in {"age_gender", "geo", "device"}:
        raise HTTPException(status_code=400, detail="Unsupported group")
    with get_conn() as conn:
        if account_id:
            row = conn.execute(
                "SELECT * FROM ad_accounts WHERE id=? AND user_id=? AND platform='google'",
                (account_id, current_user["id"]),
            ).fetchone()
            accounts = [dict(row)] if row else []
        else:
            rows = conn.execute(
                "SELECT * FROM ad_accounts WHERE user_id=? AND platform='google'",
                (current_user["id"],),
            ).fetchall()
            accounts = [dict(r) for r in rows]

    results = []
    for acc in accounts:
        customer_id = acc.get("external_id") or acc.get("account_code")
        if not customer_id:
            continue
        payload: Dict[str, object] = {"account_id": acc.get("id"), "name": acc.get("name") or customer_id}
        try:
            if group == "age_gender":
                rows, _ = _google_fetch_audience_age_gender(str(customer_id), date_from, date_to)
                payload["age_gender"] = rows
            elif group == "device":
                rows = _google_fetch_audience_device(str(customer_id), date_from, date_to)
                payload["device"] = rows
            else:
                payload["country"] = _google_fetch_audience_geo(str(customer_id), date_from, date_to, "country")
                payload["region"] = _google_fetch_audience_geo(str(customer_id), date_from, date_to, "region")
                payload["city"] = _google_fetch_audience_geo(str(customer_id), date_from, date_to, "city")
        except Exception as exc:
            payload["error"] = str(exc)
        results.append(payload)
    return {"accounts": results}

@app.get("/insights/overview")
def insights_overview(
    date_from: str,
    date_to: str,
    meta_account_id: Optional[int] = None,
    google_account_id: Optional[int] = None,
    tiktok_account_id: Optional[int] = None,
    current_user=Depends(get_current_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="date_from and date_to are required")

    def _to_float(value: object) -> float:
        try:
            return float(value)
        except Exception:
            return 0.0

    def _merge_daily(target: Dict[str, Dict[str, object]], date_key: str, row: Dict[str, object]) -> None:
        date_val = row.get(date_key)
        if not date_val:
            return
        if date_val not in target:
            target[date_val] = {"date": date_val, "spend": 0.0, "impressions": 0.0, "clicks": 0.0}
        target[date_val]["spend"] += _to_float(row.get("spend"))
        target[date_val]["impressions"] += _to_float(row.get("impressions"))
        target[date_val]["clicks"] += _to_float(row.get("clicks"))

    def _fetch_accounts(conn, platform: str, account_id: Optional[int]) -> List[Dict[str, object]]:
        if account_id:
            row = conn.execute(
                "SELECT * FROM ad_accounts WHERE id=? AND user_id=? AND platform=?",
                (account_id, current_user["id"], platform),
            ).fetchone()
            return [dict(row)] if row else []
        rows = conn.execute(
            "SELECT * FROM ad_accounts WHERE user_id=? AND platform=?",
            (current_user["id"], platform),
        ).fetchall()
        return [dict(r) for r in rows]

    with get_conn() as conn:
        meta_accounts = _fetch_accounts(conn, "meta", meta_account_id)
        google_accounts = _fetch_accounts(conn, "google", google_account_id)
        tiktok_accounts = _fetch_accounts(conn, "tiktok", tiktok_account_id)

    totals = {"meta": {"spend": 0.0, "impressions": 0.0, "clicks": 0.0},
              "google": {"spend": 0.0, "impressions": 0.0, "clicks": 0.0},
              "tiktok": {"spend": 0.0, "impressions": 0.0, "clicks": 0.0}}

    daily_meta: Dict[str, Dict[str, object]] = {}
    daily_google: Dict[str, Dict[str, object]] = {}
    daily_tiktok: Dict[str, Dict[str, object]] = {}

    for acc in meta_accounts:
        external_id = acc.get("external_id") or acc.get("account_code")
        if not external_id:
            continue
        try:
            rows = _meta_fetch_daily(str(external_id), date_from, date_to)
            for row in rows:
                _merge_daily(daily_meta, "date_start", row)
        except Exception:
            continue

    for acc in google_accounts:
        external_id = acc.get("external_id") or acc.get("account_code")
        if not external_id:
            continue
        try:
            rows = _google_fetch_daily(str(external_id), date_from, date_to)
            for row in rows:
                _merge_daily(daily_google, "date", row)
        except Exception:
            continue

    advertiser_ids: List[str] = []
    for acc in tiktok_accounts:
        adv_id = acc.get("external_id") or acc.get("account_code")
        if adv_id:
            advertiser_ids.append(str(adv_id))
    if not advertiser_ids:
        env_adv = os.getenv("TIKTOK_ADVERTISER_ID")
        if env_adv:
            advertiser_ids.append(str(env_adv))
    for adv_id in sorted(set(advertiser_ids)):
        try:
            rows = _tiktok_fetch_daily(adv_id, date_from, date_to)
            for row in rows:
                _merge_daily(daily_tiktok, "date", row)
        except Exception:
            continue

    def _finalize(daily_map: Dict[str, Dict[str, object]], platform: str) -> List[Dict[str, object]]:
        rows = [daily_map[k] for k in sorted(daily_map.keys())]
        totals[platform]["spend"] = sum(_to_float(r.get("spend")) for r in rows)
        totals[platform]["impressions"] = sum(_to_float(r.get("impressions")) for r in rows)
        totals[platform]["clicks"] = sum(_to_float(r.get("clicks")) for r in rows)
        return rows

    daily = {
        "meta": _finalize(daily_meta, "meta"),
        "google": _finalize(daily_google, "google"),
        "tiktok": _finalize(daily_tiktok, "tiktok"),
    }

    return {"totals": totals, "daily": daily}


@app.post("/admin/documents/upload")
def admin_upload_document(
    email: str = Form(...),
    title: str = Form(...),
    file: UploadFile = File(...),
    admin_user=Depends(get_admin_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if not title.strip():
        raise HTTPException(status_code=400, detail="title is required")
    with get_conn() as conn:
        user = conn.execute("SELECT id FROM users WHERE email=?", (email.strip().lower(),)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        path = _save_document(file)
        cur = conn.execute(
            "INSERT INTO user_documents (user_id, title, file_path) VALUES (?, ?, ?)",
            (user["id"], title.strip(), path),
        )
        conn.commit()
        return {"id": cur.lastrowid, "status": "ok"}


@app.get("/admin/company-profile")
def admin_get_company_profile(admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        return _get_company_profile(conn)


@app.put("/admin/company-profile")
def admin_update_company_profile(payload: CompanyProfilePayload, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        profile = _get_company_profile(conn)
        if payload.name is not None:
            profile["name"] = payload.name.strip() or None
        if payload.bin is not None:
            profile["bin"] = payload.bin.strip() or None
        if payload.iin is not None:
            profile["iin"] = payload.iin.strip() or None
        if payload.legal_address is not None:
            profile["legal_address"] = payload.legal_address.strip() or None
        if payload.factual_address is not None:
            profile["factual_address"] = payload.factual_address.strip() or None
        conn.execute(
            """
            INSERT INTO company_profile
            (id, name, bin, iin, legal_address, factual_address, bank, iban, bic, kbe, currency, updated_at)
            VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(id) DO UPDATE SET
              name=excluded.name,
              bin=excluded.bin,
              iin=excluded.iin,
              legal_address=excluded.legal_address,
              factual_address=excluded.factual_address,
              bank=excluded.bank,
              iban=excluded.iban,
              bic=excluded.bic,
              kbe=excluded.kbe,
              currency=excluded.currency,
              updated_at=CURRENT_TIMESTAMP
            """,
            (
                profile.get("name"),
                profile.get("bin"),
                profile.get("iin"),
                profile.get("legal_address"),
                profile.get("factual_address"),
                profile.get("bank"),
                profile.get("iban"),
                profile.get("bic"),
                profile.get("kbe"),
                profile.get("currency"),
            ),
        )
        conn.commit()
        return profile


@app.get("/admin/legal-entities")
def admin_list_legal_entities(admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT le.*, u.email as user_email
            FROM legal_entities le
            JOIN user_legal_entities ule ON ule.legal_entity_id = le.id
            JOIN users u ON u.id = ule.user_id
            ORDER BY le.created_at DESC
            """
        ).fetchall()
        return [dict(row) for row in rows]


@app.post("/admin/legal-entities")
def admin_create_legal_entity(payload: AdminLegalEntityPayload, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    user_email = payload.user_email.strip().lower()
    if not user_email:
        raise HTTPException(status_code=400, detail="user_email is required")
    if not payload.bin.strip() or not payload.short_name.strip() or not payload.full_name.strip() or not payload.legal_address.strip():
        raise HTTPException(status_code=400, detail="bin, short_name, full_name and legal_address are required")
    with get_conn() as conn:
        user = conn.execute("SELECT id FROM users WHERE email=?", (user_email,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        cur = conn.execute(
            """
            INSERT INTO legal_entities (name, short_name, full_name, bin, address, legal_address)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                payload.short_name.strip(),
                payload.short_name.strip(),
                payload.full_name.strip(),
                payload.bin.strip(),
                payload.legal_address.strip(),
                payload.legal_address.strip(),
            ),
        )
        conn.execute(
            """
            INSERT OR IGNORE INTO user_legal_entities (user_id, legal_entity_id, is_default)
            VALUES (?, ?, 0)
            """,
            (user["id"], cur.lastrowid),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM legal_entities WHERE id=?", (cur.lastrowid,)).fetchone()
        result = dict(row)
        result["user_email"] = user_email
        return result


@app.put("/admin/legal-entities/{entity_id}")
def admin_update_legal_entity(entity_id: int, payload: AdminLegalEntityPayload, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    user_email = payload.user_email.strip().lower()
    if not user_email:
        raise HTTPException(status_code=400, detail="user_email is required")
    if not payload.bin.strip() or not payload.short_name.strip() or not payload.full_name.strip() or not payload.legal_address.strip():
        raise HTTPException(status_code=400, detail="bin, short_name, full_name and legal_address are required")
    with get_conn() as conn:
        entity = conn.execute("SELECT id FROM legal_entities WHERE id=?", (entity_id,)).fetchone()
        if not entity:
            raise HTTPException(status_code=404, detail="Legal entity not found")
        user = conn.execute("SELECT id FROM users WHERE email=?", (user_email,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        conn.execute(
            """
            UPDATE legal_entities
            SET name=?, short_name=?, full_name=?, bin=?, address=?, legal_address=?
            WHERE id=?
            """,
            (
                payload.short_name.strip(),
                payload.short_name.strip(),
                payload.full_name.strip(),
                payload.bin.strip(),
                payload.legal_address.strip(),
                payload.legal_address.strip(),
                entity_id,
            ),
        )
        conn.execute(
            """
            INSERT OR IGNORE INTO user_legal_entities (user_id, legal_entity_id, is_default)
            VALUES (?, ?, 0)
            """,
            (user["id"], entity_id),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM legal_entities WHERE id=?", (entity_id,)).fetchone()
        result = dict(row)
        result["user_email"] = user_email
        return result


@app.post("/wallet/topup-requests")
def create_wallet_topup_request(payload: WalletTopupRequestPayload, current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        legal_entity = None
        if payload.legal_entity_id:
            legal_entity = conn.execute(
                """
                SELECT le.*
                FROM legal_entities le
                JOIN user_legal_entities ule ON ule.legal_entity_id = le.id
                WHERE le.id=? AND ule.user_id=?
                """,
                (payload.legal_entity_id, current_user["id"]),
            ).fetchone()
            if not legal_entity:
                raise HTTPException(status_code=404, detail="Legal entity not found")
            legal_entity = dict(legal_entity)
        entity_name = _format_legal_entity_name(legal_entity) if legal_entity else None
        entity_address = (legal_entity.get("legal_address") or legal_entity.get("address")) if legal_entity else None
        client_name = payload.client_name or entity_name
        client_bin = payload.client_bin or (legal_entity.get("bin") if legal_entity else None)
        client_address = payload.client_address or entity_address
        client_email = payload.client_email or (legal_entity.get("email") if legal_entity else None)
        cur = conn.execute(
            """
            INSERT INTO wallet_topup_requests
            (user_id, amount, currency, note, status, legal_entity_id, client_name, client_bin, client_address, client_email, order_ref)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                current_user["id"],
                payload.amount,
                payload.currency,
                payload.note,
                "requested",
                payload.legal_entity_id,
                client_name,
                client_bin,
                client_address,
                client_email,
                payload.order_ref,
            ),
        )
        conn.commit()
        request_id = cur.lastrowid
        one_c_url = os.getenv("ONEC_REQUEST_URL")
        if one_c_url:
            try:
                httpx.post(
                    one_c_url,
                    json={
                        "request_id": request_id,
                        "user_id": current_user["id"],
                        "user_email": current_user["email"],
                        "amount": payload.amount,
                        "currency": payload.currency,
                        "note": payload.note,
                        "legal_entity_id": payload.legal_entity_id,
                        "client_name": client_name,
                        "client_bin": client_bin,
                        "client_address": client_address,
                        "client_email": client_email,
                        "order_ref": payload.order_ref,
                    },
                    timeout=10,
                )
            except Exception:
                pass
        _send_telegram_alert(
            "\n".join(
                [
                    "🧾 <b>Запрос на пополнение кошелька</b>",
                    f"ID: <code>{request_id}</code>",
                    f"Пользователь: <code>{current_user['email']}</code> (id={current_user['id']})",
                    f"Сумма: <b>{payload.amount:.2f} {payload.currency}</b>",
                    f"Контрагент: <b>{client_name or '—'}</b>",
                    f"БИН/ИИН: <code>{client_bin or '—'}</code>",
                    f"Order Ref: <code>{payload.order_ref or '—'}</code>",
                ]
            )
        )
        return {
            "id": request_id,
            "status": "requested",
            "invoice_url": f"/wallet/topup-requests/{request_id}/invoice",
        }


@app.get("/wallet/topup-requests")
def list_wallet_topup_requests(current_user=Depends(get_current_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT r.*, le.name as legal_entity_name, le.bin as legal_entity_bin,
                   i.invoice_number, i.invoice_date, i.amount as invoice_amount, i.currency as invoice_currency
            FROM wallet_topup_requests r
            LEFT JOIN legal_entities le ON le.id = r.legal_entity_id
            LEFT JOIN invoice_uploads i ON i.id = (
                SELECT id FROM invoice_uploads WHERE request_id = r.id ORDER BY created_at DESC LIMIT 1
            )
            WHERE r.user_id=?
            ORDER BY r.created_at DESC
            """,
            (current_user["id"],),
        ).fetchall()
        return [dict(row) for row in rows]


@app.get("/wallet/topup-requests/{request_id}/invoice", response_class=HTMLResponse)
def wallet_topup_invoice_page(
    request_id: int,
    token: Optional[str] = None,
    current_user=Depends(get_optional_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if token:
        current_user = _get_user_by_token(token)
    if not current_user:
        raise HTTPException(status_code=401, detail="Missing token")
    with get_conn() as conn:
        request_row = conn.execute(
            "SELECT * FROM wallet_topup_requests WHERE id=? AND user_id=?",
            (request_id, current_user["id"]),
        ).fetchone()
        if not request_row:
            raise HTTPException(status_code=404, detail="Request not found")
        invoice_row = conn.execute(
            "SELECT * FROM invoice_uploads WHERE request_id=? ORDER BY created_at DESC LIMIT 1",
            (request_id,),
        ).fetchone()
        if invoice_row and invoice_row.get("pdf_path"):
            return HTMLResponse(
                content=_wallet_invoice_page_html(
                    dict(request_row),
                    dict(invoice_row),
                    token,
                )
            )
        req = dict(request_row)
        try:
            created_at = req["created_at"]
            dt = created_at if isinstance(created_at, datetime) else datetime.fromisoformat(created_at)
        except ValueError:
            dt = datetime.utcnow()
        number = req.get("invoice_number")
        invoice_date = req.get("invoice_date")
        if not number:
            number = _next_invoice_number(conn)
            invoice_date = datetime.utcnow().date().isoformat()
            conn.execute(
                "UPDATE wallet_topup_requests SET invoice_number=?, invoice_date=? WHERE id=?",
                (number, invoice_date, request_id),
            )
            conn.commit()
        if invoice_date:
            try:
                dt = datetime.fromisoformat(invoice_date)
            except ValueError:
                pass
        date_str = _format_date_ru(dt)
        amount = _format_amount(req.get("amount") or 0)
        currency = req.get("currency") or "KZT"
        amount_words = _amount_to_words_ru(req.get("amount") or 0)
        date_str = f"{date_str} Рі."
        company = _get_company_profile(conn)
        company_name = company.get("name") or BENEFICIARY["name"]
        description = (
            f"За услуги по использованию Программного обеспечения Исполнителя "
            f"\"{company_name}\" по счету {number} от {dt.strftime('%d.%m.%Y')} г."
        )
        beneficiary_bin = company.get("bin") or company.get("iin") or BENEFICIARY["bin"]
        payload = {
            "request_id": request_id,
            "number": number,
            "date": date_str,
            "beneficiary_name": company.get("name") or BENEFICIARY["name"],
            "beneficiary_bin": beneficiary_bin,
            "beneficiary_bank": company.get("bank") or BENEFICIARY["bank"],
            "beneficiary_iban": company.get("iban") or BENEFICIARY["iban"],
            "beneficiary_bic": company.get("bic") or BENEFICIARY["bic"],
            "beneficiary_kbe": company.get("kbe") or BENEFICIARY["kbe"],
            "beneficiary_address": company.get("legal_address") or company.get("factual_address") or "",
            "payment_code": "853",
            "payer_name": req.get("client_name") or "Плательщик не указан",
            "payer_bin": req.get("client_bin") or "ИИН/БИН не указан",
            "payer_address": req.get("client_address") or "Адрес не указан",
            "description": description,
            "contract_note": "Публичный договор возмездного оказания услуг от 22.04.2025 г.",
            "amount": amount,
            "currency": currency,
            "amount_words": amount_words,
            "token": token or "",
        }
        return HTMLResponse(content=_invoice_1c_html(payload))


@app.get("/wallet/topup-requests/{request_id}/pdf")
def wallet_topup_invoice_pdf(
    request_id: int,
    token: Optional[str] = None,
    current_user=Depends(get_optional_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if token:
        current_user = _get_user_by_token(token)
    if not current_user:
        raise HTTPException(status_code=401, detail="Missing token")
    with get_conn() as conn:
        request_row = conn.execute(
            "SELECT id FROM wallet_topup_requests WHERE id=? AND user_id=?",
            (request_id, current_user["id"]),
        ).fetchone()
        if not request_row:
            raise HTTPException(status_code=404, detail="Request not found")
        invoice_row = conn.execute(
            "SELECT * FROM invoice_uploads WHERE request_id=? ORDER BY created_at DESC LIMIT 1",
            (request_id,),
        ).fetchone()
        if not invoice_row or not invoice_row["pdf_path"]:
            raise HTTPException(status_code=404, detail="Invoice PDF not found")
        pdf_path = invoice_row["pdf_path"]
        r2_ref = _r2_parse_path(pdf_path)
        if r2_ref:
            bucket, key = r2_ref
            url = _r2_presigned_url(key, bucket=bucket)
            if not url:
                raise HTTPException(status_code=500, detail="R2 not configured")
            return RedirectResponse(url)
        return FileResponse(pdf_path, media_type="application/pdf")


@app.get("/wallet/topup-requests/{request_id}/pdf-generated")
def wallet_topup_invoice_generated_pdf(
    request_id: int,
    token: Optional[str] = None,
    current_user=Depends(get_optional_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if token:
        current_user = _get_user_by_token(token)
    if not current_user:
        raise HTTPException(status_code=401, detail="Missing token")
    with get_conn() as conn:
        request_row = conn.execute(
            "SELECT * FROM wallet_topup_requests WHERE id=? AND user_id=?",
            (request_id, current_user["id"]),
        ).fetchone()
        if not request_row:
            raise HTTPException(status_code=404, detail="Request not found")
        req = dict(request_row)
        number = req.get("invoice_number")
        invoice_date = req.get("invoice_date")
        if not number:
            number = _next_invoice_number(conn)
            invoice_date = datetime.utcnow().date().isoformat()
            conn.execute(
                "UPDATE wallet_topup_requests SET invoice_number=?, invoice_date=? WHERE id=?",
                (number, invoice_date, request_id),
            )
            conn.commit()
        if invoice_date:
            try:
                dt = datetime.fromisoformat(invoice_date)
            except ValueError:
                dt = datetime.utcnow()
        else:
            dt = datetime.utcnow()
        date_str = _format_date_ru(dt) + " Рі."
        amount = _format_amount(req.get("amount") or 0)
        currency = req.get("currency") or "KZT"
        amount_words = _amount_to_words_ru(req.get("amount") or 0)
        company = _get_company_profile(conn)
        beneficiary_bin = company.get("bin") or company.get("iin") or BENEFICIARY["bin"]
        description = (
            f"За услуги по использованию Программного обеспечения Исполнителя "
            f"\"{company.get('name') or BENEFICIARY['name']}\" по счету {number} от {dt.strftime('%d.%m.%Y')} г."
        )
        payload = {
            "request_id": request_id,
            "number": number,
            "date": date_str,
            "beneficiary_name": company.get("name") or BENEFICIARY["name"],
            "beneficiary_bin": beneficiary_bin,
            "beneficiary_bank": company.get("bank") or BENEFICIARY["bank"],
            "beneficiary_iban": company.get("iban") or BENEFICIARY["iban"],
            "beneficiary_bic": company.get("bic") or BENEFICIARY["bic"],
            "beneficiary_kbe": company.get("kbe") or BENEFICIARY["kbe"],
            "beneficiary_address": company.get("legal_address") or company.get("factual_address") or "",
            "payment_code": "853",
            "payer_name": req.get("client_name") or "Плательщик не указан",
            "payer_bin": req.get("client_bin") or "ИИН/БИН не указан",
            "payer_address": req.get("client_address") or "Адрес не указан",
            "description": description,
            "contract_note": "Публичный договор возмездного оказания услуг от 22.04.2025 г.",
            "amount": amount,
            "currency": currency,
            "amount_words": amount_words,
        }
        html = _invoice_1c_html(payload)
        try:
            from weasyprint import HTML
        except Exception:
            raise HTTPException(status_code=500, detail="PDF renderer is not available")
        if _r2_enabled():
            buffer = BytesIO()
            HTML(string=html, base_url=os.path.dirname(__file__)).write_pdf(buffer)
            key = f"wallet_invoices/wallet_invoice_{request_id}.pdf"
            _r2_upload_bytes(key, buffer.getvalue(), "application/pdf")
            url = _r2_presigned_url(key)
            if not url:
                raise HTTPException(status_code=500, detail="R2 not configured")
            return RedirectResponse(url)
        pdf_path = _wallet_invoice_pdf_path(request_id)
        HTML(string=html, base_url=os.path.dirname(__file__)).write_pdf(pdf_path)
        return FileResponse(pdf_path, media_type="application/pdf")


@app.get("/legal-entities")
def list_legal_entities(current_user=Depends(get_current_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT le.*
            FROM legal_entities le
            JOIN user_legal_entities ule ON ule.legal_entity_id = le.id
            WHERE ule.user_id=?
            ORDER BY le.created_at DESC
            """,
            (current_user["id"],),
        ).fetchall()
        return [dict(row) for row in rows]


@app.post("/legal-entities")
def create_legal_entity(payload: LegalEntityPayload, current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    short_name = (payload.short_name or name).strip()
    full_name = (payload.full_name or name).strip()
    legal_address = (payload.legal_address or payload.address or "").strip() or None
    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO legal_entities (name, short_name, full_name, bin, address, legal_address, email, bank, iban, bic, kbe)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name,
                short_name,
                full_name,
                payload.bin,
                payload.address,
                legal_address,
                payload.email,
                payload.bank,
                payload.iban,
                payload.bic,
                payload.kbe,
            ),
        )
        conn.execute(
            """
            INSERT OR IGNORE INTO user_legal_entities (user_id, legal_entity_id, is_default)
            VALUES (?, ?, ?)
            """,
            (current_user["id"], cur.lastrowid, 0),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM legal_entities WHERE id=?", (cur.lastrowid,)).fetchone()
        return dict(row)


@app.put("/legal-entities/{entity_id}")
def update_legal_entity(entity_id: int, payload: LegalEntityPayload, current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    short_name = (payload.short_name or name).strip()
    full_name = (payload.full_name or name).strip()
    legal_address = (payload.legal_address or payload.address or "").strip() or None
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT le.id FROM legal_entities le
            JOIN user_legal_entities ule ON ule.legal_entity_id = le.id
            WHERE le.id=? AND ule.user_id=?
            """,
            (entity_id, current_user["id"]),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Legal entity not found")
        conn.execute(
            """
            UPDATE legal_entities
            SET name=?, short_name=?, full_name=?, bin=?, address=?, legal_address=?, email=?, bank=?, iban=?, bic=?, kbe=?
            WHERE id=?
            """,
            (
                name,
                short_name,
                full_name,
                payload.bin,
                payload.address,
                legal_address,
                payload.email,
                payload.bank,
                payload.iban,
                payload.bic,
                payload.kbe,
                entity_id,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM legal_entities WHERE id=?", (entity_id,)).fetchone()
        return dict(row)


@app.delete("/legal-entities/{entity_id}")
def delete_legal_entity(entity_id: int, current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        link = conn.execute(
            """
            SELECT 1 FROM user_legal_entities
            WHERE user_id=? AND legal_entity_id=?
            """,
            (current_user["id"], entity_id),
        ).fetchone()
        if not link:
            raise HTTPException(status_code=404, detail="Legal entity not found")
        conn.execute(
            "DELETE FROM user_legal_entities WHERE user_id=? AND legal_entity_id=?",
            (current_user["id"], entity_id),
        )
        other_links = conn.execute(
            "SELECT 1 FROM user_legal_entities WHERE legal_entity_id=? LIMIT 1",
            (entity_id,),
        ).fetchone()
        if not other_links:
            conn.execute("DELETE FROM legal_entities WHERE id=?", (entity_id,))
        conn.commit()
        return {"status": "ok"}


@app.post("/api/invoices")
def upload_invoice(
    pdf: Optional[UploadFile] = File(None),
    file: Optional[UploadFile] = File(None),
    invoice_id: Optional[int] = Form(None),
    request_id: Optional[int] = Form(None),
    invoice_number: Optional[str] = Form(None),
    invoice_date: Optional[str] = Form(None),
    amount: Optional[float] = Form(None),
    currency: Optional[str] = Form(None),
    client_name: Optional[str] = Form(None),
    client_bin: Optional[str] = Form(None),
    client_address: Optional[str] = Form(None),
    order_ref: Optional[str] = Form(None),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    upload = pdf or file
    if not upload:
        raise HTTPException(status_code=400, detail="PDF file is required")
    pdf_path = _save_invoice_pdf(upload)
    with get_conn() as conn:
        if invoice_id:
            row = conn.execute("SELECT * FROM invoice_uploads WHERE id=?", (invoice_id,)).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Invoice not found")
            existing = dict(row)
            effective_request_id = request_id or existing.get("request_id")
            conn.execute(
                """
                UPDATE invoice_uploads
                SET request_id=?, invoice_number=?, invoice_date=?, amount=?, currency=?,
                    client_name=?, client_bin=?, client_address=?, order_ref=?,
                    pdf_path=?, status=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
                """,
                (
                    effective_request_id,
                    invoice_number or existing.get("invoice_number"),
                    invoice_date or existing.get("invoice_date"),
                    amount if amount is not None else existing.get("amount"),
                    currency or existing.get("currency"),
                    client_name or existing.get("client_name"),
                    client_bin or existing.get("client_bin"),
                    client_address or existing.get("client_address"),
                    order_ref or existing.get("order_ref"),
                    pdf_path,
                    "ready",
                    invoice_id,
                ),
            )
            if effective_request_id:
                conn.execute(
                    "UPDATE wallet_topup_requests SET status=? WHERE id=?",
                    ("invoice_ready", effective_request_id),
                )
            conn.commit()
            return {"id": invoice_id, "request_id": effective_request_id, "status": "ok"}
        cur = conn.execute(
            """
            INSERT INTO invoice_uploads
            (request_id, invoice_number, invoice_date, amount, currency, client_name, client_bin, client_address, order_ref, pdf_path, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                request_id,
                invoice_number,
                invoice_date,
                amount,
                currency,
                client_name,
                client_bin,
                client_address,
                order_ref,
                pdf_path,
                "ready",
            ),
        )
        if request_id:
            conn.execute(
                "UPDATE wallet_topup_requests SET status=? WHERE id=?",
                ("invoice_ready", request_id),
            )
        conn.commit()
        return {"id": cur.lastrowid, "request_id": request_id, "status": "ok"}


@app.post("/api/invoices/pending")
def invoice_pending(payload: InvoicePendingPayload):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO invoice_uploads
            (request_id, invoice_number, invoice_date, amount, currency, client_name, client_bin, client_address, order_ref, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload.request_id,
                payload.invoice_number,
                payload.invoice_date,
                payload.amount,
                payload.currency,
                payload.client_name,
                payload.client_bin,
                payload.client_address,
                payload.order_ref,
                "pending",
            ),
        )
        if payload.request_id:
            conn.execute(
                "UPDATE wallet_topup_requests SET status=? WHERE id=?",
                ("invoice_pending", payload.request_id),
            )
        conn.commit()
        return {"status": "ok", "invoice_id": cur.lastrowid, "request_id": payload.request_id}


@app.get("/me")
def me(current_user=Depends(get_current_user)):
    return {"id": current_user["id"], "email": current_user["email"]}


@app.get("/account-requests")
def list_account_requests(current_user=Depends(get_current_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT r.*,
                   a.id as account_id,
                   a.account_code as account_code_db,
                   a.budget_total as budget_total,
                   a.currency as account_currency,
                   COALESCE((SELECT SUM(t.amount_input)
                             FROM topups t
                             WHERE t.account_id = a.id AND t.status='completed'), 0) as topup_completed_total
            FROM account_requests r
            LEFT JOIN ad_accounts a ON a.user_id = r.user_id AND a.platform = r.platform AND a.name = r.name
            WHERE r.user_id=?
            ORDER BY r.created_at DESC
            """,
            (current_user["id"],),
        ).fetchall()
        return [dict(row) for row in rows]


@app.post("/account-requests")
def create_account_request(payload: AccountRequestCreate, current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO account_requests (user_id, platform, name, payload, status) VALUES (?, ?, ?, json(?), ?)",
            (current_user["id"], payload.platform, payload.name, json.dumps(payload.payload, ensure_ascii=False), "new"),
        )
        conn.commit()
        request_id = cur.lastrowid
        _send_telegram_alert(
            "\n".join(
                [
                    "🆕 <b>Заявка на открытие аккаунта</b>",
                    f"ID: <code>{request_id}</code>",
                    f"Пользователь: <code>{current_user['email']}</code> (id={current_user['id']})",
                    f"Платформа: <b>{payload.platform}</b>",
                    f"Название: <b>{payload.name}</b>",
                ]
            )
        )
        return {
            "id": request_id,
            "user_id": current_user["id"],
            "platform": payload.platform,
            "name": payload.name,
            "status": "new",
        }


@app.get("/admin/account-requests")
def admin_list_account_requests(admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT r.*,
                   u.email as user_email,
                   a.id as account_id,
                   a.account_code as account_code_db,
                   a.budget_total as budget_total,
                   a.currency as account_currency,
                   COALESCE((SELECT SUM(t.amount_input)
                             FROM topups t
                             WHERE t.account_id = a.id AND t.status='completed'), 0) as topup_completed_total
            FROM account_requests r
            JOIN users u ON u.id = r.user_id
            LEFT JOIN ad_accounts a ON a.user_id = r.user_id AND a.platform = r.platform AND a.name = r.name
            ORDER BY r.created_at DESC
            """
        ).fetchall()
        return [dict(row) for row in rows]


@app.get("/admin/accounts")
def admin_list_accounts(admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT a.*, u.email as user_email
            FROM ad_accounts a
            JOIN users u ON u.id = a.user_id
            ORDER BY a.created_at DESC
            """
        ).fetchall()
        return _attach_live_billing_many([dict(row) for row in rows])


@app.post("/admin/accounts")
def admin_create_account(payload: AdminAccountCreate, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        user = conn.execute("SELECT id FROM users WHERE id=?", (payload.user_id,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        cur = conn.execute(
            """
            INSERT INTO ad_accounts (user_id, platform, name, external_id, account_code, currency, status)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload.user_id,
                payload.platform,
                payload.name,
                payload.external_id,
                payload.account_code,
                payload.currency,
                payload.status or "pending",
            ),
        )
        conn.commit()
        return {
            "id": cur.lastrowid,
            "user_id": payload.user_id,
            "platform": payload.platform,
            "name": payload.name,
            "external_id": payload.external_id,
            "account_code": payload.account_code,
            "currency": payload.currency,
            "status": payload.status or "pending",
        }


@app.patch("/admin/accounts/{account_id}")
def admin_update_account(account_id: int, payload: AdminAccountUpdate, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM ad_accounts WHERE id=?", (account_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Account not found")
        updates = []
        params: List[object] = []
        if payload.user_id is not None:
            user = conn.execute("SELECT id FROM users WHERE id=?", (payload.user_id,)).fetchone()
            if not user:
                raise HTTPException(status_code=404, detail="User not found")
            updates.append("user_id=?")
            params.append(payload.user_id)
        if payload.platform is not None:
            if payload.platform not in {"meta", "google", "tiktok", "yandex", "telegram", "monochrome"}:
                raise HTTPException(status_code=400, detail="Unsupported platform")
            updates.append("platform=?")
            params.append(payload.platform)
        if payload.name is not None:
            updates.append("name=?")
            params.append(payload.name)
        if payload.external_id is not None:
            updates.append("external_id=?")
            params.append(payload.external_id)
        if payload.account_code is not None:
            updates.append("account_code=?")
            params.append(payload.account_code)
        if payload.currency is not None:
            updates.append("currency=?")
            params.append(payload.currency)
        if payload.status is not None:
            updates.append("status=?")
            params.append(payload.status)
        if not updates:
            raise HTTPException(status_code=400, detail="No fields to update")
        params.append(account_id)
        conn.execute(f"UPDATE ad_accounts SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
        return {"id": account_id, "status": "updated"}


@app.get("/admin/topups")
def admin_list_topups(admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT t.*, a.name as account_name, a.platform as account_platform, a.currency as account_currency, u.email as user_email
            FROM topups t
            JOIN ad_accounts a ON a.id = t.account_id
            JOIN users u ON u.id = t.user_id
            ORDER BY t.created_at DESC
            """
        ).fetchall()
        return _attach_topup_account_amount([dict(row) for row in rows])


@app.get("/admin/clients")
def admin_list_clients(admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        clients: List[Dict[str, object]] = []
        try:
            try:
                rows = conn.execute(
                    """
                    SELECT
                      u.id,
                      u.email,
                      COALESCE(ts.unread_topups, 0) as unread_topups,
                      COALESCE(ts.pending_requests, 0) as pending_requests,
                      COALESCE(ts.completed_total_kzt, 0) as completed_total,
                      COALESCE(ts.completed_count, 0) as completed_count,
                      COALESCE(ts.last_topup_at, ws.last_funding_at) as last_activity
                    FROM users u
                    LEFT JOIN (
                      SELECT
                        user_id,
                        COALESCE(SUM(CASE WHEN seen_by_admin=0 THEN 1 ELSE 0 END), 0) as unread_topups,
                        COALESCE(SUM(CASE WHEN status!='completed' THEN 1 ELSE 0 END), 0) as pending_requests,
                        COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END), 0) as completed_count,
                        COALESCE(SUM(CASE WHEN status='completed' THEN COALESCE(amount_input, 0) ELSE 0 END), 0) as completed_total_kzt,
                        MAX(created_at) as last_topup_at
                      FROM topups
                      GROUP BY user_id
                    ) ts ON ts.user_id = u.id
                    LEFT JOIN (
                      SELECT
                        user_id,
                        COALESCE(SUM(CASE WHEN type='adjustment' AND amount > 0 THEN amount ELSE 0 END), 0) as completed_total,
                        MAX(created_at) as last_funding_at
                      FROM wallet_transactions
                      GROUP BY user_id
                    ) ws ON ws.user_id = u.id
                    WHERE COALESCE(ts.completed_count, 0) > 0 OR COALESCE(u.is_client, 0) = 1
                    ORDER BY unread_topups DESC, u.email ASC
                    """
                ).fetchall()
            except Exception:
                rows = conn.execute(
                    """
                    SELECT
                      u.id,
                      u.email,
                      0 as unread_topups,
                      COALESCE(ts.pending_requests, 0) as pending_requests,
                      COALESCE(ts.completed_total_kzt, 0) as completed_total,
                      COALESCE(ts.completed_count, 0) as completed_count,
                      COALESCE(ts.last_topup_at, ws.last_funding_at) as last_activity
                    FROM users u
                    LEFT JOIN (
                      SELECT
                        user_id,
                        COALESCE(SUM(CASE WHEN status!='completed' THEN 1 ELSE 0 END), 0) as pending_requests,
                        COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END), 0) as completed_count,
                        COALESCE(SUM(CASE WHEN status='completed' THEN COALESCE(amount_input, 0) ELSE 0 END), 0) as completed_total_kzt,
                        MAX(created_at) as last_topup_at
                      FROM topups
                      GROUP BY user_id
                    ) ts ON ts.user_id = u.id
                    LEFT JOIN (
                      SELECT
                        user_id,
                        COALESCE(SUM(CASE WHEN type='adjustment' AND amount > 0 THEN amount ELSE 0 END), 0) as completed_total,
                        MAX(created_at) as last_funding_at
                      FROM wallet_transactions
                      GROUP BY user_id
                    ) ws ON ws.user_id = u.id
                    WHERE COALESCE(ts.completed_count, 0) > 0 OR COALESCE(u.is_client, 0) = 1
                    ORDER BY u.email ASC
                    """
                ).fetchall()
            clients = [dict(row) for row in rows]
            for row in clients:
                row["completed_total_kzt"] = float(row.get("completed_total") or 0.0)
        except Exception:
            fallback_rows = conn.execute(
                """
                SELECT u.id, u.email, 0 as unread_topups,
                  COALESCE(SUM(CASE WHEN t.status!='completed' THEN 1 ELSE 0 END), 0) as pending_requests,
                  COALESCE(SUM(CASE WHEN t.status='completed' THEN 1 ELSE 0 END), 0) as completed_count,
                  MAX(t.created_at) as last_activity
                FROM users u
                LEFT JOIN topups t ON t.user_id = u.id
                GROUP BY u.id, u.email
                HAVING COALESCE(SUM(CASE WHEN t.status='completed' THEN 1 ELSE 0 END), 0) > 0
                ORDER BY u.email ASC
                """
            ).fetchall()
            clients = []
            for row in fallback_rows:
                payload = dict(row)
                payload["completed_total"] = 0.0
                payload["completed_total_kzt"] = 0.0
                clients.append(payload)
        return clients


@app.get("/admin/users")
def admin_list_users(admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        admin_emails = sorted(ADMIN_EMAILS)
        email_filter = ""
        params: List[object] = []
        if admin_emails:
            placeholders = ", ".join(["?"] * len(admin_emails))
            email_filter = f"WHERE u.email NOT IN ({placeholders})"
            params.extend(admin_emails)
        rows = conn.execute(
            """
            SELECT u.id, u.email, u.created_at,
              COALESCE(SUM(CASE WHEN t.status='completed' THEN 1 ELSE 0 END), 0) as completed_count
            FROM users u
            LEFT JOIN topups t ON t.user_id = u.id
            {email_filter}
            GROUP BY u.id, u.email, u.created_at
            HAVING COALESCE(SUM(CASE WHEN t.status='completed' THEN 1 ELSE 0 END), 0) = 0
               AND COALESCE(u.is_client, 0) = 0
            ORDER BY u.created_at DESC
            """.format(email_filter=email_filter),
            params,
        ).fetchall()
        return [dict(row) for row in rows]


@app.get("/admin/users/{user_id}/fees")
def admin_get_user_fees(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        return _default_fee_config()
    with get_conn() as conn:
        user = conn.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        profile = _get_or_create_profile(conn, user_id)
        return _load_fee_config(profile.get("fee_config"))


@app.put("/admin/users/{user_id}/fees")
def admin_update_user_fees(user_id: int, payload: FeeConfigPayload, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    updates = payload.model_dump()
    for key, value in updates.items():
        if value is None:
            continue
        if value < 0 or value > 100:
            raise HTTPException(status_code=400, detail="Fee percent must be between 0 and 100")
    with get_conn() as conn:
        user = conn.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        profile = _get_or_create_profile(conn, user_id)
        current = _load_fee_config(profile.get("fee_config"))
        for key, value in updates.items():
            if key in current:
                current[key] = value
        conn.execute(
            "UPDATE user_profiles SET fee_config=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
            (json.dumps(current, ensure_ascii=False), user_id),
        )
        conn.commit()
        return current


@app.post("/admin/users/{user_id}/make-client")
def admin_make_user_client(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="User not found")
        conn.execute("UPDATE users SET is_client=1 WHERE id=?", (user_id,))
        conn.commit()
        return {"id": user_id, "status": "client"}


@app.get("/admin/clients/{user_id}/allocations")
def admin_client_allocations(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT t.*, a.name as account_name, a.platform as account_platform
            FROM topups t
            JOIN ad_accounts a ON a.id = t.account_id
            WHERE t.user_id=?
            ORDER BY t.created_at DESC
            """,
            (user_id,),
        ).fetchall()
        return [dict(row) for row in rows]


@app.get("/admin/clients/{user_id}/requests")
def admin_client_requests(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT t.*, a.name as account_name, a.platform as account_platform, a.currency as account_currency
            FROM topups t
            JOIN ad_accounts a ON a.id = t.account_id
            WHERE t.user_id=? AND t.status!='completed'
            ORDER BY t.created_at DESC
            """,
            (user_id,),
        ).fetchall()
        return _attach_topup_account_amount([dict(row) for row in rows])


@app.get("/admin/clients/{user_id}/topups")
def admin_client_topups(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT t.*, a.name as account_name, a.platform as account_platform, a.currency as account_currency
            FROM topups t
            LEFT JOIN ad_accounts a ON a.id = t.account_id
            WHERE t.user_id=? AND t.status='completed'
            ORDER BY t.created_at DESC
            """,
            (user_id,),
        ).fetchall()
        return _attach_topup_account_amount([dict(row) for row in rows])


@app.get("/admin/clients/{user_id}/wallet-transactions")
def admin_client_wallet_transactions(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT wt.*, a.name as account_name, a.platform as account_platform
            FROM wallet_transactions wt
            LEFT JOIN ad_accounts a ON a.id = wt.account_id
            WHERE wt.user_id=?
            ORDER BY wt.created_at DESC
            """,
            (user_id,),
        ).fetchall()
        result = []
        try:
            rates_data = _fetch_bcc_rates()
        except Exception:
            rates_data = None
        for row in rows:
            payload = dict(row)
            payload["amount_usd"] = _convert_amount_to_usd(payload.get("amount"), payload.get("currency"), rates_data)
            result.append(payload)
        return result


@app.get("/admin/clients/{user_id}/invoice-summary")
def admin_client_invoice_summary(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        return {"invoice_total_kzt": 0.0, "invoice_count": 0}
    with get_conn() as conn:
        user = conn.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        rows = conn.execute(
            """
            SELECT
              i.id as invoice_id,
              i.amount as invoice_amount,
              COALESCE(i.currency, 'KZT') as invoice_currency
            FROM wallet_topup_requests r
            JOIN invoice_uploads i ON i.id = (
              SELECT id
              FROM invoice_uploads
              WHERE request_id = r.id AND COALESCE(status, 'pending') = 'ready'
              ORDER BY created_at DESC
              LIMIT 1
            )
            WHERE r.user_id=? AND COALESCE(r.status, 'requested') = 'invoice_ready'
            ORDER BY r.created_at DESC
            """,
            (user_id,),
        ).fetchall()

        def _to_float(value: object) -> float:
            if value is None:
                return 0.0
            try:
                return float(value)
            except (TypeError, ValueError):
                text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
                try:
                    return float(text)
                except (TypeError, ValueError):
                    return 0.0

        total_kzt = 0.0
        count = 0
        for row in rows:
            amount = _to_float(row.get("invoice_amount"))
            if amount <= 0:
                continue
            # Wallet invoices are treated as KZT source of truth for admin summary.
            total_kzt += float(amount)
            count += 1
        return {"invoice_total_kzt": round(total_kzt, 2), "invoice_count": count}


@app.get("/admin/clients/{user_id}/accounts")
def admin_client_accounts(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT a.*
            FROM ad_accounts a
            WHERE a.user_id=?
            ORDER BY a.created_at DESC
            """,
            (user_id,),
        ).fetchall()
        return _attach_live_billing_many([dict(row) for row in rows])


@app.get("/admin/clients/{user_id}/profile")
def admin_client_profile(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT u.id, u.email, p.name, p.company, p.language, p.whatsapp_phone, p.telegram_handle
            FROM users u
            LEFT JOIN user_profiles p ON p.user_id = u.id
            WHERE u.id=?
            """,
            (user_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="User not found")
        return dict(row)


@app.post("/admin/clients/{user_id}/mark-seen")
def admin_mark_client_seen(user_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        conn.execute("UPDATE topups SET seen_by_admin=1 WHERE user_id=?", (user_id,))
        conn.commit()
        return {"status": "ok"}


@app.get("/admin/export/requests.xlsx")
def admin_export_requests(admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Requests"
    ws.append(["Дата", "Клиент", "Платформа", "Название", "Статус", "Менеджер"])
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT r.*, u.email as user_email
            FROM account_requests r
            JOIN users u ON u.id = r.user_id
            ORDER BY r.created_at DESC
            """
        ).fetchall()
        for row in rows:
            ws.append(
                [
                    row["created_at"],
                    row["user_email"],
                    row["platform"],
                    row["name"],
                    row["status"],
                    row["manager_email"] or "",
                ]
            )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="account_requests.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/admin/export/accounts.xlsx")
def admin_export_accounts(admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"
    ws.append(["Дата", "Клиент", "Платформа", "Название", "Договор/код", "External ID"])
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT a.*, u.email as user_email
            FROM ad_accounts a
            JOIN users u ON u.id = a.user_id
            ORDER BY a.created_at DESC
            """
        ).fetchall()
        for row in rows:
            ws.append(
                [
                    row["created_at"],
                    row["user_email"],
                    row["platform"],
                    row["name"],
                    row["account_code"] or "",
                    row["external_id"] or "",
                ]
            )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="accounts.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/admin/export/topups.xlsx")
def admin_export_topups(admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    wb = Workbook()
    ws = wb.active
    ws.title = "Topups"
    ws.append(
        [
            "Дата",
            "Клиент",
            "Платформа",
            "Аккаунт",
            "Сумма",
            "Комиссия",
            "НДС",
            "К оплате",
            "Валюта",
            "Статус",
        ]
    )
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT t.*, a.name as account_name, a.platform as account_platform, u.email as user_email
            FROM topups t
            JOIN ad_accounts a ON a.id = t.account_id
            JOIN users u ON u.id = t.user_id
            ORDER BY t.created_at DESC
            """
        ).fetchall()
        for row in rows:
            fee = (row["amount_input"] or 0) * (row["fee_percent"] or 0) / 100.0
            vat = (row["amount_input"] or 0) * (row["vat_percent"] or 0) / 100.0
            gross = (row["amount_input"] or 0) + fee + vat
            ws.append(
                [
                    row["created_at"],
                    row["user_email"],
                    row["account_platform"],
                    row["account_name"],
                    row["amount_input"],
                    round(fee, 2),
                    round(vat, 2),
                    round(gross, 2),
                    row["currency"],
                    row["status"],
                ]
            )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="topups.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/admin/topups/{topup_id}/status")
def admin_update_topup_status(topup_id: int, status: TopUpStatus, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, status, account_id, amount_input, amount_net, fee_percent, vat_percent, currency, user_id, hold_applied FROM topups WHERE id=?",
            (topup_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Topup not found")

        previous_status = row["status"]
        next_status = status.value
        hold_applied = int(row["hold_applied"] or 0) == 1
        gross_amount = _topup_gross_amount(
            float(row["amount_input"] or 0),
            float(row["fee_percent"] or 0),
            float(row["vat_percent"] or 0),
        )

        if previous_status != "completed" and next_status == "completed":
            if hold_applied:
                conn.execute("UPDATE topups SET status=?, hold_applied=0 WHERE id=?", (next_status, topup_id))
            else:
                wallet = _get_or_create_wallet(conn, row["user_id"])
                if float(wallet["balance"]) < gross_amount:
                    raise HTTPException(status_code=400, detail="Недостаточно средств на кошельке для завершения пополнения")
                new_balance = float(wallet["balance"]) - gross_amount
                conn.execute(
                    "UPDATE wallets SET balance=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
                    (new_balance, row["user_id"]),
                )
                conn.execute(
                    """
                    INSERT INTO wallet_transactions (user_id, account_id, amount, currency, type, note)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (row["user_id"], row["account_id"], -gross_amount, row["currency"], "topup", "Account topup"),
                )
                conn.execute("UPDATE topups SET status=? WHERE id=?", (next_status, topup_id))

            acc = conn.execute("SELECT budget_total FROM ad_accounts WHERE id=?", (row["account_id"],)).fetchone()
            base_amount = row["amount_net"] if row["amount_net"] else row["amount_input"]
            new_total = (acc["budget_total"] or 0) + (base_amount or 0)
            conn.execute(
                "UPDATE ad_accounts SET budget_total=? WHERE id=?",
                (new_total, row["account_id"]),
            )
            conn.execute("UPDATE users SET is_client=1 WHERE id=?", (row["user_id"],))
            user_row = conn.execute("SELECT email FROM users WHERE id=?", (row["user_id"],)).fetchone()
            account_row = conn.execute("SELECT platform, name FROM ad_accounts WHERE id=?", (row["account_id"],)).fetchone()
            _send_telegram_alert(
                "\n".join(
                    [
                        "✅ <b>Пополнение оплачено/завершено</b>",
                        f"Topup ID: <code>{topup_id}</code>",
                        f"Пользователь: <code>{user_row['email'] if user_row else row['user_id']}</code>",
                        f"Платформа: <b>{account_row['platform'] if account_row else '—'}</b>",
                        f"Аккаунт: <b>{account_row['name'] if account_row else row['account_id']}</b>",
                        f"Сумма: <b>{(row['amount_net'] or row['amount_input'] or 0):.2f} {row['currency']}</b>",
                    ]
                )
            )
        elif previous_status == "pending" and next_status == "failed":
            if hold_applied:
                wallet = _get_or_create_wallet(conn, row["user_id"])
                new_balance = float(wallet["balance"]) + gross_amount
                conn.execute(
                    "UPDATE wallets SET balance=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
                    (new_balance, row["user_id"]),
                )
                conn.execute(
                    """
                    INSERT INTO wallet_transactions (user_id, account_id, amount, currency, type, note)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (row["user_id"], row["account_id"], gross_amount, row["currency"], "topup_hold_release", f"Topup hold released #{topup_id}"),
                )
                conn.execute("UPDATE topups SET status=?, hold_applied=0 WHERE id=?", (next_status, topup_id))
            else:
                conn.execute("UPDATE topups SET status=? WHERE id=?", (next_status, topup_id))
        else:
            conn.execute("UPDATE topups SET status=? WHERE id=?", (next_status, topup_id))
        conn.commit()
        return {"id": topup_id, "status": next_status}


@app.get("/admin/topups/profit-summary")
def admin_topups_profit_summary(admin_user=Depends(get_admin_user)):
    if not get_conn:
        return {"overall": {}, "by_platform": []}
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
              a.platform as platform,
              t.currency as currency,
              COUNT(*) as completed_count,
              COALESCE(SUM(t.amount_input), 0) as amount_input_total,
              COALESCE(SUM((t.amount_input * t.fee_percent) / 100.0), 0) as fee_total
            FROM topups t
            JOIN ad_accounts a ON a.id = t.account_id
            WHERE t.status='completed'
            GROUP BY a.platform, t.currency
            ORDER BY fee_total DESC
            """
        ).fetchall()
        by_platform = [dict(row) for row in rows]

        overall_rows = conn.execute(
            """
            SELECT
              t.currency as currency,
              COUNT(*) as completed_count,
              COALESCE(SUM(t.amount_input), 0) as amount_input_total,
              COALESCE(SUM((t.amount_input * t.fee_percent) / 100.0), 0) as fee_total
            FROM topups t
            WHERE t.status='completed'
            GROUP BY t.currency
            ORDER BY fee_total DESC
            """
        ).fetchall()
        overall = [dict(row) for row in overall_rows]
        return {"overall": overall, "by_platform": by_platform}


@app.patch("/admin/topups/{topup_id}")
def admin_update_topup(topup_id: int, payload: AdminTopupUpdate, admin_user=Depends(get_admin_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if payload.amount_net is None and payload.fx_rate is None:
        raise HTTPException(status_code=400, detail="No fields to update")
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM topups WHERE id=?", (topup_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Topup not found")
        updates = []
        params: List[object] = []
        if payload.amount_net is not None:
            updates.append("amount_net=?")
            params.append(payload.amount_net)
        if payload.fx_rate is not None:
            updates.append("fx_rate=?")
            params.append(payload.fx_rate)
        params.append(topup_id)
        conn.execute(f"UPDATE topups SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
        return {"id": topup_id, "status": "updated"}


@app.post("/admin/account-requests/{request_id}/status")
def admin_update_account_request_status(
    request_id: int, payload: AccountRequestUpdate, admin_user=Depends(get_admin_user)
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM account_requests WHERE id=?", (request_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Request not found")
        if payload.manager_email:
            conn.execute(
                "UPDATE account_requests SET manager_email=? WHERE id=?",
                (payload.manager_email, request_id),
            )
        if payload.comment is not None:
            conn.execute(
                "UPDATE account_requests SET comment=? WHERE id=?",
                (payload.comment, request_id),
            )
        if payload.account_code is not None:
            conn.execute(
                "UPDATE account_requests SET account_code=? WHERE id=?",
                (payload.account_code, request_id),
            )
        default_currency = "EUR" if row["platform"] == "telegram" else "USD"
        if payload.budget_total is not None:
            existing_acc = conn.execute(
                "SELECT id, currency FROM ad_accounts WHERE user_id=? AND platform=? AND name=?",
                (row["user_id"], row["platform"], row["name"]),
            ).fetchone()
            if existing_acc:
                conn.execute(
                    "UPDATE ad_accounts SET budget_total=? WHERE id=?",
                    (payload.budget_total, existing_acc["id"]),
                )
            elif payload.status == "approved":
                conn.execute(
                    "INSERT INTO ad_accounts (user_id, platform, name, external_id, currency, account_code, budget_total) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (row["user_id"], row["platform"], row["name"], None, default_currency, payload.account_code, payload.budget_total),
                )
        if payload.status == "approved":
            existing = conn.execute(
                "SELECT id FROM ad_accounts WHERE user_id=? AND platform=? AND name=?",
                (row["user_id"], row["platform"], row["name"]),
            ).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO ad_accounts (user_id, platform, name, external_id, currency, account_code, budget_total) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (row["user_id"], row["platform"], row["name"], None, default_currency, payload.account_code, payload.budget_total),
                )
            elif payload.account_code:
                conn.execute(
                    "UPDATE ad_accounts SET account_code=? WHERE id=?",
                    (payload.account_code, existing["id"]),
                )
        conn.execute("UPDATE account_requests SET status=? WHERE id=?", (payload.status, request_id))
        _insert_request_event(
            conn,
            request_id=request_id,
            admin_email=admin_user["email"],
            event_type="status",
            status=payload.status,
            comment=payload.comment,
            manager_email=payload.manager_email,
        )
        conn.commit()
        return {"id": request_id, "status": payload.status}


@app.get("/admin/account-requests/{request_id}/events")
def admin_list_account_request_events(request_id: int, admin_user=Depends(get_admin_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT * FROM account_request_events
            WHERE request_id=?
            ORDER BY created_at DESC
            """,
            (request_id,),
        ).fetchall()
        return [dict(row) for row in rows]


@app.post("/admin/account-requests/{request_id}/events")
def admin_create_account_request_event(
    request_id: int, payload: AccountRequestEventCreate, admin_user=Depends(get_admin_user)
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM account_requests WHERE id=?", (request_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Request not found")
        if payload.manager_email:
            conn.execute(
                "UPDATE account_requests SET manager_email=? WHERE id=?",
                (payload.manager_email, request_id),
            )
        if payload.comment is not None:
            conn.execute(
                "UPDATE account_requests SET comment=? WHERE id=?",
                (payload.comment, request_id),
            )
        _insert_request_event(
            conn,
            request_id=request_id,
            admin_email=admin_user["email"],
            event_type=payload.type,
            status=payload.status,
            comment=payload.comment,
            manager_email=payload.manager_email,
        )
        conn.commit()
        return {"status": "ok"}


def _insert_request_event(
    conn,
    request_id: int,
    admin_email: str,
    event_type: str,
    status: Optional[str],
    comment: Optional[str],
    manager_email: Optional[str],
) -> None:
    conn.execute(
        """
        INSERT INTO account_request_events (request_id, admin_email, manager_email, type, status, comment)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (request_id, admin_email, manager_email, event_type, status, comment),
    )

@app.get("/invoices/preview", response_class=HTMLResponse)
def invoice_preview():
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        company = _get_company_profile(conn)
    beneficiary_bin = company.get("bin") or company.get("iin") or BENEFICIARY["bin"]
    payload = {
        "number": "INV-20240101-0001",
        "date": "2024-01-01",
        "beneficiary_name": company.get("name") or BENEFICIARY["name"],
        "beneficiary_bin": beneficiary_bin,
        "beneficiary_bank": company.get("bank") or BENEFICIARY["bank"],
        "beneficiary_iban": company.get("iban") or BENEFICIARY["iban"],
        "beneficiary_bic": company.get("bic") or BENEFICIARY["bic"],
        "beneficiary_kbe": company.get("kbe") or BENEFICIARY["kbe"],
        "payer_name": "ООО Клиент",
        "payer_bin": "ИИН/БИН не указан",
        "payer_address": "Адрес не указан",
        "description": "Пополнение рекламного аккаунта",
        "amount": _format_amount(150000),
        "currency": company.get("currency") or BENEFICIARY["currency"],
        "items": [
            {
                "description": "Пополнение рекламного аккаунта",
                "qty": "1",
                "unit": "усл.",
                "price": _format_amount(150000),
                "amount": _format_amount(150000),
            }
        ],
        "items_count": 1,
    }
    return HTMLResponse(content=_invoice_html(payload))


@app.get("/accounts")
def list_accounts(current_user=Depends(get_current_user)):
    if not get_conn:
        return []
    with get_conn() as conn:
        query = "SELECT * FROM ad_accounts WHERE user_id=?"
        params: List[object] = [current_user["id"]]
        query += " ORDER BY created_at DESC"
        rows = conn.execute(query, params).fetchall()
        return _attach_live_billing_many([dict(row) for row in rows])


@app.get("/accounts/spend")
def list_accounts_period_spend(
    date_from: str,
    date_to: str,
    current_user=Depends(get_current_user),
):
    if not get_conn:
        return {"date_from": date_from, "date_to": date_to, "items": []}
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="date_from and date_to are required")
    try:
        from_dt = datetime.strptime(date_from, "%Y-%m-%d").date()
        to_dt = datetime.strptime(date_to, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="date_from/date_to must be in YYYY-MM-DD format")
    if from_dt > to_dt:
        raise HTTPException(status_code=400, detail="date_from must be less than or equal to date_to")

    def _to_float(value: object) -> float:
        try:
            return float(value)
        except Exception:
            return 0.0

    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, platform, external_id, account_code, name, currency FROM ad_accounts WHERE user_id=? ORDER BY created_at DESC",
            (current_user["id"],),
        ).fetchall()
        accounts = [dict(row) for row in rows]

    items: List[Dict[str, object]] = []
    for acc in accounts:
        account_id = acc.get("id")
        platform = str(acc.get("platform") or "").lower().strip()
        external_id = acc.get("external_id") or acc.get("account_code") or acc.get("name")
        currency = acc.get("currency") or "USD"
        payload: Dict[str, object] = {
            "account_id": account_id,
            "platform": platform,
            "currency": currency,
            "spend": None,
        }

        if platform not in {"meta", "google", "tiktok"}:
            items.append(payload)
            continue
        if not external_id:
            items.append(payload)
            continue

        try:
            if platform == "meta":
                daily_rows = _meta_fetch_daily(str(external_id), date_from, date_to)
                payload["spend"] = sum(_to_float(row.get("spend")) for row in daily_rows)
            elif platform == "google":
                daily_rows = _google_fetch_daily(_google_normalize_customer_id(str(external_id)), date_from, date_to)
                payload["spend"] = sum(_to_float(row.get("spend")) for row in daily_rows)
            elif platform == "tiktok":
                daily_rows = _tiktok_fetch_daily(_tiktok_normalize_advertiser_id(str(external_id)), date_from, date_to)
                payload["spend"] = sum(_to_float(row.get("spend")) for row in daily_rows)
        except Exception as exc:
            payload["error"] = str(exc)

        items.append(payload)

    return {"date_from": date_from, "date_to": date_to, "items": items}


@app.post("/accounts/{account_id}/refresh-live-billing")
def refresh_account_live_billing(account_id: int, current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ad_accounts WHERE id=? AND user_id=?",
            (account_id, current_user["id"]),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Account not found")
        payload = _attach_live_billing(dict(row), force_refresh=True)
        return {
            "id": payload.get("id"),
            "platform": payload.get("platform"),
            "live_billing": payload.get("live_billing"),
            "updated_at": datetime.utcnow().isoformat() + "Z",
        }


@app.post("/accounts")
def create_account(
    platform: str,
    name: str,
    external_id: Optional[str] = None,
    currency: str = "USD",
    current_user=Depends(get_current_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if platform not in {"meta", "google", "tiktok", "yandex", "telegram", "monochrome"}:
        raise HTTPException(status_code=400, detail="Unsupported platform")
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO ad_accounts (user_id, platform, name, external_id, currency) VALUES (?, ?, ?, ?, ?)",
            (current_user["id"], platform, name, external_id, currency),
        )
        conn.commit()
        return {
            "id": cur.lastrowid,
            "user_id": current_user["id"],
            "platform": platform,
            "name": name,
            "external_id": external_id,
            "currency": currency,
        }


@app.get("/topups")
def list_topups(account_id: Optional[int] = None, status: Optional[str] = None, current_user=Depends(get_current_user)):
    if not get_conn:
        return []
    try:
        with get_conn() as conn:
            query = "SELECT t.*, a.name as account_name, a.platform as account_platform, a.currency as account_currency FROM topups t JOIN ad_accounts a ON a.id=t.account_id WHERE 1=1"
            params: List[object] = []
            if account_id:
                query += " AND t.account_id=?"
                params.append(account_id)
            if status:
                query += " AND t.status=?"
                params.append(status)
            query += " AND t.user_id=?"
            params.append(current_user["id"])
            query += " ORDER BY t.created_at DESC"
            rows = conn.execute(query, params).fetchall()
            return _attach_topup_account_amount([dict(row) for row in rows])
    except Exception as exc:
        logging.exception("Failed to list topups for user_id=%s: %s", current_user.get("id"), exc)
        return []


class TopupCreatePayload(BaseModel):
    account_id: int
    amount_input: float
    fee_percent: float = 0.0
    vat_percent: float = 0.0
    currency: str = "KZT"
    fx_rate: Optional[float] = None


def _topup_gross_amount(amount_input: float, fee_percent: float, vat_percent: float) -> float:
    fee_amount = amount_input * (fee_percent / 100.0)
    vat_amount = amount_input * (vat_percent / 100.0)
    return amount_input + fee_amount + vat_amount


@app.post("/topups")
def create_topup(payload: TopupCreatePayload, current_user=Depends(get_current_user)):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    amount_input = payload.amount_input
    if amount_input <= 0:
        raise HTTPException(status_code=400, detail="amount_input must be positive")
    account_id = payload.account_id
    fee_percent = payload.fee_percent
    vat_percent = payload.vat_percent
    currency = payload.currency
    fx_rate = payload.fx_rate
    with get_conn() as conn:
        acc = conn.execute("SELECT platform, currency, user_id FROM ad_accounts WHERE id=?", (account_id,)).fetchone()
        if not acc:
            raise HTTPException(status_code=404, detail="Account not found")
        # validate platform enum
        if acc["platform"] not in {"meta", "google", "tiktok", "yandex", "telegram", "monochrome"}:
            raise HTTPException(status_code=400, detail="Unsupported platform for topup")
        if acc["user_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Account belongs to a different user")
        resolved_user_id = current_user["id"]
        profile = _get_or_create_profile(conn, resolved_user_id)
        fee_config = _load_fee_config(profile.get("fee_config"))
        platform_fee = fee_config.get(acc["platform"])
        if platform_fee is None:
            raise HTTPException(status_code=400, detail="Commission is not set for this platform")
        fee_percent = float(platform_fee)
        wallet = _get_or_create_wallet(conn, resolved_user_id)
        wallet_balance = float(wallet["balance"] or 0)
        gross_amount = _topup_gross_amount(amount_input, fee_percent, vat_percent)
        if gross_amount > wallet_balance:
            denom = 1.0 + (fee_percent / 100.0) + (vat_percent / 100.0)
            max_input = (wallet_balance / denom) if denom > 0 else 0.0
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Недостаточно средств на кошельке: требуется {gross_amount:.2f} {currency}, "
                    f"доступно {wallet_balance:.2f} {currency}. "
                    f"Максимальная сумма пополнения при текущей комиссии: {max_input:.2f} {currency}."
                ),
            )
        if fx_rate and fx_rate > 0 and str(acc["currency"] or currency).upper() != str(currency).upper():
            amount_net = amount_input / fx_rate
        else:
            amount_net = amount_input
        cur = conn.execute(
            """
            INSERT INTO topups (account_id, user_id, amount_input, fee_percent, vat_percent, amount_net, currency, fx_rate, hold_applied, status, seen_by_admin)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (account_id, resolved_user_id, amount_input, fee_percent, vat_percent, amount_net, currency, fx_rate, 1, "pending", 0),
        )
        topup_id = cur.lastrowid
        new_balance = wallet_balance - gross_amount
        conn.execute(
            "UPDATE wallets SET balance=?, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
            (new_balance, resolved_user_id),
        )
        conn.execute(
            """
            INSERT INTO wallet_transactions (user_id, account_id, amount, currency, type, note)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (resolved_user_id, account_id, -gross_amount, currency, "topup_hold", f"Topup hold #{topup_id}"),
        )
        conn.commit()
        account_name = conn.execute("SELECT name FROM ad_accounts WHERE id=?", (account_id,)).fetchone()
        _send_telegram_alert(
            "\n".join(
                [
                    "💳 <b>Новая заявка на пополнение</b>",
                    f"ID: <code>{topup_id}</code>",
                    f"Пользователь: <code>{current_user['email']}</code> (id={resolved_user_id})",
                    f"Платформа: <b>{acc['platform']}</b>",
                    f"Аккаунт: <b>{account_name['name'] if account_name else account_id}</b> (id={account_id})",
                    f"Сумма: <b>{amount_input:.2f} {currency}</b>",
                    f"Комиссия: <b>{fee_percent:.2f}%</b>",
                    f"Холд в кошельке: <b>{gross_amount:.2f} {currency}</b>",
                ]
            )
        )
        return {
            "id": topup_id,
            "account_id": account_id,
            "user_id": resolved_user_id,
            "amount_input": amount_input,
            "fee_percent": fee_percent,
            "vat_percent": vat_percent,
            "amount_net": amount_net,
            "currency": currency,
            "fx_rate": fx_rate,
            "hold_applied": 1,
            "hold_amount": gross_amount,
            "status": "pending",
        }


@app.get("/invoices/{topup_id}", response_class=HTMLResponse)
def invoice_by_topup(
    topup_id: int,
    token: Optional[str] = None,
    current_user=Depends(get_optional_user),
):
    if not get_conn:
        raise HTTPException(status_code=500, detail="DB not initialized")
    if token:
        current_user = _get_user_by_token(token)
    if not current_user:
        raise HTTPException(status_code=401, detail="Missing token")
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT t.*, a.name as account_name
            FROM topups t
            JOIN ad_accounts a ON a.id=t.account_id
            WHERE t.id=? AND t.user_id=?
            """,
            (topup_id, current_user["id"]),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Invoice not found")
        number = _invoice_number("INV", row["created_at"], row["id"])
        company = _get_company_profile(conn)
        beneficiary_bin = company.get("bin") or company.get("iin") or BENEFICIARY["bin"]
        created_at = row["created_at"]
        date_str = created_at.date().isoformat() if isinstance(created_at, datetime) else str(created_at).split(" ")[0]
        payload = {
            "number": number,
            "date": date_str,
            "beneficiary_name": company.get("name") or BENEFICIARY["name"],
            "beneficiary_bin": beneficiary_bin,
            "beneficiary_bank": company.get("bank") or BENEFICIARY["bank"],
            "beneficiary_iban": company.get("iban") or BENEFICIARY["iban"],
            "beneficiary_bic": company.get("bic") or BENEFICIARY["bic"],
            "beneficiary_kbe": company.get("kbe") or BENEFICIARY["kbe"],
            "payer_name": "Плательщик не указан",
            "payer_bin": "ИИН/БИН не указан",
            "payer_address": "Адрес не указан",
            "description": f"Пополнение аккаунта {row['account_name']}",
            "amount": _format_amount(row["amount_input"]),
            "currency": company.get("currency") or BENEFICIARY["currency"],
            "items": [
                {
                    "description": f"Пополнение аккаунта {row['account_name']}",
                    "qty": "1",
                    "unit": "усл.",
                    "price": _format_amount(row["amount_input"]),
                    "amount": _format_amount(row["amount_input"]),
                }
            ],
            "items_count": 1,
        }
        return HTMLResponse(content=_invoice_html(payload))


# Local run: uvicorn app.main:app --reload
